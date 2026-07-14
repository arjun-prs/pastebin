#!/usr/bin/env python3
"""Generate June validation and Codex/script effort tracker workbook."""

from __future__ import annotations

import argparse
import collections
import csv
import datetime as dt
import getpass
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


QCLIHCDATA_ROOT = Path.home() / "qclihcdata"
ROMA_REPORTS_ROOT = Path.home() / "tools/ROMA/roma-deployment-scripts/reports"
ROMA_REPORTS_BACKUP_ROOT = Path.home() / "tools/roma/roma-deployment-scripts/reports"
TERMINAL_LOGS_ROOT = Path.home() / "OneDrive - Oracle Corporation/OneDrive Documents/Logs"
DEFAULT_OUTPUT_DIR = QCLIHCDATA_ROOT

MINUTES_PER_VALIDATION = 5
LOG_SESSION_GAP_MINUTES = 15
LOG_SESSION_CAP_MINUTES = 90

SKIP_NAME_PARTS = (
    "pp_matrix",
    "pp_info",
    "devices.txt",
    "previous",
    "storekeeper",
    "host-serial",
    "hostname-validation",
    "problematic_output",
)

LOG_LINE_TIMESTAMP_RE = re.compile(
    r"^\[(\d{2})/(\d{2})/(\d{4}), (\d{1,2}):(\d{2}):(\d{2})\.\d+ (AM|PM)\]"
)
TERMINAL_LOG_HEADERS = ["Day", "Date", "Log files", "Timestamped lines", "Estimated effort"]


def format_minutes(minutes: int) -> str:
    hours, mins = divmod(minutes, 60)
    parts = []
    if hours:
        parts.append(f"{hours} hour" + ("" if hours == 1 else "s"))
    if mins:
        parts.append(f"{mins} minute" + ("" if mins == 1 else "s"))
    return " ".join(parts) if parts else "0 minutes"


def format_hours_range(min_hours: float, max_hours: float) -> str:
    def one(value: float) -> str:
        minutes = int(round(value * 60))
        return format_minutes(minutes)

    if min_hours == max_hours:
        return one(min_hours)
    return f"{one(min_hours)} - {one(max_hours)}"


def add_effort_range(validation_minutes: int, min_hours: float, max_hours: float) -> str:
    min_total = validation_minutes + int(round(min_hours * 60))
    max_total = validation_minutes + int(round(max_hours * 60))
    if min_total == max_total:
        return format_minutes(min_total)
    return f"{format_minutes(min_total)} - {format_minutes(max_total)}"


def week_label(date_text: str) -> str:
    day = int(date_text[-2:])
    if day <= 7:
        return "Jun 1-7"
    if day <= 14:
        return "Jun 8-14"
    if day <= 21:
        return "Jun 15-21"
    return "Jun 22+"


def period_label(date_text: str) -> str:
    date = dt.date.fromisoformat(date_text)
    return f"Jun {date.day}"


def display_date(date_text: str) -> str:
    date = dt.date.fromisoformat(date_text)
    return f"Jun {date.day}"


def normalize_build(raw: str) -> str:
    build = raw.upper()
    if build.startswith("AGA5.2"):
        return "AGA5.2"
    if build.startswith("AGA5-Q2-P-T1"):
        return "AGA5-Q2-P-T1-REPORTS"
    for prefix in (
        "AGA5-Q2-P",
        "AGA5-Q3-IP",
        "AGA5-Q3-P",
        "HSG17",
        "JBP15",
        "JBP19",
        "PHX20",
        "PHX23",
        "IAD65",
        "IAD77",
        "SYD20",
    ):
        if build.startswith(prefix):
            return prefix
    return build


def qclihcdata_record(path: Path, root: Path) -> tuple[str, str, str] | None:
    rel = path.relative_to(root)
    name = path.name

    patterns = (
        r"(?P<build>[A-Za-z]{3}\d+(?:\.\d+)?(?:-[A-Za-z0-9]+)*?)-(?P<date>2026-06-\d{2})[_-](?P<time>\d{4,6})",
        r"(?P<build>AGA5\.2)-DG.*?(?P<date>2026-06-\d{2})-(?P<time>\d{4})",
        r"(?P<build>[A-Za-z]{3}\d+)-DG.*?(?P<date>2026-06-\d{2})-(?P<time>\d{4})",
        r"(?P<build>JBP15)-DG.*?(?P<date>2026-06-\d{2})-(?P<time>\d{4})",
        r"(?P<build>JBP19)-DG.*?(?P<date>2026-06-\d{2})-(?P<time>\d{4})",
    )
    for pattern in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            return (
                normalize_build(match.group("build")),
                match.group("date"),
                match.group("time")[:6],
            )

    match = re.search(
        r"data-(?P<date>2026-06-\d{2})[ _-](?P<h>\d{2})_(?P<mi>\d{2})(?:_(?P<s>\d{2}))?",
        name,
        re.IGNORECASE,
    )
    if match:
        site = rel.parts[0].upper()
        build = {
            "AGA": "AGA-CSV-EXPORTS",
            "SYD": "SYD-CSV-EXPORTS",
            "JBP": "JBP-CSV-EXPORTS",
            "IAD": "IAD-CSV-EXPORTS",
        }.get(site, f"{site}-CSV-EXPORTS")
        time = match.group("h") + match.group("mi") + (match.group("s") or "00")
        return build, match.group("date"), time

    return None


def roma_record(path: Path) -> tuple[str, str, str] | None:
    match = re.search(
        r"(?P<build>ABL\d+_DH\d+).*?(?P<day>\d{2})Jun26_?(?P<time>\d{4})?",
        path.name,
        re.IGNORECASE,
    )
    if not match:
        return None
    date = dt.date(2026, 6, int(match.group("day"))).isoformat()
    return match.group("build").upper(), date, match.group("time") or "0000"


def record_datetime(date_text: str, run_time: str) -> dt.datetime:
    padded = (run_time or "0000").ljust(6, "0")[:6]
    return dt.datetime.combine(
        dt.date.fromisoformat(date_text),
        dt.time(int(padded[:2]), int(padded[2:4]), int(padded[4:6])),
    )


def scan_records(
    start: dt.date,
    end: dt.date,
    end_datetime: dt.datetime | None = None,
) -> set[tuple[str, str, str, str]]:
    records: set[tuple[str, str, str, str]] = set()
    scan_roots = (
        (QCLIHCDATA_ROOT, "Commercial Builds"),
        (ROMA_REPORTS_ROOT, "ROMA reports"),
        (ROMA_REPORTS_BACKUP_ROOT, "ROMA reports"),
    )
    for root, source in scan_roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            low_name = path.name.lower()
            if path.suffix.lower() not in {".xlsx", ".csv", ".json"}:
                continue
            if any(part in low_name for part in SKIP_NAME_PARTS):
                continue
            if re.search(r"_full_report_dg\d+\.xlsx$", low_name):
                continue
            parsed = roma_record(path) if source == "ROMA reports" else qclihcdata_record(path, root)
            if not parsed:
                continue
            build, date_text, run_time = parsed
            date = dt.date.fromisoformat(date_text)
            if not (start <= date <= end):
                continue
            if end_datetime and record_datetime(date_text, run_time) > end_datetime:
                continue
            records.add((source, build, date_text, run_time))
    return records


def parse_log_line_timestamp(line: str) -> dt.datetime | None:
    match = LOG_LINE_TIMESTAMP_RE.match(line)
    if not match:
        return None
    day, month, year, hour, minute, second, am_pm = match.groups()
    hour_int = int(hour)
    if am_pm == "PM" and hour_int != 12:
        hour_int += 12
    if am_pm == "AM" and hour_int == 12:
        hour_int = 0
    return dt.datetime(
        int(year),
        int(month),
        int(day),
        hour_int,
        int(minute),
        int(second),
    )


def merge_intervals(
    intervals: list[tuple[dt.datetime, dt.datetime]],
    gap: dt.timedelta,
) -> list[tuple[dt.datetime, dt.datetime]]:
    merged: list[list[dt.datetime]] = []
    for start, end in sorted(intervals):
        if not merged or start > merged[-1][1] + gap:
            merged.append([start, end])
        else:
            merged[-1][1] = max(merged[-1][1], end)
    return [(start, end) for start, end in merged]


def scan_terminal_log_effort(start: dt.date, end: dt.date) -> dict[str, dict[str, int]]:
    """Estimate terminal-active effort from timestamped OneDrive terminal logs.

    This is intentionally conservative: timestamps are grouped by actual line
    date, each log file's per-day contribution is capped, and nearby intervals
    are merged so parallel terminal tabs do not inflate the same active window.
    """
    if not TERMINAL_LOGS_ROOT.exists():
        return {}

    intervals_by_day: dict[str, list[tuple[dt.datetime, dt.datetime]]] = collections.defaultdict(list)
    files_by_day: dict[str, set[str]] = collections.defaultdict(set)
    lines_by_day: collections.Counter[str] = collections.Counter()
    session_cap = dt.timedelta(minutes=LOG_SESSION_CAP_MINUTES)

    for path in sorted(TERMINAL_LOGS_ROOT.glob("*.log")):
        per_day: dict[str, list[dt.datetime]] = collections.defaultdict(list)
        try:
            with path.open(errors="ignore") as handle:
                for line in handle:
                    timestamp = parse_log_line_timestamp(line)
                    if not timestamp or not (start <= timestamp.date() <= end):
                        continue
                    per_day[timestamp.date().isoformat()].append(timestamp)
        except OSError:
            continue

        for date_text, timestamps in per_day.items():
            timestamps.sort()
            interval_start = timestamps[0]
            interval_end = timestamps[-1]
            if interval_end <= interval_start:
                interval_end = interval_start + dt.timedelta(minutes=5)
            if interval_end - interval_start > session_cap:
                interval_end = interval_start + session_cap
            intervals_by_day[date_text].append((interval_start, interval_end))
            files_by_day[date_text].add(path.name)
            lines_by_day[date_text] += len(timestamps)

    effort: dict[str, dict[str, int]] = {}
    gap = dt.timedelta(minutes=LOG_SESSION_GAP_MINUTES)
    for date_text, intervals in intervals_by_day.items():
        merged = merge_intervals(intervals, gap)
        active_minutes = round(sum((end - start).total_seconds() / 60 for start, end in merged))
        effort[date_text] = {
            "log_files": len(files_by_day[date_text]),
            "timestamped_lines": lines_by_day[date_text],
            "active_minutes": int(active_minutes),
        }
    return effort


def parse_meeting_minutes(row: dict[str, str]) -> int:
    if row.get("minutes"):
        return int(float(row["minutes"]))
    if row.get("meeting_minutes"):
        return int(float(row["meeting_minutes"]))
    if row.get("hours"):
        return int(round(float(row["hours"]) * 60))
    if row.get("meeting_hours"):
        return int(round(float(row["meeting_hours"]) * 60))
    raise ValueError("meeting CSV row must contain minutes, meeting_minutes, hours, or meeting_hours")


def load_meeting_hours(
    path: Path,
    start: dt.date,
    end: dt.date,
    min_attendees: int = 0,
) -> dict[str, dict[str, int]]:
    meeting_hours: dict[str, dict[str, int]] = {}
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "date" not in reader.fieldnames:
            raise ValueError("meeting CSV must contain a date column")
        for row in reader:
            if not row.get("date"):
                continue
            date = dt.date.fromisoformat(row["date"])
            if not (start <= date <= end):
                continue
            attendee_count = int(row.get("attendee_count") or row.get("attendees") or 0)
            if min_attendees and attendee_count and attendee_count < min_attendees:
                continue
            date_text = date.isoformat()
            minutes = parse_meeting_minutes(row)
            if date_text not in meeting_hours:
                meeting_hours[date_text] = {"meeting_count": 0, "meeting_minutes": 0}
            meeting_hours[date_text]["meeting_count"] += int(row.get("meeting_count") or row.get("meetings") or 1)
            meeting_hours[date_text]["meeting_minutes"] += minutes
    return meeting_hours


def parse_codex_effort_minutes(row: dict[str, str]) -> tuple[int, int]:
    if row.get("minutes"):
        minutes = int(float(row["minutes"]))
        return minutes, minutes
    if row.get("max_minutes"):
        min_minutes = int(float(row.get("min_minutes") or row["max_minutes"]))
        max_minutes = int(float(row["max_minutes"]))
        return min_minutes, max_minutes
    if row.get("hours"):
        minutes = int(round(float(row["hours"]) * 60))
        return minutes, minutes
    if row.get("max_hours"):
        min_minutes = int(round(float(row.get("min_hours") or row["max_hours"]) * 60))
        max_minutes = int(round(float(row["max_hours"]) * 60))
        return min_minutes, max_minutes
    raise ValueError("Codex CSV row must contain minutes, min_minutes/max_minutes, hours, or min_hours/max_hours")


def load_codex_effort(path: Path, start: dt.date, end: dt.date) -> dict[str, object]:
    min_total = 0
    max_total = 0
    count = 0
    by_period: dict[str, dict[str, int]] = collections.defaultdict(lambda: {"count": 0, "min_minutes": 0, "max_minutes": 0})
    by_date: dict[str, dict[str, int]] = collections.defaultdict(lambda: {"count": 0, "min_minutes": 0, "max_minutes": 0})
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "date" not in reader.fieldnames:
            raise ValueError("Codex CSV must contain a date column")
        for row in reader:
            if not row.get("date"):
                continue
            date = dt.date.fromisoformat(row["date"])
            if not (start <= date <= end):
                continue
            min_minutes, max_minutes = parse_codex_effort_minutes(row)
            min_total += min_minutes
            max_total += max_minutes
            count += 1
            period = period_label(date.isoformat())
            by_period[period]["count"] += 1
            by_period[period]["min_minutes"] += min_minutes
            by_period[period]["max_minutes"] += max_minutes
            date_text = date.isoformat()
            by_date[date_text]["count"] += 1
            by_date[date_text]["min_minutes"] += min_minutes
            by_date[date_text]["max_minutes"] += max_minutes
    return {
        "count": count,
        "min_minutes": min_total,
        "max_minutes": max_total,
        "by_period": dict(by_period),
        "by_date": dict(by_date),
    }


def rows_from_counts(counter: collections.Counter, key_names: tuple[str, ...]) -> list[list[object]]:
    rows = []
    for key, count in sorted(counter.items()):
        keys = key if isinstance(key, tuple) else (key,)
        minutes = count * MINUTES_PER_VALIDATION
        rows.append([*keys, count, format_minutes(minutes)])
    return rows


def effort_range_text(min_minutes: int, max_minutes: int | None = None) -> str:
    max_minutes = min_minutes if max_minutes is None else max_minutes
    return format_hours_range(min_minutes / 60, max_minutes / 60)


def consolidated_period_sort_key(label: str) -> int:
    match = re.search(r"Jun\s+(\d+)", label)
    return int(match.group(1)) if match else 99


def write_sheet(
    ws,
    headers: list[str],
    rows: list[list[object]],
    metadata: list[tuple[str, str]] | None = None,
) -> None:
    header_row = 1
    if metadata:
        for label, value in metadata:
            ws.append([label, value])
        ws.append([])
        header_row = len(metadata) + 2
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for row in ws.iter_rows(min_row=1, max_row=header_row - 2):
        for cell in row:
            cell.font = Font(bold=True)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 70)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def build_workbook(
    records: set[tuple[str, str, str, str]],
    output: Path,
    log_effort: dict[str, dict[str, int]] | None = None,
    meeting_hours: dict[str, dict[str, int]] | None = None,
    codex_effort: dict[str, object] | None = None,
) -> None:
    by_source = collections.Counter()
    by_roma_build = collections.Counter()
    by_week = collections.Counter()
    by_date = collections.Counter()
    log_effort = log_effort or {}
    meeting_hours = meeting_hours or {}
    codex_effort = codex_effort or {}

    for source, build, date_text, _run_time in records:
        by_source[source] += 1
        by_date[date_text] += 1
        by_week[week_label(date_text)] += 1
        if source == "ROMA reports":
            by_roma_build[build] += 1

    generated_by = getpass.getuser()
    metadata = [("Generated by", generated_by)]

    wb = Workbook()
    wb.properties.creator = generated_by
    ws = wb.active
    ws.title = "Updated Total"
    total_rows = rows_from_counts(by_source, ("Source",))
    total_count = sum(by_source.values())
    validation_minutes = total_count * MINUTES_PER_VALIDATION
    log_minutes = sum(item["active_minutes"] for item in log_effort.values())
    meeting_minutes = sum(item["meeting_minutes"] for item in meeting_hours.values())
    codex_min_minutes = int(codex_effort.get("min_minutes", 0))
    codex_max_minutes = int(codex_effort.get("max_minutes", codex_min_minutes))
    total_min_minutes = validation_minutes
    total_max_minutes = validation_minutes
    if log_effort:
        total_rows.append(["Terminal Log Effort", "", format_minutes(log_minutes)])
        total_min_minutes += log_minutes
        total_max_minutes += log_minutes
    if codex_effort:
        total_rows.append(
            [
                "Codex Script Work Done",
                codex_effort.get("count", ""),
                effort_range_text(codex_min_minutes, codex_max_minutes),
            ]
        )
        total_min_minutes += codex_min_minutes
        total_max_minutes += codex_max_minutes
    if meeting_hours:
        total_rows.append(["Meeting Hours", "", format_minutes(meeting_minutes)])
        total_min_minutes += meeting_minutes
        total_max_minutes += meeting_minutes
    total_rows.append(["Total", total_count, effort_range_text(total_min_minutes, total_max_minutes)])
    write_sheet(ws, ["Source", "Runs", "Effort"], total_rows, metadata)

    ws = wb.create_sheet("ROMA Buildwise")
    roma_rows = rows_from_counts(by_roma_build, ("Build",))
    roma_count = sum(by_roma_build.values())
    roma_rows.append(["ROMA total", roma_count, format_minutes(roma_count * MINUTES_PER_VALIDATION)])
    write_sheet(ws, ["Build", "Runs", "Effort"], roma_rows, metadata)

    ws = wb.create_sheet("Updated Weekwise Total")
    week_order = {"Jun 1-7": 1, "Jun 8-14": 2, "Jun 15-21": 3, "Jun 22+": 4}
    log_by_week: collections.Counter[str] = collections.Counter()
    for date_text, info in log_effort.items():
        log_by_week[week_label(date_text)] += info["active_minutes"]
    meeting_by_week: collections.Counter[str] = collections.Counter()
    for date_text, info in meeting_hours.items():
        meeting_by_week[week_label(date_text)] += info["meeting_minutes"]
    week_rows = []
    all_weeks = set(by_week) | set(log_by_week) | set(meeting_by_week)
    for week in sorted(all_weeks, key=lambda item: week_order[item]):
        count = by_week[week]
        week_validation_minutes = count * MINUTES_PER_VALIDATION
        week_log_minutes = log_by_week[week]
        week_meeting_minutes = meeting_by_week[week]
        if log_effort or meeting_hours:
            week_rows.append(
                [
                    week,
                    count,
                    format_minutes(week_validation_minutes),
                    format_minutes(week_log_minutes),
                    format_minutes(week_meeting_minutes),
                    format_minutes(week_validation_minutes + week_log_minutes + week_meeting_minutes),
                ]
            )
        else:
            week_rows.append([week, count, format_minutes(week_validation_minutes)])
    if log_effort or meeting_hours:
        week_rows.append(
            [
                "Total",
                total_count,
                format_minutes(validation_minutes),
                format_minutes(log_minutes),
                format_minutes(meeting_minutes),
                format_minutes(validation_minutes + log_minutes + meeting_minutes),
            ]
        )
        write_sheet(
            ws,
            ["Week", "Runs", "Validation effort", "Terminal log effort", "Meeting hours", "Total effort"],
            week_rows,
            metadata,
        )
    else:
        week_rows.append(["Total", total_count, format_minutes(validation_minutes)])
        write_sheet(ws, ["Week", "Runs", "Effort"], week_rows, metadata)

    ws = wb.create_sheet("Daywise Efforts")
    day_rows = []
    codex_by_date = codex_effort.get("by_date", {})
    all_dates = set(by_date) | set(log_effort) | set(meeting_hours) | set(codex_by_date)
    for date_text in sorted(all_dates):
        count = by_date[date_text]
        day_validation_minutes = count * MINUTES_PER_VALIDATION
        day_log_minutes = log_effort.get(date_text, {}).get("active_minutes", 0)
        day_meeting_minutes = meeting_hours.get(date_text, {}).get("meeting_minutes", 0)
        day_codex_info = codex_by_date.get(date_text, {})
        day_codex_min = int(day_codex_info.get("min_minutes", 0))
        day_codex_max = int(day_codex_info.get("max_minutes", day_codex_min))
        if log_effort or meeting_hours or codex_effort:
            day_min_total = day_validation_minutes + day_log_minutes + day_codex_min + day_meeting_minutes
            day_max_total = day_validation_minutes + day_log_minutes + day_codex_max + day_meeting_minutes
            day_rows.append(
                [
                    display_date(date_text),
                    date_text,
                    count,
                    format_minutes(day_validation_minutes),
                    format_minutes(day_log_minutes),
                    format_hours_range(day_codex_min / 60, day_codex_max / 60) if day_codex_info else "-",
                    format_minutes(day_meeting_minutes),
                    format_hours_range(day_min_total / 60, day_max_total / 60),
                ]
            )
        else:
            day_rows.append(
                [
                    display_date(date_text),
                    date_text,
                    count,
                    format_minutes(day_validation_minutes),
                ]
            )
    if log_effort or meeting_hours or codex_effort:
        day_rows.append(
            [
                "Total",
                "",
                total_count,
                format_minutes(validation_minutes),
                format_minutes(log_minutes),
                format_hours_range(codex_min_minutes / 60, codex_max_minutes / 60) if codex_effort else "-",
                format_minutes(meeting_minutes),
                format_hours_range(total_min_minutes / 60, total_max_minutes / 60),
            ]
        )
        write_sheet(
            ws,
            [
                "Day",
                "Date",
                "Runs",
                "Validation effort",
                "Terminal Log Effort",
                "Codex/script effort",
                "Meeting hours",
                "Total effort",
            ],
            day_rows,
            metadata,
        )
    else:
        day_rows.append(["Total", "", total_count, format_minutes(validation_minutes)])
        write_sheet(ws, ["Day", "Date", "Runs", "Effort"], day_rows, metadata)

    if log_effort:
        ws = wb.create_sheet("Terminal Log Effort")
        log_rows = []
        for date_text, info in sorted(log_effort.items()):
            log_rows.append(
                [
                    display_date(date_text),
                    date_text,
                    info["log_files"],
                    info["timestamped_lines"],
                    format_minutes(info["active_minutes"]),
                ]
            )
        log_rows.append(["Total", "", "", "", format_minutes(log_minutes)])
        write_sheet(
            ws,
            TERMINAL_LOG_HEADERS,
            log_rows,
            metadata,
        )

    if meeting_hours:
        ws = wb.create_sheet("Meeting Hours")
        meeting_rows = []
        total_meetings = 0
        for date_text, info in sorted(meeting_hours.items()):
            total_meetings += info["meeting_count"]
            meeting_rows.append(
                [
                    display_date(date_text),
                    date_text,
                    info["meeting_count"],
                    format_minutes(info["meeting_minutes"]),
                ]
            )
        meeting_rows.append(["Total", "", total_meetings, format_minutes(meeting_minutes)])
        write_sheet(
            ws,
            ["Day", "Date", "Meeting count", "Meeting hours"],
            meeting_rows,
            metadata,
        )

    if log_effort or meeting_hours or codex_effort:
        ws = wb.create_sheet("Consolidated table")
        validation_by_period: collections.Counter[str] = collections.Counter()
        runs_by_period: collections.Counter[str] = collections.Counter()
        for _source, _build, date_text, _run_time in records:
            period = period_label(date_text)
            runs_by_period[period] += 1
            validation_by_period[period] += MINUTES_PER_VALIDATION

        log_by_period: collections.Counter[str] = collections.Counter()
        for date_text, info in log_effort.items():
            log_by_period[period_label(date_text)] += info["active_minutes"]

        meeting_by_period: collections.Counter[str] = collections.Counter()
        for date_text, info in meeting_hours.items():
            meeting_by_period[period_label(date_text)] += info["meeting_minutes"]

        codex_by_period = codex_effort.get("by_period", {})
        all_periods = set(runs_by_period) | set(log_by_period) | set(meeting_by_period) | set(codex_by_period)
        consolidated_rows = []
        for period in sorted(all_periods, key=consolidated_period_sort_key):
            codex_info = codex_by_period.get(period, {})
            codex_min = int(codex_info.get("min_minutes", 0))
            codex_max = int(codex_info.get("max_minutes", codex_min))
            row_min = validation_by_period[period] + log_by_period[period] + meeting_by_period[period] + codex_min
            row_max = validation_by_period[period] + log_by_period[period] + meeting_by_period[period] + codex_max
            consolidated_rows.append(
                [
                    period,
                    runs_by_period[period],
                    format_minutes(validation_by_period[period]),
                    format_minutes(log_by_period[period]),
                    format_hours_range(codex_min / 60, codex_max / 60) if codex_info else "-",
                    format_minutes(meeting_by_period[period]),
                    format_hours_range(row_min / 60, row_max / 60),
                ]
            )
        total_min = validation_minutes + log_minutes + meeting_minutes + codex_min_minutes
        total_max = validation_minutes + log_minutes + meeting_minutes + codex_max_minutes
        consolidated_rows.append(
            [
                "Total",
                total_count,
                format_minutes(validation_minutes),
                format_minutes(log_minutes),
                format_hours_range(codex_min_minutes / 60, codex_max_minutes / 60) if codex_effort else "-",
                format_minutes(meeting_minutes),
                format_hours_range(total_min / 60, total_max / 60),
            ]
        )
        write_sheet(
            ws,
            [
                "Date / period",
                "Validation runs",
                "Validation effort",
                "Terminal Log Effort",
                "Codex/script effort",
                "Meeting hours",
                "Total effort",
            ],
            consolidated_rows,
            metadata,
        )

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--start", default="2026-06-01", help="Start date, YYYY-MM-DD")
    parser.add_argument(
        "--end",
        help="End date, YYYY-MM-DD. Defaults to today's date.",
    )
    parser.add_argument(
        "--output",
        help=(
            "Output .xlsx path. Defaults to "
            "~/qclihcdata/work_tracker_<username>_<generated_date>_<generated_time>.xlsx"
        ),
    )
    parser.add_argument(
        "--include-log-effort",
        action="store_true",
        help=(
            "Add terminal-active effort estimated from OneDrive terminal logs as a "
            "separate sheet and include it in total effort columns."
        ),
    )
    parser.add_argument(
        "--meeting-hours-csv",
        help=(
            "Optional CSV with date plus minutes/meeting_minutes or hours/meeting_hours. "
            "Adds a Meeting Hours sheet and includes those hours in totals."
        ),
    )
    parser.add_argument(
        "--min-meeting-attendees",
        type=int,
        default=0,
        help=(
            "When --meeting-hours-csv includes attendee_count/attendees, only include "
            "meeting rows with at least this many attendees."
        ),
    )
    parser.add_argument(
        "--codex-effort-csv",
        help=(
            "Optional CSV with date plus minutes/hours or min_minutes/max_minutes. "
            "Adds Codex/script effort into the Updated Total sheet only."
        ),
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    start = dt.date.fromisoformat(args.start)
    generated_at = dt.datetime.now()
    end = dt.date.fromisoformat(args.end) if args.end else generated_at.date()
    end_datetime = None if args.end else generated_at
    if args.output:
        output = Path(args.output).expanduser()
    else:
        username = getpass.getuser()
        generated_at_text = generated_at.strftime("%Y-%m-%d_%H-%M-%S")
        output = DEFAULT_OUTPUT_DIR / f"work_tracker_{username}_{generated_at_text}.xlsx"
    records = scan_records(start, end, end_datetime)
    log_effort = scan_terminal_log_effort(start, end) if args.include_log_effort else None
    meeting_hours = (
        load_meeting_hours(
            Path(args.meeting_hours_csv).expanduser(),
            start,
            end,
            args.min_meeting_attendees,
        )
        if args.meeting_hours_csv
        else None
    )
    codex_effort = load_codex_effort(Path(args.codex_effort_csv).expanduser(), start, end) if args.codex_effort_csv else None
    build_workbook(records, output, log_effort, meeting_hours, codex_effort)
    print(f"Wrote {output}")
    print(f"Distinct validation/report runs: {len(records)}")
    print(f"Estimated validation effort: {format_minutes(len(records) * MINUTES_PER_VALIDATION)}")
    if log_effort:
        log_minutes = sum(item["active_minutes"] for item in log_effort.values())
        print(f"Estimated terminal log effort: {format_minutes(log_minutes)}")
        print(
            "Combined estimated effort: "
            f"{format_minutes((len(records) * MINUTES_PER_VALIDATION) + log_minutes)}"
        )
    if meeting_hours:
        meeting_minutes = sum(item["meeting_minutes"] for item in meeting_hours.values())
        print(f"Estimated meeting hours: {format_minutes(meeting_minutes)}")
    if codex_effort:
        print(
            "Estimated Codex/script effort: "
            f"{format_hours_range(codex_effort['min_minutes'] / 60, codex_effort['max_minutes'] / 60)}"
        )


if __name__ == "__main__":
    main()
