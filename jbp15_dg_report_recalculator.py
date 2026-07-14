#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
import json
import re
import shlex
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook


class workbook_patch:
    PATCH_COLUMNS = ["Source_port", "DMARC1", "DMARC2", "Destination_port"]
    A_END_DISPLAY_COLUMNS = ["Hostname", "Interface", "Rack", "Elevation"]
    Z_COLUMNS = ["Z Hostname", "Z Interface", "Z Rack", "Z Elevation"]
    PATCH_NOT_FOUND = "PP_info_not_found"
    TRANSCEIVER_INTERFACE_SUFFIXES = ("s0", "s1", "s2", "s3")
    RACK_ELEV_RE = re.compile(r"\(Rack-(\d+):(\d+)-elevation\)", re.IGNORECASE)
    PATCH_RACK_RE = re.compile(r"rack\s+(\d+)(?:\s+u(\d+))?", re.IGNORECASE)

    @staticmethod
    def norm(value) -> str:
        if value is None:
            return ""
        if pd.isna(value):
            return ""
        text = str(value).replace("\xa0", " ").strip()
        return "" if text.lower() == "nan" else text

    @staticmethod
    def normalize_token(value: str) -> str:
        return " ".join(workbook_patch.norm(value).lower().replace(":", " ").split())

    @staticmethod
    def endpoint_key(hostname: str, interface: str) -> str:
        return f"{workbook_patch.normalize_token(hostname)} {workbook_patch.normalize_token(interface)}".strip()

    @staticmethod
    def split_device_endpoint(value: str):
        value = workbook_patch.norm(value)
        if not value:
            return "", ""
        parts = value.split()
        if len(parts) >= 2:
            return parts[0], parts[1]
        return value, ""

    @staticmethod
    def parse_hostname_rack_elevation(hostname: str):
        hostname = workbook_patch.norm(hostname)
        if not hostname:
            return "", "", ""

        match = workbook_patch.RACK_ELEV_RE.search(hostname)
        if not match:
            return hostname, "", ""

        clean_hostname = hostname.split(" (", 1)[0].strip()
        rack = match.group(1).zfill(4)
        elevation = match.group(2)
        return clean_hostname, rack, elevation

    @staticmethod
    def parse_patch_rack_elevation(value: str):
        value = workbook_patch.norm(value)
        if not value:
            return "", ""

        match = workbook_patch.PATCH_RACK_RE.search(value)
        if not match:
            return "", ""

        rack = match.group(1)
        elevation = workbook_patch.norm(match.group(2))
        return rack, elevation

    @staticmethod
    def build_interface_from_transceiver(transceiver: str, channel: str) -> str:
        transceiver = workbook_patch.norm(transceiver)
        channel = workbook_patch.norm(channel)
        if not transceiver or not channel:
            return ""

        transceiver_match = re.search(r"(\d+)", transceiver)
        if not transceiver_match:
            return ""

        try:
            channel_num = int(float(channel))
        except Exception:
            return ""

        if 1 <= channel_num <= 2:
            suffix = "s0"
        elif 3 <= channel_num <= 4:
            suffix = "s1"
        elif 5 <= channel_num <= 6:
            suffix = "s2"
        elif 7 <= channel_num <= 8:
            suffix = "s3"
        else:
            return ""

        return f"swp{transceiver_match.group(1)}{suffix}"

    @staticmethod
    def build_interfaces_from_transceiver(transceiver: str):
        transceiver = workbook_patch.norm(transceiver)
        if not transceiver:
            return []

        transceiver_match = re.search(r"(\d+)", transceiver)
        if not transceiver_match:
            return []

        interface_base = f"swp{transceiver_match.group(1)}"
        return [f"{interface_base}{suffix}" for suffix in workbook_patch.TRANSCEIVER_INTERFACE_SUFFIXES]

    @staticmethod
    def find_column(columns, wanted: str):
        wanted_normalized = wanted.lower().replace("_", "").replace(" ", "")
        for column in columns:
            candidate = str(column).strip().lower().replace("_", "").replace(" ", "")
            if candidate == wanted_normalized:
                return column
        return None

    @staticmethod
    def load_panels_df(panels_path: Path) -> pd.DataFrame:
        workbook = pd.ExcelFile(panels_path)
        first_sheet = workbook.sheet_names[0]
        raw = pd.read_excel(panels_path, sheet_name=first_sheet)

        required_map = {
            "DeviceA": "DeviceA",
            "RackA": "RackA",
            "Source_port": "Source_port",
            "DMARC1": "DMARC1",
            "DMARC2": "DMARC2",
            "Destination_port": "Destination_port",
            "DeviceB": "DeviceB",
            "RackB": "RackB",
        }

        selected = {}
        for output_name, wanted in required_map.items():
            actual = workbook_patch.find_column(raw.columns, wanted)
            if actual is None:
                raise ValueError(
                    f"Patch panel file is missing required column '{wanted}'. "
                    f"Columns found: {list(raw.columns)}"
                )
            selected[output_name] = actual

        panels_df = raw[[selected[name] for name in required_map]].copy()
        panels_df.columns = list(required_map.keys())

        for column in panels_df.columns:
            panels_df[column] = panels_df[column].map(workbook_patch.norm)

        split_a = panels_df["DeviceA"].map(workbook_patch.split_device_endpoint)
        split_b = panels_df["DeviceB"].map(workbook_patch.split_device_endpoint)
        panels_df[["DeviceA_Host", "DeviceA_Interface"]] = pd.DataFrame(split_a.tolist(), index=panels_df.index)
        panels_df[["DeviceB_Host", "DeviceB_Interface"]] = pd.DataFrame(split_b.tolist(), index=panels_df.index)

        panels_df["A_KEY"] = panels_df.apply(
            lambda row: workbook_patch.endpoint_key(row["DeviceA_Host"], row["DeviceA_Interface"]),
            axis=1,
        )
        panels_df["B_KEY"] = panels_df.apply(
            lambda row: workbook_patch.endpoint_key(row["DeviceB_Host"], row["DeviceB_Interface"]),
            axis=1,
        )
        panels_df[["DeviceA_Rack", "DeviceA_Elevation"]] = pd.DataFrame(
            panels_df["RackA"].map(workbook_patch.parse_patch_rack_elevation).tolist(),
            index=panels_df.index,
        )
        panels_df[["DeviceB_Rack", "DeviceB_Elevation"]] = pd.DataFrame(
            panels_df["RackB"].map(workbook_patch.parse_patch_rack_elevation).tolist(),
            index=panels_df.index,
        )

        return panels_df

    @staticmethod
    def panel_lookup_views(panels_df: pd.DataFrame):
        a_view = panels_df.drop_duplicates(subset=["A_KEY"], keep="first").copy()
        b_view = panels_df.drop_duplicates(subset=["B_KEY"], keep="first").copy()
        return a_view, b_view

    @staticmethod
    def assign_panel_columns(target_df: pd.DataFrame, panel_match_df: pd.DataFrame, reverse: bool):
        ordered_columns = workbook_patch.PATCH_COLUMNS if not reverse else list(reversed(workbook_patch.PATCH_COLUMNS))
        for target_column, source_column in zip(workbook_patch.PATCH_COLUMNS, ordered_columns):
            target_df[target_column] = panel_match_df[source_column].values

        z_map = {
            "Z Hostname": "DeviceB_Host",
            "Z Interface": "DeviceB_Interface",
            "Z Rack": "DeviceB_Rack",
            "Z Elevation": "DeviceB_Elevation",
        }

        for target_column, source_column in z_map.items():
            target_df[target_column] = panel_match_df[source_column].values

    @staticmethod
    def fill_missing_columns(df: pd.DataFrame, key_column: str, columns):
        for column in columns:
            df[column] = df[column].fillna(workbook_patch.PATCH_NOT_FOUND)
        no_local_endpoint = df[key_column].eq("")
        for column in columns:
            df.loc[no_local_endpoint, column] = ""

    @staticmethod
    def prepare_local_endpoint(df: pd.DataFrame, hostname_col: str, interface_col: str) -> pd.DataFrame:
        out = df.copy()
        for column in out.columns:
            out[column] = out[column].map(workbook_patch.norm)

        parsed = out[hostname_col].map(workbook_patch.parse_hostname_rack_elevation)
        parsed_df = pd.DataFrame(parsed.tolist(), columns=["_clean_host", "_rack_parsed", "_elev_parsed"])
        out[hostname_col] = parsed_df["_clean_host"]

        if "Rack" not in out.columns:
            out["Rack"] = ""
        if "Elevation" not in out.columns:
            out["Elevation"] = ""

        out.loc[out["Rack"].map(workbook_patch.norm).eq(""), "Rack"] = parsed_df["_rack_parsed"]
        out.loc[out["Elevation"].map(workbook_patch.norm).eq(""), "Elevation"] = parsed_df["_elev_parsed"]
        out["_LOCAL_KEY"] = out.apply(
            lambda row: workbook_patch.endpoint_key(row.get(hostname_col, ""), row.get(interface_col, "")),
            axis=1,
        )
        return out

    @staticmethod
    def enrich_with_patch_path(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        endpoint_columns = [
            "DeviceA_Host",
            "DeviceA_Interface",
            "DeviceA_Rack",
            "DeviceA_Elevation",
            "DeviceB_Host",
            "DeviceB_Interface",
            "DeviceB_Rack",
            "DeviceB_Elevation",
        ]

        a_match = pd.merge(
            out[["_LOCAL_KEY"]],
            a_view[["A_KEY"] + workbook_patch.PATCH_COLUMNS + endpoint_columns],
            how="left",
            left_on="_LOCAL_KEY",
            right_on="A_KEY",
        )

        workbook_patch.assign_panel_columns(out, a_match, reverse=False)

        unmatched_mask = out["Source_port"].isna()
        if unmatched_mask.any():
            b_match = pd.merge(
                out.loc[unmatched_mask, ["_LOCAL_KEY"]],
                b_view[["B_KEY"] + workbook_patch.PATCH_COLUMNS + endpoint_columns],
                how="left",
                left_on="_LOCAL_KEY",
                right_on="B_KEY",
            )
            unmatched_out = out.loc[unmatched_mask].copy()
            workbook_patch.assign_panel_columns(unmatched_out, b_match, reverse=True)
            reversed_columns = dict(zip(workbook_patch.PATCH_COLUMNS, reversed(workbook_patch.PATCH_COLUMNS)))
            for target_column, source_column in reversed_columns.items():
                out.loc[unmatched_mask, target_column] = b_match[source_column].values
            for column in workbook_patch.Z_COLUMNS:
                out.loc[unmatched_mask, column] = unmatched_out[column].values

        workbook_patch.fill_missing_columns(out, "_LOCAL_KEY", workbook_patch.PATCH_COLUMNS + workbook_patch.Z_COLUMNS)
        return out

    @staticmethod
    def enrich_lldp_sheet(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        required = ["Hostname", "Interface"]
        for column in required:
            if column not in df.columns:
                raise ValueError(f"Sheet 'lldp_sp' is missing required column '{column}'")

        out = workbook_patch.prepare_local_endpoint(df, "Hostname", "Interface")
        out = workbook_patch.enrich_with_patch_path(out, a_view, b_view)

        desired_columns = [
            "Hostname",
            "Interface",
            "Rack",
            "Elevation",
            "Source_port",
            "DMARC1",
            "DMARC2",
            "Destination_port",
            "Z Hostname",
            "Z Interface",
            "Z Rack",
            "Z Elevation",
            "Active Host",
            "Act. Interface",
            "Act. Rack",
            "Act. Elevation",
            "Expected Hostname",
            "Exp. Interface",
            "Exp. Rack",
            "Exp. Elevation",
        ]
        return out[[column for column in desired_columns if column in out.columns]]

    @staticmethod
    def enrich_interfaces_sheet(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        if "Hostname" not in df.columns:
            raise ValueError("Sheet 'interfaces_sp' is missing required column 'Hostname'")

        interface_col = None
        for candidate in ["Interface", "Interface (not up / not enabled)"]:
            if candidate in df.columns:
                interface_col = candidate
                break
        if interface_col is None:
            raise ValueError(
                "Sheet 'interfaces_sp' is missing an interface column. "
                "Expected 'Interface' or 'Interface (not up / not enabled)'."
            )

        working = df.copy()
        if interface_col != "Interface":
            working["Interface"] = working[interface_col]

        out = workbook_patch.prepare_local_endpoint(working, "Hostname", "Interface")
        out = workbook_patch.enrich_with_patch_path(out, a_view, b_view)

        desired_columns = [
            "Hostname",
            "Interface",
            "Rack",
            "Elevation",
            "Source_port",
            "DMARC1",
            "DMARC2",
            "Destination_port",
            "Z Hostname",
            "Z Interface",
            "Z Rack",
            "Z Elevation",
        ]
        return out[[column for column in desired_columns if column in out.columns]]

    @staticmethod
    def enrich_optics_sheet(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        if "Hostname" not in df.columns:
            raise ValueError("Sheet 'optics_rx_tx_threshold' is missing required column 'Hostname'")
        if "Transceiver" not in df.columns or "Channel" not in df.columns:
            raise ValueError(
                "Sheet 'optics_rx_tx_threshold' must contain 'Transceiver' and 'Channel' columns"
            )

        working = df.copy()
        for column in working.columns:
            working[column] = working[column].map(workbook_patch.norm)
        working["Interface"] = working.apply(
            lambda row: workbook_patch.build_interface_from_transceiver(row.get("Transceiver", ""), row.get("Channel", "")),
            axis=1,
        )

        out = workbook_patch.prepare_local_endpoint(working, "Hostname", "Interface")
        out = workbook_patch.enrich_with_patch_path(out, a_view, b_view)

        desired_columns = [
            "Hostname",
            "Interface",
            "Rack",
            "Elevation",
            "Transceiver",
            "Channel",
            "Measured (dBm)",
            "Min Threshold (dBm)",
            "Max Threshold (dBm)",
            "Source_port",
            "DMARC1",
            "DMARC2",
            "Destination_port",
            "Z Hostname",
            "Z Interface",
            "Z Rack",
            "Z Elevation",
        ]
        return out[[column for column in desired_columns if column in out.columns]]

    @staticmethod
    def enrich_optics_temp_sheet(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        if "Hostname" not in df.columns:
            raise ValueError("Sheet 'optics_temp' is missing required column 'Hostname'")
        if "Transceiver" not in df.columns:
            raise ValueError("Sheet 'optics_temp' is missing required column 'Transceiver'")

        working = df.copy()
        for column in working.columns:
            working[column] = working[column].map(workbook_patch.norm)

        if "Interface" in working.columns:
            working = working.drop(columns=["Interface"])

        working["_Interface_List"] = working["Transceiver"].map(
            lambda transceiver: workbook_patch.build_interfaces_from_transceiver(transceiver) or [""]
        )
        working = working.explode("_Interface_List").rename(columns={"_Interface_List": "Interface"})

        out = workbook_patch.prepare_local_endpoint(working, "Hostname", "Interface")
        out = workbook_patch.enrich_with_patch_path(out, a_view, b_view)

        desired_columns = [
            "Hostname",
            "Interface",
            "Rack",
            "Elevation",
            "Transceiver",
            "Measured (°C)",
            "Measured (C)",
            "Threshold (°C)",
            "Threshold (C)",
            "Status",
            "Source_port",
            "DMARC1",
            "DMARC2",
            "Destination_port",
            "Z Hostname",
            "Z Interface",
            "Z Rack",
            "Z Elevation",
            "Placement Group",
        ]
        return out[[column for column in desired_columns if column in out.columns]]

    @staticmethod
    def enrich_combined_fec_sheet(df: pd.DataFrame, a_view: pd.DataFrame, b_view: pd.DataFrame) -> pd.DataFrame:
        required = ["Hostname", "Interface"]
        for column in required:
            if column not in df.columns:
                raise ValueError(f"Sheet 'combined_fec' is missing required column '{column}'")

        out = workbook_patch.prepare_local_endpoint(df, "Hostname", "Interface")
        out = workbook_patch.enrich_with_patch_path(out, a_view, b_view)

        desired_columns = [
            "Hostname",
            "Interface",
            "Rack",
            "Elevation",
            "Lock Status",
            "Pre-FEC BER",
            "Pre‑FEC BER",
            "Source_port",
            "DMARC1",
            "DMARC2",
            "Destination_port",
            "Z Hostname",
            "Z Interface",
            "Z Rack",
            "Z Elevation",
            "Remote Host",
            "Remote Interface",
            "Remote Rack",
            "Remote Elevation",
        ]
        return out[[column for column in desired_columns if column in out.columns]]

    @staticmethod
    def style_workbook(output_path: Path):
        from openpyxl.styles import Font, PatternFill

        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
        a_end_fill = PatternFill(start_color="EADCF8", end_color="EADCF8", fill_type="solid")
        z_end_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        patch_fills = {
            "Source_port": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "DMARC1": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "DMARC2": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
            "Destination_port": PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid"),
        }
        lldp_group_styles = {
            "active": {
                "columns": ["Active Host", "Act. Interface", "Act. Rack", "Act. Elevation"],
                "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                "font": Font(color="9C0006"),
            },
            "expected": {
                "columns": ["Expected Hostname", "Exp. Interface", "Exp. Rack", "Exp. Elevation"],
                "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                "font": Font(color="006100"),
            },
        }

        workbook = load_workbook(output_path)
        for worksheet in workbook.worksheets:
            if worksheet.max_row == 0 or worksheet.max_column == 0:
                continue

            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            header_map = {cell.value: index + 1 for index, cell in enumerate(worksheet[1])}
            for column_name in workbook_patch.A_END_DISPLAY_COLUMNS:
                col_idx = header_map.get(column_name)
                if col_idx is None:
                    continue
                worksheet.cell(row=1, column=col_idx).fill = a_end_fill
                for row_idx in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = a_end_fill

            for column_name in workbook_patch.Z_COLUMNS:
                col_idx = header_map.get(column_name)
                if col_idx is None:
                    continue
                worksheet.cell(row=1, column=col_idx).fill = z_end_fill
                for row_idx in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = z_end_fill

            for patch_column, patch_fill in patch_fills.items():
                col_idx = header_map.get(patch_column)
                if col_idx is None:
                    continue
                worksheet.cell(row=1, column=col_idx).fill = patch_fill
                for row_idx in range(2, worksheet.max_row + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = patch_fill

            if worksheet.title == "lldp_sp":
                for style in lldp_group_styles.values():
                    for column_name in style["columns"]:
                        col_idx = header_map.get(column_name)
                        if col_idx is None:
                            continue
                        worksheet.cell(row=1, column=col_idx).fill = style["fill"]
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = style["fill"]
                            cell.font = style["font"]

            for column_cells in worksheet.columns:
                values = [workbook_patch.norm(cell.value) for cell in column_cells[:100]]
                max_len = max((len(value) for value in values), default=0)
                worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 36)

        workbook.save(output_path)


class csv_patch:
    PATCH_COLUMNS = workbook_patch.PATCH_COLUMNS
    Z_COLUMNS = workbook_patch.Z_COLUMNS
    PATCH_NOT_FOUND = workbook_patch.PATCH_NOT_FOUND
    METADATA_COLUMNS = {
        "time",
        "ad",
        "block",
        "building",
        "channel",
        "deploymentgroupinstance",
        "device",
        "elevation",
        "fabricinstance",
        "instance",
        "interface",
        "job",
        "metric",
        "name",
        "planeinstance",
        "rack",
        "role",
        "serial",
        "units",
    }

    @staticmethod
    def compact_name(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value).lower())

    @staticmethod
    def find_column(columns: Iterable[str], *wanted_names: str):
        wanted = {csv_patch.compact_name(name) for name in wanted_names}
        for column in columns:
            if csv_patch.compact_name(column) in wanted:
                return column
        return None

    @staticmethod
    def require_column(columns: Iterable[str], *wanted_names: str) -> str:
        column = csv_patch.find_column(columns, *wanted_names)
        if column is None:
            raise ValueError(f"Missing required column. Expected one of: {', '.join(wanted_names)}")
        return column

    @staticmethod
    def read_csv(path: Path) -> pd.DataFrame:
        return pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")

    @staticmethod
    def numeric_text(value) -> str:
        text = workbook_patch.norm(value)
        match = re.search(r"[-+]?\d+(?:\.\d+)?(?:e[-+]?\d+)?", text, re.IGNORECASE)
        return match.group(0) if match else text

    @staticmethod
    def format_number(value: float) -> str:
        return f"{value:g}"

    @staticmethod
    def thresholds_from_filename(path: Path):
        for group in re.findall(r"\(([^()]*)\)", path.stem):
            numbers = re.findall(r"[-+]?\d+(?:\.\d+)?", group)
            if len(numbers) >= 2:
                values = [float(number) for number in numbers[:2]]
                return csv_patch.format_number(min(values)), csv_patch.format_number(max(values))
        return "", ""

    @staticmethod
    def detect_power_value_column(columns: Iterable[str]):
        preferred = ["RX power", "TX power", "Value", "Measured", "Measured (dBm)"]
        column = csv_patch.find_column(columns, *preferred)
        if column is not None:
            return column

        for candidate in columns:
            normalized = csv_patch.compact_name(candidate)
            if "power" in normalized and normalized not in csv_patch.METADATA_COLUMNS:
                return candidate

        for candidate in columns:
            if csv_patch.compact_name(candidate) not in csv_patch.METADATA_COLUMNS:
                return candidate

        return None

    @staticmethod
    def prepare_optics_csv(df: pd.DataFrame, csv_path: Path) -> pd.DataFrame:
        device_col = csv_patch.require_column(df.columns, "device", "Hostname")
        transceiver_col = csv_patch.require_column(df.columns, "name", "Transceiver")
        channel_col = csv_patch.require_column(df.columns, "channel", "Channel")
        rack_col = csv_patch.find_column(df.columns, "rack", "Rack")
        elevation_col = csv_patch.find_column(df.columns, "elevation", "Elevation")
        value_col = csv_patch.detect_power_value_column(df.columns)
        min_threshold, max_threshold = csv_patch.thresholds_from_filename(csv_path)

        out = df.copy()
        out["Hostname"] = out[device_col].map(workbook_patch.norm)
        out["Transceiver"] = out[transceiver_col].map(workbook_patch.norm)
        out["Channel"] = out[channel_col].map(workbook_patch.norm)
        out["Interface"] = out.apply(
            lambda row: workbook_patch.build_interface_from_transceiver(row["Transceiver"], row["Channel"]),
            axis=1,
        )
        out["Rack"] = out[rack_col].map(workbook_patch.norm) if rack_col else ""
        out["Elevation"] = out[elevation_col].map(workbook_patch.norm) if elevation_col else ""
        if value_col:
            out["Measured (dBm)"] = out[value_col].map(csv_patch.numeric_text)
        else:
            out["Measured (dBm)"] = ""
        out["Min Threshold (dBm)"] = min_threshold
        out["Max Threshold (dBm)"] = max_threshold
        out["_CSV_KIND"] = "optics"
        return out

    @staticmethod
    def prepare_endpoint_csv(df: pd.DataFrame, csv_path: Path) -> pd.DataFrame:
        del csv_path
        device_col = csv_patch.require_column(df.columns, "device", "Hostname")
        interface_col = csv_patch.require_column(df.columns, "interface", "Interface")
        rack_col = csv_patch.find_column(df.columns, "rack", "Rack")
        elevation_col = csv_patch.find_column(df.columns, "elevation", "Elevation")
        remote_device_col = csv_patch.find_column(df.columns, "remote_device", "remote device", "Remote Host")
        remote_interface_col = csv_patch.find_column(df.columns, "remote_interface", "remote interface", "Remote Interface")

        out = df.copy()
        out["Hostname"] = out[device_col].map(workbook_patch.norm)
        out["Interface"] = out[interface_col].map(workbook_patch.norm)
        out["Rack"] = out[rack_col].map(workbook_patch.norm) if rack_col else ""
        out["Elevation"] = out[elevation_col].map(workbook_patch.norm) if elevation_col else ""
        if remote_device_col:
            out["Remote Host"] = out[remote_device_col].map(workbook_patch.norm)
        if remote_interface_col:
            out["Remote Interface"] = out[remote_interface_col].map(workbook_patch.norm)
        out["_CSV_KIND"] = "endpoint"
        return out

    @staticmethod
    def prepare_csv(df: pd.DataFrame, csv_path: Path) -> pd.DataFrame:
        has_device = csv_patch.find_column(df.columns, "device", "Hostname") is not None
        has_interface = csv_patch.find_column(df.columns, "interface", "Interface") is not None
        has_transceiver = csv_patch.find_column(df.columns, "name", "Transceiver") is not None
        has_channel = csv_patch.find_column(df.columns, "channel", "Channel") is not None

        if has_device and has_transceiver and has_channel:
            return csv_patch.prepare_optics_csv(df, csv_path)
        if has_device and has_interface:
            return csv_patch.prepare_endpoint_csv(df, csv_path)

        raise ValueError(
            "Unsupported CSV format. Expected either device/name/channel optics rows "
            "or device/interface endpoint rows."
        )

    @staticmethod
    def merge_panel_side(out: pd.DataFrame, panels_df: pd.DataFrame, side: str) -> pd.DataFrame:
        key_column = "A_KEY" if side == "A" else "B_KEY"
        panel_columns = [
            "DeviceA_Host",
            "DeviceA_Interface",
            "DeviceA_Rack",
            "DeviceA_Elevation",
            "DeviceB_Host",
            "DeviceB_Interface",
            "DeviceB_Rack",
            "DeviceB_Elevation",
        ]
        view = panels_df.drop_duplicates(subset=[key_column], keep="first")
        return pd.merge(
            out[["_LOCAL_KEY"]],
            view[[key_column] + csv_patch.PATCH_COLUMNS + panel_columns],
            how="left",
            left_on="_LOCAL_KEY",
            right_on=key_column,
        )

    @staticmethod
    def fill_local_rack_elevation(
        out: pd.DataFrame,
        match: pd.DataFrame,
        mask: pd.Series,
        rack_source: str,
        elevation_source: str,
    ):
        for target_column, source_column in [("Rack", rack_source), ("Elevation", elevation_source)]:
            blank_target = out[target_column].map(workbook_patch.norm).eq("")
            fill_mask = mask & blank_target
            out.loc[fill_mask, target_column] = match.loc[fill_mask, source_column].values

    @staticmethod
    def assign_panel_columns(
        out: pd.DataFrame,
        match: pd.DataFrame,
        mask: pd.Series,
        patch_map: dict,
        z_map: dict,
    ):
        for target_column, source_column in patch_map.items():
            out.loc[mask, target_column] = match.loc[mask, source_column].values
        for target_column, source_column in z_map.items():
            out.loc[mask, target_column] = match.loc[mask, source_column].values

    @staticmethod
    def enrich_with_patch_panel(df: pd.DataFrame, panels_df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy().reset_index(drop=True)
        out["_LOCAL_KEY"] = out.apply(
            lambda row: workbook_patch.endpoint_key(row.get("Hostname", ""), row.get("Interface", "")),
            axis=1,
        )
        out["_PANEL_MATCH_SIDE"] = ""

        for column in csv_patch.PATCH_COLUMNS + csv_patch.Z_COLUMNS:
            out[column] = ""

        a_match = csv_patch.merge_panel_side(out, panels_df, "A").reset_index(drop=True)
        a_mask = a_match["A_KEY"].notna()
        csv_patch.assign_panel_columns(
            out,
            a_match,
            a_mask,
            {
                "Source_port": "Source_port",
                "DMARC1": "DMARC1",
                "DMARC2": "DMARC2",
                "Destination_port": "Destination_port",
            },
            {
                "Z Hostname": "DeviceB_Host",
                "Z Interface": "DeviceB_Interface",
                "Z Rack": "DeviceB_Rack",
                "Z Elevation": "DeviceB_Elevation",
            },
        )
        csv_patch.fill_local_rack_elevation(out, a_match, a_mask, "DeviceA_Rack", "DeviceA_Elevation")
        out.loc[a_mask, "_PANEL_MATCH_SIDE"] = "DeviceA"

        unmatched_mask = out["_PANEL_MATCH_SIDE"].eq("") & out["_LOCAL_KEY"].ne("")
        if unmatched_mask.any():
            b_match = csv_patch.merge_panel_side(out.loc[unmatched_mask].copy(), panels_df, "B").reset_index(drop=True)
            b_match_positions = b_match.index[b_match["B_KEY"].notna()]
            unmatched_indexes = out.index[unmatched_mask].to_series().reset_index(drop=True)
            b_indexes = unmatched_indexes.loc[b_match_positions].to_numpy()

            if len(b_indexes):
                for target_column, source_column in {
                    "Source_port": "Destination_port",
                    "DMARC1": "DMARC2",
                    "DMARC2": "DMARC1",
                    "Destination_port": "Source_port",
                }.items():
                    out.loc[b_indexes, target_column] = b_match.loc[b_match_positions, source_column].values

                for target_column, source_column in {
                    "Z Hostname": "DeviceA_Host",
                    "Z Interface": "DeviceA_Interface",
                    "Z Rack": "DeviceA_Rack",
                    "Z Elevation": "DeviceA_Elevation",
                }.items():
                    out.loc[b_indexes, target_column] = b_match.loc[b_match_positions, source_column].values

                rack_fill = out.loc[b_indexes, "Rack"].map(workbook_patch.norm).eq("")
                elevation_fill = out.loc[b_indexes, "Elevation"].map(workbook_patch.norm).eq("")
                out.loc[b_indexes[rack_fill.to_numpy()], "Rack"] = b_match.loc[
                    b_match_positions, "DeviceB_Rack"
                ].values[rack_fill.to_numpy()]
                out.loc[b_indexes[elevation_fill.to_numpy()], "Elevation"] = b_match.loc[
                    b_match_positions, "DeviceB_Elevation"
                ].values[elevation_fill.to_numpy()]
                out.loc[b_indexes, "_PANEL_MATCH_SIDE"] = "DeviceB"

        no_endpoint = out["_LOCAL_KEY"].eq("")
        no_match = out["_PANEL_MATCH_SIDE"].eq("") & ~no_endpoint
        out.loc[no_match, csv_patch.PATCH_COLUMNS + csv_patch.Z_COLUMNS] = csv_patch.PATCH_NOT_FOUND
        out.loc[no_endpoint, csv_patch.PATCH_COLUMNS + csv_patch.Z_COLUMNS] = ""
        return out

    @staticmethod
    def ordered_columns(df: pd.DataFrame) -> list[str]:
        kind = df["_CSV_KIND"].iloc[0] if "_CSV_KIND" in df.columns and len(df) else "endpoint"
        if kind == "optics":
            preferred = [
                "Hostname",
                "Interface",
                "Rack",
                "Elevation",
                "Transceiver",
                "Channel",
                "Measured (dBm)",
                "Min Threshold (dBm)",
                "Max Threshold (dBm)",
            ]
        else:
            preferred = ["Hostname", "Interface", "Rack", "Elevation", "Value"]

        preferred += csv_patch.PATCH_COLUMNS + csv_patch.Z_COLUMNS + ["Remote Host", "Remote Interface"]
        visible_preferred = [column for column in preferred if column in df.columns]
        rest = [
            column
            for column in df.columns
            if column not in visible_preferred and not column.startswith("_")
        ]
        return visible_preferred + rest

    @staticmethod
    def enrich_csv(csv_path: Path, panels_df: pd.DataFrame) -> pd.DataFrame:
        source_df = csv_patch.read_csv(csv_path)
        prepared_df = csv_patch.prepare_csv(source_df, csv_path)
        enriched_df = csv_patch.enrich_with_patch_panel(prepared_df, panels_df)
        return enriched_df[csv_patch.ordered_columns(enriched_df)]


SCRIPT_DIR = Path(__file__).resolve().parent
SUMMARY_SHEET = "summary"
FULL_REPORT_SUFFIX = "_full_report"
WITH_PP_SUFFIX = "_with_pp"
OUTPUT_TIMESTAMP_FORMAT = "%Y%m%d_%H%M%S"
DEFAULT_SITE_PREFIX = "jbp15"
JBP19_SITE_PREFIX = "jbp19"
DEFAULT_BATCH_SITE_PREFIX = DEFAULT_SITE_PREFIX
DEFAULT_INVENTORY_PATH = Path("~/.codex/memories/jbp_planar_ai2nd_inventory.json").expanduser()
DEFAULT_REGION = "jbp"
DEFAULT_BUILDING = "15"
DEFAULT_FABRIC_INSTANCE = "2"
DEFAULT_QCLI_SITE = DEFAULT_SITE_PREFIX
DEFAULT_DASHBOARD_GROUP_WINDOW_MINUTES = 15
JBP_QCLI_SITE_ALIASES = {
    "15": DEFAULT_SITE_PREFIX,
    DEFAULT_SITE_PREFIX: DEFAULT_SITE_PREFIX,
    "19": JBP19_SITE_PREFIX,
    JBP19_SITE_PREFIX: JBP19_SITE_PREFIX,
}
JBP19_QCLI_TARGET_ORDER = (
    "ipr",
    "dg1",
    "dg2",
    "dg3",
    "dg4",
    "dg5",
    "dg6",
    "dg7",
    "dg8",
    "dg9",
    "dg10",
    "dg11",
    "dg12",
    "dg13",
    "dg14",
    "dg15",
    "dg16",
)
JBP19_QCLI_DH_TAG_BY_TARGET = {
    "ipr": "IPR",
    "dg1": "J1-DH-GA",
    "dg2": "J1-DH-GA",
    "dg3": "J1-DH-GA",
    "dg4": "J1-DH-GA",
    "dg5": "J1-DH-GC",
    "dg6": "J1-DH-GC",
    "dg7": "J1-DH-GC",
    "dg8": "J1-DH-1A",
    "dg9": "J1-DH-1A",
    "dg10": "J1-DH-1A",
    "dg11": "J1-DH-1B",
    "dg12": "J1-DH-1B",
    "dg13": "J1-DH-1B",
}
JBP19_QCLI_DEFAULT_DH_TAG = "DG"
JBP19_QCLI_RACKS_BY_TARGET = {
    "dg1": "0205,0304",
    "dg2": "0206,0305",
    "dg3": "0604,0705",
    "dg4": "0605,0706",
    "dg5": "2405,2504",
    "dg6": "2406,2505",
    "dg7": "2813,2913",
    "dg8": "3405,3504",
    "dg9": "3406,3505",
    "dg10": "3813,3913",
    "dg11": "4405,4504",
    "dg12": "4406,4505",
    "dg13": "4813,4913",
}
JBP19_QFABT0_PANELS_NAME = "JBP19_pp_info (1).xlsx"
JBP19_QFABT1_IPR_PANELS_NAME = "JBP19_pp_info_all_q1_q2_q3_ipr.xlsx"
PHASE1_OPTICS_FILTER_THREAD_URL = "https://dyn.slack.com/archives/C0B4PJUTGCS/p1780807820198769"
PHASE1_RX_INPUT_WORSE_THAN_DBM = -3.0
PHASE1_TX_OUTPUT_WORSE_THAN_DBM = 0.0
NON_IPR_MAX_DBM = 4.5
NON_IPR_RELAX_RX_MIN_DBM = -4.0
NON_IPR_RELAX_TX_MIN_DBM = -1.0
NON_IPR_RELAX_MAX_DBM = 5.5
IPR_RX_MIN_DBM = -4.0
IPR_TX_MIN_DBM = -2.0
IPR_MAX_DBM = 2.0
IPR_RELAX_RX_MIN_DBM = -5.0
IPR_RELAX_TX_MIN_DBM = -3.0
IPR_RELAX_MAX_DBM = 3.0
RELAX_NON_IPR_OPTICS = False
RELAX_IPR_OPTICS = False
PHASE1_DEFAULT_DG_SCOPE = {
    "DG1": {"milestone": "JBP15.1", "data_hall": "B2 DH1C"},
    "DG2": {"milestone": "JBP15.1", "data_hall": "B2 DH1C"},
    "DG3": {"milestone": "JBP15.1", "data_hall": "B2 DH1C"},
    "DG4": {"milestone": "JBP15.1", "data_hall": "B2 DH1C"},
    "DG5": {"milestone": "JBP15.1", "data_hall": "B2 DH1C"},
    "DG6": {"milestone": "JBP15.4", "data_hall": "C1 DH1D"},
    "DG7": {"milestone": "JBP15.4", "data_hall": "C1 DH1D"},
    "DG8": {"milestone": "JBP15.5", "data_hall": "C1 DHGC"},
    "DG9": {"milestone": "JBP15.5", "data_hall": "C1 DHGC"},
    "DG10": {"milestone": "JBP15.6", "data_hall": "C1 DHGD"},
    "DG11": {"milestone": "JBP15.6", "data_hall": "C1 DHGD"},
}
PHASE1_SPECIAL_OPTICS_SCOPES = {"IPR", "SPINE"}
DASHBOARD_DATA_TIMESTAMP_RE = re.compile(
    r"data-(\d{4}-\d{2}-\d{2}) (\d{2})_(\d{2})_(\d{2})"
)
DASHBOARD_REPORT_COLUMNS = [
    "Time",
    "Issue Type",
    "Power Type",
    "fabric_instance",
    "device",
    "rack",
    "elevation",
    "name",
    "channel",
    "interface",
    "DG",
    "serial",
    "Value",
    "remote_device",
    "remote_interface",
    "Source File",
    "Local PP",
    "Peer PP",
    "Peer Endpoint",
    "Peer Rack",
    "PP_A",
    "PP_Z",
    "Matched Endpoint",
]
DASHBOARD_SUMMARY_COLUMNS = [
    "DG",
    "RX Optics",
    "TX Optics",
    "Total Optics",
    "Pre-FEC",
    "FEC Bin",
    "Grand Total",
]
DASHBOARD_KIND_ORDER = {
    "rx": 0,
    "tx": 1,
    "pre_fec": 2,
    "fec_bin": 3,
}
COUNT_SUMMARY_SHEETS = {
    "lldp": ("lldp_sp", "full_path_lldp_with_int_down"),
    "interface": ("interfaces_sp", "interfaces_sp_with_pp"),
    "optics": ("optics_rx_tx_threshold", "optics_rx_tx_threshold_with_pp"),
    "optics_temp": ("optics_temp", "optics_temp_with_pp"),
    "combined_fec": ("combined_fec", "combined_fec_with_pp"),
}
COUNT_SUMMARY_KEYS = ("lldp", "interface", "optics", "optics_temp", "combined_fec")
DETAILED_COUNT_SUMMARY_KEYS = (
    "lldp",
    "interface",
    "optics_tx",
    "optics_rx",
    "optics_temp",
    "fec_bin",
    "pre_fec",
)
COUNT_SUMMARY_LABELS = {
    "t0": "T0 Side",
    "ipr": "T1<>IPR Side",
}
COUNT_SUMMARY_SHEET = "DG Error Counts"
COUNT_SUMMARY_TABLE_SHEET = "Latest Counts"
COUNT_SUMMARY_THREAD_URL = (
    "https://dyn.slack.com/archives/C0AU7KB98JH/"
    "p1780665720527989?thread_ts=1780580597.386099&cid=C0AU7KB98JH"
)
COUNT_SUMMARY_PHASE_BASELINES = {
    "Phase 1": {
        "DG1": 70,
        "DG2": 24,
        "DG3": 48,
        "DG4": 18,
        "DG5": 18,
        "DG6": 7,
        "DG7": 44,
        "DG8": 34,
        "DG9": 86,
        "DG10": 61,
        "DG11": 43,
    }
}
COUNT_SUMMARY_PHASE_HEADERS = [
    "Phase number",
    "DG number",
    "Past Total",
    "lldp",
    "interface",
    "optics(tx)",
    "optics(rx)",
    "optics(temp)",
    "fec bin",
    "pre fec",
    "Present Total",
    "Total Delta",
]

EXCEL_SHEET_ENRICHERS: dict[str, tuple[str, Callable[[pd.DataFrame, pd.DataFrame, pd.DataFrame], pd.DataFrame]]] = {
    "lldp_sp": ("lldp_sp", workbook_patch.enrich_lldp_sheet),
    "interfaces_sp": ("interfaces_sp", workbook_patch.enrich_interfaces_sheet),
    "pre_fec_ber_threshold": ("pre_fec_ber_threshold", workbook_patch.enrich_combined_fec_sheet),
    "combined_fec": ("combined_fec", workbook_patch.enrich_combined_fec_sheet),
    "optics_temp": ("optics_temp", workbook_patch.enrich_optics_temp_sheet),
    "optics_rx_tx_threshold": ("optics_rx_tx_threshold", workbook_patch.enrich_optics_sheet),
}

QCLI_FULL_REPORT_SHEET_ENRICHERS: dict[str, tuple[str, Callable[[pd.DataFrame, pd.DataFrame, pd.DataFrame], pd.DataFrame]]] = {
    "full_path_lldp_with_int_down": ("lldp_sp", workbook_patch.enrich_lldp_sheet),
    "interfaces_sp_with_pp": ("interfaces_sp", workbook_patch.enrich_interfaces_sheet),
    "pre_fec_ber_threshold_with_pp": ("pre_fec_ber_threshold", workbook_patch.enrich_combined_fec_sheet),
    "combined_fec_with_pp": ("combined_fec", workbook_patch.enrich_combined_fec_sheet),
    "optics_temp_with_pp": ("optics_temp", workbook_patch.enrich_optics_temp_sheet),
    "optics_rx_tx_threshold_with_pp": ("optics_rx_tx_threshold", workbook_patch.enrich_optics_sheet),
}

ALL_EXCEL_SHEET_ENRICHERS = {
    **EXCEL_SHEET_ENRICHERS,
    **QCLI_FULL_REPORT_SHEET_ENRICHERS,
}
REPLACED_INPUT_SHEETS = set(ALL_EXCEL_SHEET_ENRICHERS)

GENERATED_SHEETS = {output for output, _ in EXCEL_SHEET_ENRICHERS.values()}
GENERATED_SHEETS.add("fec_bin_threshold_sp")
REQUIRED_OUTPUT_SHEETS = (
    "lldp_sp",
    "interfaces_sp",
    "optics_temp",
    "optics_rx_tx_threshold",
    "pre_fec_ber_threshold",
    "combined_fec",
    "fec_bin_threshold_sp",
    "endpoint_threshold",
)
PATCH_COLUMNS = set(workbook_patch.PATCH_COLUMNS + workbook_patch.Z_COLUMNS)


@dataclass(frozen=True)
class DashboardCsvSource:
    path: Path
    kind: str
    timestamp: datetime


@dataclass
class DashboardCsvGroup:
    sources: dict[str, DashboardCsvSource]

    @property
    def start(self) -> datetime:
        return min(source.timestamp for source in self.sources.values())

    @property
    def end(self) -> datetime:
        return max(source.timestamp for source in self.sources.values())

    @property
    def output_timestamp(self) -> datetime:
        return self.end


@dataclass(frozen=True)
class DashboardPanelRecord:
    a_endpoint: str
    z_endpoint: str
    a_rack: str
    z_rack: str
    pp_a: str
    pp_z: str


def norm(value) -> str:
    return workbook_patch.norm(value)


def compact_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def site_prefix_values(site_prefix: str) -> list[str]:
    return [
        item.strip().lower()
        for item in str(site_prefix or "").split(",")
        if item.strip()
    ]


def filename_site_name(path: Path) -> str | None:
    match = re.search(r"\b(jbp\d+)-", path.name, re.IGNORECASE)
    return match.group(1).lower() if match else None


def path_matches_site_prefix(path: Path, site_prefix: str) -> bool:
    prefixes = site_prefix_values(site_prefix)
    if not prefixes:
        return True
    lower_name = path.name.lower()
    return any(lower_name.startswith(prefix) for prefix in prefixes)


def qcli_tag_key(value: str) -> str:
    return compact_name(value).removeprefix("jbp15")


def normalize_jbp_qcli_site(value: str | None = None) -> str:
    key = compact_name(value or DEFAULT_QCLI_SITE)
    site = JBP_QCLI_SITE_ALIASES.get(key)
    if site is None:
        available = ", ".join(sorted(JBP_QCLI_SITE_ALIASES))
        raise ValueError(f"Unknown JBP qcli site {value!r}. Available: {available}")
    return site


def ordered_unique(values: Iterable[object]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def load_planar_inventory(inventory_path: Path) -> dict:
    path = inventory_path.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"JBP planar inventory JSON not found: {path}")
    with path.open() as handle:
        return json.load(handle)


def jbp15_planar_build(inventory: dict) -> dict:
    try:
        return inventory["planar_builds"]["jbp15"]
    except KeyError as exc:
        raise ValueError("Inventory JSON does not contain planar_builds.jbp15") from exc


def t1_plane_from_tag(tag: str) -> int | None:
    key = qcli_tag_key(tag)
    pg_match = re.fullmatch(r"(?:pg)?15([1-4])", key)
    if pg_match:
        return int(pg_match.group(1))

    plane_match = re.fullmatch(r"(?:t1)?(?:plane|p)([1-4])", key)
    if plane_match:
        return int(plane_match.group(1))

    return None


def jbp15_dg_number_from_target(target: str) -> str | None:
    key = qcli_tag_key(target)
    dg_match = re.fullmatch(r"(?:pg|dg)?([0-9]+)", key)
    if not dg_match:
        return None
    return str(int(dg_match.group(1)))


def racks_for_qcli_tag(inventory: dict, tag: str) -> list[str]:
    build = jbp15_planar_build(inventory)
    key = qcli_tag_key(tag)
    design_scope = build.get("design_rack_scope", {})

    direct_tags = {
        "all": "network_racks",
        "network": "network_racks",
        "networkracks": "network_racks",
        "qfabric": "network_racks",
        "qfabt0": "qfabt0_racks",
        "t0": "qfabt0_racks",
        "qfabt1": "qfabt1_racks",
        "t1": "qfabt1_racks",
        "qfabip": "qfabip_racks",
        "qfabipr": "qfabip_racks",
        "ip": "qfabip_racks",
        "ipr": "qfabip_racks",
        "gpu": "gpu_racks",
        "gpuracks": "gpu_racks",
    }
    if key in direct_tags:
        racks = design_scope.get(direct_tags[key], [])
        if not racks and key == "ipr":
            racks = build.get("ipr_rack", {}).get("rack_positions", [])
        return ordered_unique(racks)

    if key in {"t1spine", "t1spines", "spine", "spines"}:
        racks: list[str] = []
        for group in build.get("t1_spine_groups", []):
            racks.extend(group.get("rack_positions", []))
        return ordered_unique(racks)

    pg_match = re.fullmatch(r"(?:pg|dg)?([0-9]+)", key)
    if pg_match:
        placement_group = pg_match.group(1)
        for group in build.get("placement_groups", []):
            aliases = {
                qcli_tag_key(str(group.get("placement_group", ""))),
                qcli_tag_key(str(group.get("source_placement_group_label", ""))),
                qcli_tag_key(str(group.get("deployment_group_alias", ""))),
            }
            if key in aliases or placement_group in aliases:
                return ordered_unique(group.get("qfabt0_racks", []))

    available = [
        "qfabt1",
        "qfabt0",
        "qfabip",
        "network",
        "gpu",
        "t1-spine",
        "PG1",
        "DG1",
        "plane1",
        "PG151",
    ]
    raise ValueError(f"Unknown qcli command tag {tag!r}. Examples: {', '.join(available)}")


def qcli_command_parts_for_tag(
    inventory: dict,
    tag: str,
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    plane = t1_plane_from_tag(tag)
    command = [
        "qcli",
        "hc-summary",
        "--region",
        DEFAULT_REGION,
        "--building",
        DEFAULT_BUILDING,
        "--instance",
        DEFAULT_FABRIC_INSTANCE,
    ]
    if state:
        command.extend(["--state", state])

    if plane is not None:
        command.extend(["--spectrum", "--t1-reports", "--nvidia-linkflap-clear", "--planar", str(plane), "--tier", "1"])
    else:
        dg_number = jbp15_dg_number_from_target(tag)
        if dg_number is not None:
            command.extend(["--spectrum", "--customtag", "test_spectrum_fec_ber", "--failures-only", "--filter_pg", "--t1-reports", "--nvidia-linkflap-clear", "--slack"])
            if no_apex_update:
                command.append("--no-apex-update")
            command.extend(["--deployment-group", dg_number])
            return command
        else:
            racks = racks_for_qcli_tag(inventory, tag)
            if not racks:
                raise ValueError(f"Inventory tag {tag!r} resolved to no racks")
            rack_list = ",".join(racks)
            command.extend([
                "--rack",
                rack_list,
                "--spectrum",
                "--customtag",
                "test_spectrum_fec_ber",
                "--failures-only",
                "--filter_rack",
                rack_list,
                "--filter_pg",
                "--t1-reports",
                "--nvidia-linkflap-clear",
            ])

    if no_apex_update:
        command.append("--no-apex-update")
    return command


def shell_command(parts: Iterable[str]) -> str:
    return " ".join(shlex.quote(part) for part in parts)


def qcli_commands_with_optics_relax(commands: str, relax: bool) -> str:
    if not relax:
        return commands
    lines: list[str] = []
    for line in commands.splitlines():
        if " --column 1" in f" {line}":
            lines.append(line)
        elif " --deployment-group " in line:
            before_dg, dg_part = line.rsplit(" --deployment-group ", 1)
            lines.append(f"{before_dg} --optics-relax --deployment-group {dg_part}")
        else:
            lines.append(f"{line} --optics-relax")
    return "\n".join(lines)


def jbp15_qcli_all_targets(inventory: dict) -> list[str]:
    build = jbp15_planar_build(inventory)
    targets: list[str] = []
    for group in build.get("placement_groups", []):
        alias = norm(group.get("deployment_group_alias"))
        if not alias:
            placement_group = norm(group.get("placement_group"))
            alias = f"DG{placement_group}" if placement_group else ""
        if alias:
            targets.append(alias.upper())
    return ordered_unique(targets)


def expand_jbp15_qcli_target_value(value: str, inventory: dict) -> list[str]:
    text = str(value).strip()
    if not text:
        return []

    key = compact_name(text)
    if key in {"all", "alltargets", "commands", "qclicommands"}:
        return jbp15_qcli_all_targets(inventory)

    range_match = re.fullmatch(
        r"(?:jbp15)?(?:pg|dg)?([0-9]+)\s*-\s*(?:jbp15)?(?:pg|dg)?([0-9]+)",
        text,
        flags=re.IGNORECASE,
    )
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        step = 1 if end >= start else -1
        return [f"DG{number}" for number in range(start, end + step, step)]

    if t1_plane_from_tag(text) is not None:
        return [text]

    single_dg_match = re.fullmatch(r"(?:jbp15)?(?:pg|dg)?([0-9]+)", key)
    if single_dg_match:
        return [f"DG{int(single_dg_match.group(1))}"]

    return [text]


def parse_jbp15_qcli_targets(targets: str | Iterable[str] | None, inventory: dict) -> list[str]:
    if targets is None:
        return jbp15_qcli_all_targets(inventory)
    if isinstance(targets, str):
        raw_values = re.split(r"[\s,]+", targets.strip())
    else:
        raw_values = [str(value).strip() for value in targets]

    expanded: list[str] = []
    for value in raw_values:
        expanded.extend(expand_jbp15_qcli_target_value(value, inventory))
    if not expanded:
        return jbp15_qcli_all_targets(inventory)

    result: list[str] = []
    seen: set[str] = set()
    for target in expanded:
        key = qcli_tag_key(target)
        if key in seen:
            continue
        seen.add(key)
        result.append(target)
    return result


def jbp15_qcli_label_for_target(target: str) -> str:
    plane = t1_plane_from_tag(target)
    if plane is not None:
        return f"PG15{plane}"

    key = qcli_tag_key(target)
    dg_match = re.fullmatch(r"(?:pg|dg)?([0-9]+)", key)
    if dg_match:
        return f"DG{int(dg_match.group(1))}"
    return str(target).upper()


def jbp15_qcli_label_for_targets(targets: Iterable[str]) -> str:
    labels = [jbp15_qcli_label_for_target(target) for target in targets]
    if len(labels) == 1:
        return labels[0]

    dg_numbers = []
    for label in labels:
        match = re.fullmatch(r"DG([0-9]+)", label)
        if not match:
            return ",".join(labels)
        dg_numbers.append(int(match.group(1)))

    if dg_numbers == list(range(dg_numbers[0], dg_numbers[-1] + 1)):
        return f"DG{dg_numbers[0]}-DG{dg_numbers[-1]}"
    return ",".join(f"DG{number}" for number in dg_numbers)


def jbp15_qcli_command_parts_for_racks(
    racks: Iterable[str],
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    rack_list = ",".join(ordered_unique(racks))
    if not rack_list:
        raise ValueError("JBP15 qcli target group resolved to no racks")

    command = [
        "qcli",
        "hc-summary",
        "--region",
        DEFAULT_REGION,
        "--building",
        DEFAULT_BUILDING,
        "--instance",
        DEFAULT_FABRIC_INSTANCE,
    ]
    if state:
        command.extend(["--state", state])
    command.extend([
        "--rack",
        rack_list,
        "--spectrum",
        "--customtag",
        "test_spectrum_fec_ber",
        "--failures-only",
        "--filter_rack",
        rack_list,
        "--filter_pg",
        "--t1-reports",
        "--nvidia-linkflap-clear",
    ])
    if no_apex_update:
        command.append("--no-apex-update")
    return command


def jbp15_qcli_command_parts_for_dgs(
    dg_numbers: Iterable[str],
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    dg_list = ",".join(ordered_unique(str(int(number)) for number in dg_numbers))
    if not dg_list:
        raise ValueError("JBP15 qcli target group resolved to no DGs")

    command = [
        "qcli",
        "hc-summary",
        "--region",
        DEFAULT_REGION,
        "--building",
        DEFAULT_BUILDING,
        "--instance",
        DEFAULT_FABRIC_INSTANCE,
    ]
    if state:
        command.extend(["--state", state])
    command.extend(["--spectrum", "--customtag", "test_spectrum_fec_ber", "--failures-only", "--filter_pg", "--t1-reports", "--nvidia-linkflap-clear", "--slack"])
    if no_apex_update:
        command.append("--no-apex-update")
    command.extend(["--deployment-group", dg_list])
    return command


def jbp15_qcli_command_groups(targets: Iterable[str]) -> list[tuple[str, list[str]]]:
    groups: list[tuple[str, list[str]]] = []
    current_rack_targets: list[str] = []

    def flush_rack_targets() -> None:
        if current_rack_targets:
            groups.append(("racks", current_rack_targets.copy()))
            current_rack_targets.clear()

    for target in targets:
        if t1_plane_from_tag(target) is not None:
            flush_rack_targets()
            groups.append(("plane", [target]))
        else:
            current_rack_targets.append(target)
    flush_rack_targets()
    return groups


def jbp15_qcli_command_lines(
    targets: str | Iterable[str] | None = None,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    inventory = load_planar_inventory(inventory_path)
    target_keys = parse_jbp15_qcli_targets(targets, inventory)

    lines: list[str] = []
    for group_kind, group_targets in jbp15_qcli_command_groups(target_keys):
        if group_kind == "plane":
            command_parts = qcli_command_parts_for_tag(inventory, group_targets[0], state, no_apex_update)
        else:
            dg_numbers = [jbp15_dg_number_from_target(target) for target in group_targets]
            if all(number is not None for number in dg_numbers):
                command_parts = jbp15_qcli_command_parts_for_dgs(
                    [number for number in dg_numbers if number is not None],
                    state,
                    no_apex_update,
                )
            else:
                rack_values: list[str] = []
                for target in group_targets:
                    rack_values.extend(racks_for_qcli_tag(inventory, target))
                command_parts = jbp15_qcli_command_parts_for_racks(rack_values, state, no_apex_update)
        lines.append(shell_command(command_parts))
    return lines


def jbp15_qcli_commands_text(
    targets: str | Iterable[str] | None = None,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    state: str = "deployed",
    no_apex_update: bool = False,
) -> str:
    return "\n".join(jbp15_qcli_command_lines(targets, inventory_path, state, no_apex_update))


def jbp19_qcli_target_key(value: str) -> str:
    key = compact_name(value).removeprefix("jbp19")
    aliases = {
        "alltargets": "all",
        "commands": "all",
        "qclicommands": "all",
        "ip": "ipr",
        "iprack": "ipr",
        "ipracks": "ipr",
    }
    if key in aliases:
        return aliases[key]

    dg_match = re.fullmatch(r"(?:dg|pg)?([0-9]+)", key)
    if dg_match:
        return f"dg{int(dg_match.group(1))}"
    return key


def expand_jbp19_qcli_target_value(value: str) -> list[str]:
    text = str(value).strip()
    if not text:
        return []

    range_match = re.fullmatch(
        r"(?:jbp19)?(?:pg|dg)?([0-9]+)\s*-\s*(?:jbp19)?(?:pg|dg)?([0-9]+)",
        text,
        flags=re.IGNORECASE,
    )
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        step = 1 if end >= start else -1
        return [f"dg{number}" for number in range(start, end + step, step)]

    return [jbp19_qcli_target_key(text)]


def parse_jbp19_qcli_targets(targets: str | Iterable[str] | None) -> list[str]:
    if targets is None:
        return list(JBP19_QCLI_TARGET_ORDER)
    if isinstance(targets, str):
        raw_values = re.split(r"[\s,]+", targets.strip())
    else:
        raw_values = [str(value).strip() for value in targets]

    keys: list[str] = []
    for value in raw_values:
        keys.extend(expand_jbp19_qcli_target_value(value))
    if not keys or "all" in keys:
        return list(JBP19_QCLI_TARGET_ORDER)

    unknown = [key for key in keys if key not in JBP19_QCLI_TARGET_ORDER]
    if unknown:
        available = ", ".join(
            ["all", "ipr"]
            + [target.upper() for target in JBP19_QCLI_TARGET_ORDER if target.startswith("dg")]
        )
        raise ValueError(f"Unknown JBP19 qcli target(s): {', '.join(unknown)}. Available: {available}")

    result: list[str] = []
    seen: set[str] = set()
    for key in keys:
        if key in seen:
            continue
        seen.add(key)
        result.append(key)
    return result


def jbp19_qcli_dh_key(value: str) -> str:
    key = compact_name(value)
    return key.removeprefix("room")


def jbp19_qcli_dh_matches(value: str, wanted: str) -> bool:
    value_key = jbp19_qcli_dh_key(value)
    wanted_key = jbp19_qcli_dh_key(wanted)
    aliases = {
        value_key,
        value_key.removeprefix("j1"),
        value_key.removeprefix("dh"),
        value_key.removeprefix("j1dh"),
    }
    return wanted_key in aliases


def jbp19_qcli_dh_tag_for_target(target: str) -> str:
    return JBP19_QCLI_DH_TAG_BY_TARGET.get(target, JBP19_QCLI_DEFAULT_DH_TAG)


def jbp19_qcli_command_parts_for_racks(
    racks: str | None = None,
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    command = [
        "qcli",
        "hc-summary",
        "--region",
        DEFAULT_REGION,
        "--building",
        "19",
        "--instance",
        DEFAULT_FABRIC_INSTANCE,
    ]
    if state:
        command.extend(["--state", state])
    if not racks:
        command.extend(["--spectrum", "--customtag", "test_spectrum_fec_ber", "--failures-only", "--t1-reports", "--nvidia-linkflap-clear", "--column", "1"])
        if no_apex_update:
            command.append("--no-apex-update")
        return command

    command.extend([
        "--rack",
        racks,
        "--spectrum",
        "--customtag",
        "test_spectrum_fec_ber",
        "--failures-only",
        "--filter_rack",
        racks,
        "--filter_pg",
        "--t1-reports",
        "--nvidia-linkflap-clear",
    ])
    if no_apex_update:
        command.append("--no-apex-update")
    return command


def jbp19_qcli_command_parts_for_dgs(
    targets: Iterable[str],
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    dg_numbers: list[str] = []
    for target in targets:
        match = re.fullmatch(r"dg([0-9]+)", target)
        if not match:
            raise ValueError(f"JBP19 target {target!r} is not a DG target")
        dg_numbers.append(match.group(1))

    dg_list = ",".join(ordered_unique(dg_numbers))
    if not dg_list:
        raise ValueError("JBP19 qcli target group resolved to no DGs")

    command = [
        "qcli",
        "hc-summary",
        "--region",
        DEFAULT_REGION,
        "--building",
        "19",
        "--instance",
        DEFAULT_FABRIC_INSTANCE,
    ]
    if state:
        command.extend(["--state", state])
    command.extend(["--spectrum", "--customtag", "test_spectrum_fec_ber", "--failures-only", "--filter_pg", "--t1-reports", "--nvidia-linkflap-clear", "--slack"])
    if no_apex_update:
        command.append("--no-apex-update")
    command.extend(["--deployment-group", dg_list])
    return command


def jbp19_qcli_command_parts_for_target(
    target: str,
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    if target == "ipr":
        return jbp19_qcli_command_parts_for_racks(None, state, no_apex_update)
    return jbp19_qcli_command_parts_for_dgs([target], state, no_apex_update)


def jbp19_qcli_label_for_targets(targets: Iterable[str]) -> str:
    target_labels = ["IPR Racks" if target == "ipr" else target.upper() for target in targets]
    if len(target_labels) == 1:
        return target_labels[0]

    dg_numbers = []
    for target in targets:
        match = re.fullmatch(r"dg([0-9]+)", target)
        if not match:
            return ",".join(target_labels)
        dg_numbers.append(int(match.group(1)))

    if dg_numbers == list(range(dg_numbers[0], dg_numbers[-1] + 1)):
        return f"DG{dg_numbers[0]}-DG{dg_numbers[-1]}"
    return ",".join(f"DG{number}" for number in dg_numbers)


def jbp19_qcli_command_parts_for_targets(
    targets: Iterable[str],
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    return jbp19_qcli_command_parts_for_dgs(targets, state, no_apex_update)


def jbp19_qcli_command_groups(targets: Iterable[str]) -> list[tuple[str, list[str]]]:
    groups: list[tuple[str, list[str]]] = []
    dg_targets: list[str] = []

    def flush_dg_targets() -> None:
        if dg_targets:
            groups.append(("dg", dg_targets.copy()))
            dg_targets.clear()

    for target in targets:
        if target == "ipr":
            flush_dg_targets()
            groups.append(("ipr", [target]))
            continue
        dg_targets.append(target)
    flush_dg_targets()
    return groups


def jbp19_qcli_command_lines(
    targets: str | Iterable[str] | None = None,
    dh_tag: str = "",
    state: str = "deployed",
    no_apex_update: bool = False,
) -> list[str]:
    target_keys = parse_jbp19_qcli_targets(targets)
    if dh_tag.strip():
        target_keys = [
            key
            for key in target_keys
            if jbp19_qcli_dh_matches(jbp19_qcli_dh_tag_for_target(key), dh_tag)
        ]
        if not target_keys:
            available = ", ".join(
                ordered_unique(
                    [jbp19_qcli_dh_tag_for_target(target) for target in JBP19_QCLI_TARGET_ORDER]
                )
            )
            raise ValueError(
                f"No JBP19 qcli targets found for dh tag {dh_tag!r}. "
                f"Available dh tags: {available}"
            )

    lines: list[str] = []
    for dh_tag_value, group_targets in jbp19_qcli_command_groups(target_keys):
        if group_targets == ["ipr"]:
            command_parts = jbp19_qcli_command_parts_for_target(
                "ipr",
                state,
                no_apex_update,
            )
        else:
            command_parts = jbp19_qcli_command_parts_for_targets(
                group_targets,
                state,
                no_apex_update,
            )
        lines.append(shell_command(command_parts))
    return lines


def jbp19_qcli_commands_text(
    targets: str | Iterable[str] | None = None,
    dh_tag: str = "",
    state: str = "deployed",
    no_apex_update: bool = False,
) -> str:
    return "\n".join(jbp19_qcli_command_lines(targets, dh_tag, state, no_apex_update))


def jbp_site_qcli_commands_text(
    site: str | None = None,
    targets: str | Iterable[str] | None = None,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    dh_tag: str = "",
    state: str = "deployed",
    no_apex_update: bool = False,
) -> str:
    site_key = normalize_jbp_qcli_site(site)
    if site_key == DEFAULT_SITE_PREFIX:
        return jbp15_qcli_commands_text(targets, inventory_path, state, no_apex_update)
    return jbp19_qcli_commands_text(targets, dh_tag, state, no_apex_update)


def output_stem_prefix(source_path: Path) -> str:
    stem = source_path.stem
    if stem.endswith(FULL_REPORT_SUFFIX):
        return f"{stem[:-len(FULL_REPORT_SUFFIX)]}{WITH_PP_SUFFIX}"
    elif stem.endswith(WITH_PP_SUFFIX):
        return f"{stem}_refreshed"
    return f"{stem}{WITH_PP_SUFFIX}"


def default_output_path(source_path: Path, timestamp: str | None = None) -> Path:
    timestamp = timestamp or datetime.now().strftime(OUTPUT_TIMESTAMP_FORMAT)
    return source_path.with_name(f"{output_stem_prefix(source_path)}_{timestamp}.xlsx")


def existing_timestamped_outputs(source_path: Path) -> list[Path]:
    return sorted(source_path.parent.glob(f"{output_stem_prefix(source_path)}_[0-9]" + "*.xlsx"))


def discover_panels_path(batch_dir: Path, explicit_path: str | None = None) -> Path:
    if explicit_path:
        panels_path = Path(explicit_path).expanduser().resolve()
        if not panels_path.exists():
            raise FileNotFoundError(f"Patch panel workbook not found: {panels_path}")
        return panels_path

    patterns = [
        "jbp_combined_pp*.xlsx",
        "*combined*pp*.xlsx",
        "JBP*_pp_info.xlsx",
        "*pp_info*.xlsx",
    ]
    seen: set[Path] = set()
    for pattern in patterns:
        candidates: list[Path] = []
        for hit in batch_dir.glob(pattern):
            if not hit.is_file():
                continue
            resolved = hit.resolve()
            if resolved not in seen:
                seen.add(resolved)
                candidates.append(resolved)
        if candidates:
            return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))

    raise FileNotFoundError(
        f"No PP matrix workbook found in {batch_dir}. Expected a file like jbp_combined_pp*.xlsx."
    )


def discover_named_panels_path(batch_dir: Path, file_name: str) -> Path | None:
    for base_dir in (batch_dir, SCRIPT_DIR):
        candidate = base_dir / file_name
        if candidate.exists():
            return candidate.resolve()
    return None


def discover_source_panels_path(
    batch_dir: Path,
    source_path: Path,
    explicit_path: str | None = None,
) -> Path:
    if explicit_path:
        return discover_panels_path(batch_dir, explicit_path)

    site_name = filename_site_name(source_path)
    lower_name = source_path.name.lower()
    if site_name == JBP19_SITE_PREFIX:
        preferred_name = (
            JBP19_QFABT1_IPR_PANELS_NAME
            if "-q2-ip-" in lower_name
            else JBP19_QFABT0_PANELS_NAME
        )
        panels_path = discover_named_panels_path(batch_dir, preferred_name)
        if panels_path is not None:
            return panels_path
        raise FileNotFoundError(f"JBP19 PP matrix workbook not found: {batch_dir / preferred_name}")

    return discover_panels_path(batch_dir)


def load_pp_a_z_panels_df(panels_path: Path) -> pd.DataFrame | None:
    workbook = pd.ExcelFile(panels_path)
    sheet_name = "T0 PANELS" if "T0 PANELS" in workbook.sheet_names else workbook.sheet_names[0]
    raw = pd.read_excel(panels_path, sheet_name=sheet_name, dtype=str, keep_default_na=False)
    required_map = {
        "DeviceA": ("DEVICE_A", "DeviceA"),
        "RackA": ("RackA", "Rack A"),
        "Source_port": ("PP_A", "Source_port"),
        "Destination_port": ("PP_Z", "Destination_port"),
        "DeviceB": ("DEVICE_Z", "DeviceB"),
        "RackB": ("RACK_Z", "RackB", "Rack B"),
    }
    optional_map = {
        "DMARC1": ("DMARC1", "DMARC 1"),
        "DMARC2": ("DMARC2", "DMARC 2"),
    }

    selected: dict[str, str] = {}
    for output_name, wanted_names in required_map.items():
        actual = find_dashboard_column(raw.columns, *wanted_names)
        if actual is None:
            return None
        selected[output_name] = actual
    for output_name, wanted_names in optional_map.items():
        actual = find_dashboard_column(raw.columns, *wanted_names)
        if actual is not None:
            selected[output_name] = actual

    panels_df = pd.DataFrame(
        {
            "DeviceA": raw[selected["DeviceA"]],
            "RackA": raw[selected["RackA"]],
            "Source_port": raw[selected["Source_port"]],
            "DMARC1": raw[selected["DMARC1"]] if "DMARC1" in selected else "",
            "DMARC2": raw[selected["DMARC2"]] if "DMARC2" in selected else "",
            "Destination_port": raw[selected["Destination_port"]],
            "DeviceB": raw[selected["DeviceB"]],
            "RackB": raw[selected["RackB"]],
        }
    )

    for column in panels_df.columns:
        panels_df[column] = panels_df[column].map(norm)

    split_a = panels_df["DeviceA"].map(workbook_patch.split_device_endpoint)
    split_b = panels_df["DeviceB"].map(workbook_patch.split_device_endpoint)
    panels_df[["DeviceA_Host", "DeviceA_Interface"]] = pd.DataFrame(split_a.tolist(), index=panels_df.index)
    panels_df[["DeviceB_Host", "DeviceB_Interface"]] = pd.DataFrame(split_b.tolist(), index=panels_df.index)
    panels_df["A_KEY"] = panels_df.apply(
        lambda row: workbook_patch.endpoint_key(row["DeviceA_Host"], row["DeviceA_Interface"]),
        axis=1,
    )
    panels_df["B_KEY"] = panels_df.apply(
        lambda row: workbook_patch.endpoint_key(row["DeviceB_Host"], row["DeviceB_Interface"]),
        axis=1,
    )
    panels_df[["DeviceA_Rack", "DeviceA_Elevation"]] = pd.DataFrame(
        panels_df["RackA"].map(workbook_patch.parse_patch_rack_elevation).tolist(),
        index=panels_df.index,
    )
    panels_df[["DeviceB_Rack", "DeviceB_Elevation"]] = pd.DataFrame(
        panels_df["RackB"].map(workbook_patch.parse_patch_rack_elevation).tolist(),
        index=panels_df.index,
    )
    return panels_df


def load_panels_df(panels_path: Path) -> pd.DataFrame:
    normalized = load_pp_a_z_panels_df(panels_path)
    if normalized is not None:
        return normalized
    return workbook_patch.load_panels_df(panels_path)


def is_generated_or_support_file(path: Path) -> bool:
    stem = path.stem
    name_key = compact_name(path.name)
    if stem.endswith(WITH_PP_SUFFIX) or "_with_pp_" in stem or "_pp_filled" in stem:
        return True
    if stem.endswith("_unmatched"):
        return True
    if "jbpcombinedpp" in name_key or "ppinfo" in name_key or "cables" in name_key:
        return True
    return False


def workbook_has_supported_input_sheet(path: Path) -> bool:
    try:
        xls = pd.ExcelFile(path)
    except Exception:
        return False
    return any(sheet in ALL_EXCEL_SHEET_ENRICHERS for sheet in xls.sheet_names)


def is_qcli_full_report_source(path: Path, site_prefix: str = "") -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    if not path_matches_site_prefix(path, site_prefix):
        return False
    if not path.stem.endswith(FULL_REPORT_SUFFIX):
        return False
    if is_generated_or_support_file(path):
        return False
    return workbook_has_supported_input_sheet(path)


def is_excel_source(path: Path) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    if is_generated_or_support_file(path):
        return False
    return workbook_has_supported_input_sheet(path)


def is_default_csv_source(path: Path) -> bool:
    if path.suffix.lower() != ".csv" or is_generated_or_support_file(path):
        return False
    name = path.name
    compact = " ".join(name.split())
    is_rack_all_export = (
        "Rack All-data-" in name
        or re.search(r"Rack All \([^)]*greater than 0[^)]*\)-data-", name) is not None
    )
    if not is_rack_all_export:
        return False
    return (
        compact.startswith("Pre Fec Ber issues T1 ")
        or compact.startswith("T1 -_ T0 RX power")
        or compact.startswith("T1 -_ T0 TX power")
        or compact.startswith("Fec Bin issues T1 -_ T0 ")
    )


def compact_spaces(value: str) -> str:
    return " ".join(str(value).split())


def dashboard_csv_kind(path: Path) -> str | None:
    if path.suffix.lower() != ".csv" or is_generated_or_support_file(path):
        return None
    name = path.name
    compact = compact_spaces(name)
    is_rack_all_export = (
        "Rack All-data-" in name
        or re.search(r"Rack All \([^)]*greater than 0[^)]*\)-data-", name) is not None
    )
    if not is_rack_all_export:
        return None
    if compact.startswith("T1 -_ T0 RX power"):
        return "rx"
    if compact.startswith("T1 -_ T0 TX power"):
        return "tx"
    if compact.startswith("Pre Fec Ber issues T1 -_ T0 "):
        return "pre_fec"
    if compact.startswith("Fec Bin issues T1 -_ T0 "):
        return "fec_bin"
    return None


def dashboard_csv_timestamp(path: Path) -> datetime:
    match = DASHBOARD_DATA_TIMESTAMP_RE.search(path.name)
    if match:
        return datetime.strptime(
            f"{match.group(1)} {match.group(2)}:{match.group(3)}:{match.group(4)}",
            "%Y-%m-%d %H:%M:%S",
        )
    return datetime.fromtimestamp(path.stat().st_mtime)


def dashboard_csv_sources(batch_dir: Path) -> list[DashboardCsvSource]:
    sources: list[DashboardCsvSource] = []
    for path in sorted(batch_dir.iterdir()):
        if not path.is_file():
            continue
        kind = dashboard_csv_kind(path)
        if kind is None:
            continue
        sources.append(DashboardCsvSource(path.resolve(), kind, dashboard_csv_timestamp(path)))
    return sorted(sources, key=lambda source: (source.timestamp, DASHBOARD_KIND_ORDER[source.kind], source.path.name))


def cluster_dashboard_csv_sources(
    sources: list[DashboardCsvSource],
    window_minutes: int = DEFAULT_DASHBOARD_GROUP_WINDOW_MINUTES,
) -> list[DashboardCsvGroup]:
    groups: list[DashboardCsvGroup] = []
    window = timedelta(minutes=window_minutes)
    for source in sources:
        candidates: list[DashboardCsvGroup] = []
        for group in groups:
            if source.kind in group.sources:
                continue
            start = min(group.start, source.timestamp)
            end = max(group.end, source.timestamp)
            if end - start <= window:
                candidates.append(group)

        if candidates:
            target = min(
                candidates,
                key=lambda group: (
                    abs((source.timestamp - group.end).total_seconds()),
                    group.start,
                ),
            )
            target.sources[source.kind] = source
        else:
            groups.append(DashboardCsvGroup({source.kind: source}))

    return sorted(groups, key=lambda group: (group.start, group.end))


def existing_dashboard_outputs(batch_dir: Path, group: DashboardCsvGroup) -> list[Path]:
    timestamp = group.output_timestamp.strftime("%Y-%m-%d-%H%M")
    return sorted(batch_dir.glob(f"JBP*-DG*-t1-t0-rx-tx-pre-fec-fec-bin-{timestamp}-with-pp*.xlsx"))


def dashboard_group_needs_processing(
    batch_dir: Path,
    group: DashboardCsvGroup,
    force_refresh: bool = False,
) -> bool:
    if force_refresh:
        return True
    return not existing_dashboard_outputs(batch_dir, group)


def dashboard_group_matches_site_prefix(group: DashboardCsvGroup, site_prefix: str) -> bool:
    prefixes = set(site_prefix_values(site_prefix))
    if not prefixes:
        return True
    return dashboard_site_name(group).lower() in prefixes


def discover_dashboard_csv_groups(
    batch_dir: Path,
    force_refresh: bool = False,
    window_minutes: int = DEFAULT_DASHBOARD_GROUP_WINDOW_MINUTES,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> tuple[list[DashboardCsvGroup], list[DashboardCsvGroup]]:
    groups = cluster_dashboard_csv_sources(dashboard_csv_sources(batch_dir), window_minutes)
    groups = [group for group in groups if dashboard_group_matches_site_prefix(group, site_prefix)]
    pending = [group for group in groups if dashboard_group_needs_processing(batch_dir, group, force_refresh)]
    skipped = [group for group in groups if group not in pending]
    return pending, skipped


def dashboard_group_label(group: DashboardCsvGroup) -> str:
    parts = [
        f"{kind}:{source.path.name}"
        for kind, source in sorted(group.sources.items(), key=lambda item: DASHBOARD_KIND_ORDER[item[0]])
    ]
    return ", ".join(parts)


def workbook_has_pp_info(workbook_path: Path) -> bool:
    if not workbook_path.exists():
        return False
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        if any(sheet in wb.sheetnames for sheet in GENERATED_SHEETS):
            return True
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = {norm(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())}
            if {"Source_port", "DMARC1", "DMARC2", "Destination_port"}.issubset(headers):
                return True
            if {"PP_A", "PP_Z"}.issubset(headers):
                return True
    finally:
        wb.close()
    return False


def source_needs_processing(source_path: Path, force_refresh: bool = False) -> bool:
    if force_refresh:
        return True
    return not any(workbook_has_pp_info(path) for path in existing_timestamped_outputs(source_path))


def discover_batch_sources(
    batch_dir: Path,
    force_refresh: bool = False,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> tuple[list[Path], list[Path]]:
    sources = [
        path for path in sorted(batch_dir.iterdir())
        if path.is_file() and is_qcli_full_report_source(path, site_prefix)
    ]
    pending = [path for path in sources if source_needs_processing(path, force_refresh)]
    skipped = [path for path in sources if path not in pending]
    return pending, skipped


def pp_match_counts(df: pd.DataFrame) -> tuple[int, int, int]:
    rows = len(df)
    missing_values = {"", "none", "nan", workbook_patch.PATCH_NOT_FOUND.lower()}
    if all(column in df.columns for column in workbook_patch.PATCH_COLUMNS):
        matched_mask = df["Source_port"].map(norm).str.lower().map(lambda value: value not in missing_values)
    elif "PP_A" in df.columns:
        matched_mask = df["PP_A"].map(norm).str.lower().map(lambda value: value not in missing_values)
    else:
        return rows, 0, rows
    matches = int(matched_mask.sum())
    return rows, matches, rows - matches


def build_summary(source_path: Path, output_path: Path, panels_path: Path, generated: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, str | int]] = [
        {"Section": "Run Info", "Key": "Run Timestamp", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Section": "Run Info", "Key": "Input Report", "Value": str(source_path)},
        {"Section": "Run Info", "Key": "Output Report", "Value": str(output_path)},
        {"Section": "Run Info", "Key": "Panels File", "Value": str(panels_path)},
    ]
    for sheet_name, df in generated.items():
        total, matches, missing = pp_match_counts(df)
        rows.extend(
            [
                {"Section": "Generated", "Key": f"{sheet_name} rows", "Value": total},
                {"Section": "Generated", "Key": f"{sheet_name} PP matches", "Value": matches},
                {"Section": "Generated", "Key": f"{sheet_name} PP missing", "Value": missing},
            ]
        )
    return pd.DataFrame(rows)


def move_summary_to_end(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    if SUMMARY_SHEET in workbook.sheetnames:
        worksheet = workbook[SUMMARY_SHEET]
        workbook._sheets.remove(worksheet)
        workbook._sheets.append(worksheet)
    workbook.save(output_path)


def write_full_report(
    source_path: Path,
    output_path: Path,
    panels_path: Path,
    generated: dict[str, pd.DataFrame],
    original_workbook: pd.ExcelFile | None = None,
) -> None:
    exported_sheet_names = [sheet for sheet in REQUIRED_OUTPUT_SHEETS if sheet in generated]
    exported_sheet_names.extend(sheet for sheet in generated if sheet not in exported_sheet_names)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name in exported_sheet_names:
            generated[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    workbook_patch.style_workbook(output_path)


def process_excel(
    source_path: Path,
    panels_path: Path,
    panels_df: pd.DataFrame,
    output_path: Path,
    inventory: dict | None = None,
) -> Path:
    xls = pd.ExcelFile(source_path)
    a_view, b_view = workbook_patch.panel_lookup_views(panels_df)
    generated: dict[str, pd.DataFrame] = {}
    site_name = count_summary_site_name(source_path) or DEFAULT_SITE_PREFIX

    for input_sheet, (output_sheet, enricher) in EXCEL_SHEET_ENRICHERS.items():
        if input_sheet not in xls.sheet_names:
            continue
        source_df = pd.read_excel(source_path, sheet_name=input_sheet)
        if output_sheet == "optics_rx_tx_threshold" and inventory is not None:
            source_df = filter_site_optics_rows(
                source_df,
                inventory,
                site_name,
                report_side=count_summary_side(source_path),
            )
        generated[output_sheet] = enricher(source_df, a_view, b_view)

    for input_sheet, (output_sheet, enricher) in QCLI_FULL_REPORT_SHEET_ENRICHERS.items():
        if input_sheet not in xls.sheet_names:
            continue
        source_df = pd.read_excel(source_path, sheet_name=input_sheet)
        if output_sheet == "optics_rx_tx_threshold" and inventory is not None:
            source_df = filter_site_optics_rows(
                source_df,
                inventory,
                site_name,
                report_side=count_summary_side(source_path),
            )
        generated[output_sheet] = enricher(source_df, a_view, b_view)

    if not generated:
        raise ValueError(f"No supported report sheets found in {source_path}")

    write_full_report(source_path, output_path, panels_path, generated, original_workbook=xls)
    return output_path


def csv_output_sheet(csv_path: Path, enriched_df: pd.DataFrame) -> str:
    compact = " ".join(csv_path.name.split())
    if compact.startswith("Pre Fec Ber issues"):
        return "pre_fec_ber_threshold"
    if compact.startswith("Fec Bin issues"):
        return "fec_bin_threshold_sp"
    if "RX power" in compact or "TX power" in compact:
        return "optics_rx_tx_threshold"
    kind = enriched_df["_CSV_KIND"].iloc[0] if "_CSV_KIND" in enriched_df.columns and len(enriched_df) else ""
    return "optics_rx_tx_threshold" if kind == "optics" else "endpoint_threshold"


def process_csv(
    source_path: Path,
    panels_path: Path,
    panels_df: pd.DataFrame,
    output_path: Path,
    inventory: dict | None = None,
) -> Path:
    enriched_df = csv_patch.enrich_csv(source_path, panels_df)
    source_kind = dashboard_csv_kind(source_path)
    if source_kind in {"rx", "tx"} and inventory is not None:
        site_name = site_name_from_path_or_frame(source_path, enriched_df) or DEFAULT_SITE_PREFIX
        enriched_df = filter_site_optics_rows(enriched_df, inventory, site_name, source_kind)
    sheet_name = csv_output_sheet(source_path, enriched_df)
    write_full_report(source_path, output_path, panels_path, {sheet_name: enriched_df})
    return output_path


def find_dashboard_column(columns: Iterable[str], *wanted_names: str) -> str | None:
    wanted = {compact_name(name) for name in wanted_names}
    for column in columns:
        if compact_name(column) in wanted:
            return str(column)
    return None


def require_dashboard_column(columns: Iterable[str], *wanted_names: str) -> str:
    column = find_dashboard_column(columns, *wanted_names)
    if column is None:
        raise ValueError(f"Missing required column. Expected one of: {', '.join(wanted_names)}")
    return column


def split_dashboard_endpoint(value: str) -> tuple[str, str]:
    parts = norm(value).split()
    if len(parts) >= 2:
        return parts[0], parts[1]
    if parts:
        return parts[0], ""
    return "", ""


def dashboard_endpoint_key(value: str) -> str:
    hostname, interface = split_dashboard_endpoint(value)
    return workbook_patch.endpoint_key(hostname, interface)


def rack_number_from_text(value: str) -> str:
    text = norm(value)
    if re.fullmatch(r"\d{3,5}", text):
        return text
    match = re.search(r"(?:R|Rack\s+)(\d{3,5})", text, re.IGNORECASE)
    return match.group(1).zfill(4) if match else ""


def read_dashboard_csv(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    if df.empty:
        return df
    nonblank = df.apply(lambda row: any(norm(value) for value in row), axis=1)
    return df.loc[nonblank].reset_index(drop=True)


def dashboard_site_name(group: DashboardCsvGroup) -> str:
    for source in sorted(group.sources.values(), key=lambda item: (item.timestamp, DASHBOARD_KIND_ORDER[item.kind])):
        try:
            df = read_dashboard_csv(source.path)
        except Exception:
            continue
        device_col = find_dashboard_column(df.columns, "device", "Hostname")
        if device_col is None:
            continue
        for value in df[device_col].tolist():
            match = re.search(r"\b(jbp\d+)-", norm(value), re.IGNORECASE)
            if match:
                return match.group(1).upper()
    return "JBP"


def discover_dashboard_panels_path(
    batch_dir: Path,
    site_name: str,
    explicit_path: str | None = None,
) -> Path:
    if explicit_path:
        panels_path = Path(explicit_path).expanduser().resolve()
        if not panels_path.exists():
            raise FileNotFoundError(f"Patch panel workbook not found: {panels_path}")
        return panels_path

    if site_name.lower() == JBP19_SITE_PREFIX:
        preferred_names = [
            JBP19_QFABT1_IPR_PANELS_NAME,
            "JBP19_pp_info.xlsx",
            "jbp19_pp_info.xlsx",
        ]
    else:
        preferred_names = [
            f"{site_name}_pp_info.xlsx",
            f"{site_name.lower()}_pp_info.xlsx",
            f"{site_name}_combined_pp.xlsx",
            f"{site_name.lower()}_combined_pp.xlsx",
        ]
    for name in preferred_names:
        candidate = batch_dir / name
        if candidate.exists():
            return candidate.resolve()
    return discover_panels_path(batch_dir)


def load_dashboard_panel_lookup(
    panels_path: Path,
) -> tuple[dict[str, DashboardPanelRecord], dict[str, DashboardPanelRecord]]:
    xls = pd.ExcelFile(panels_path)
    sheet_name = "T0 PANELS" if "T0 PANELS" in xls.sheet_names else xls.sheet_names[0]
    raw = pd.read_excel(panels_path, sheet_name=sheet_name, dtype=str, keep_default_na=False)

    device_a_col = require_dashboard_column(raw.columns, "DEVICE_A", "DeviceA")
    rack_a_col = require_dashboard_column(raw.columns, "RackA", "Rack A")

    pp_a_col = find_dashboard_column(raw.columns, "PP_A")
    pp_z_col = find_dashboard_column(raw.columns, "PP_Z")
    device_z_col = find_dashboard_column(raw.columns, "DEVICE_Z")
    rack_z_col = find_dashboard_column(raw.columns, "RACK_Z")

    if pp_a_col and pp_z_col and device_z_col and rack_z_col:
        device_b_col = device_z_col
        rack_b_col = rack_z_col
    else:
        pp_a_col = require_dashboard_column(raw.columns, "Source_port")
        pp_z_col = require_dashboard_column(raw.columns, "Destination_port")
        device_b_col = require_dashboard_column(raw.columns, "DeviceB")
        rack_b_col = require_dashboard_column(raw.columns, "RackB")

    a_lookup: dict[str, DashboardPanelRecord] = {}
    z_lookup: dict[str, DashboardPanelRecord] = {}
    for _, row in raw.iterrows():
        record = DashboardPanelRecord(
            a_endpoint=norm(row[device_a_col]),
            z_endpoint=norm(row[device_b_col]),
            a_rack=rack_number_from_text(row[rack_a_col]),
            z_rack=rack_number_from_text(row[rack_b_col]),
            pp_a=norm(row[pp_a_col]),
            pp_z=norm(row[pp_z_col]),
        )
        a_key = dashboard_endpoint_key(record.a_endpoint)
        z_key = dashboard_endpoint_key(record.z_endpoint)
        if a_key:
            a_lookup.setdefault(a_key, record)
        if z_key:
            z_lookup.setdefault(z_key, record)
    return a_lookup, z_lookup


def dashboard_dg_by_qfabt0_rack(inventory: dict, site_name: str) -> dict[str, str]:
    site_key = site_name.lower()
    try:
        build = inventory["planar_builds"][site_key]
    except KeyError as exc:
        raise ValueError(f"Inventory JSON does not contain planar_builds.{site_key}") from exc

    pg_to_dg: dict[str, str] = {}
    for group in build.get("placement_groups", []):
        placement_group = norm(group.get("placement_group", ""))
        if not placement_group:
            continue
        pg_to_dg[placement_group] = norm(group.get("deployment_group_alias", "")) or f"DG{placement_group}"

    raw_map = build.get("design_rack_scope", {}).get("placement_group_by_qfabt0_rack", {})
    rack_to_dg: dict[str, str] = {}
    for rack, placement_group in raw_map.items():
        dg = pg_to_dg.get(str(placement_group), f"DG{placement_group}")
        rack_text = str(rack)
        rack_to_dg[rack_text] = dg
        try:
            rack_to_dg[str(int(rack_text))] = dg
            rack_to_dg[rack_text.zfill(4)] = dg
        except ValueError:
            pass
    return rack_to_dg


def inventory_build_for_site(inventory: dict, site_name: str) -> dict:
    site_key = site_name.lower()
    try:
        return inventory["planar_builds"][site_key]
    except KeyError as exc:
        raise ValueError(f"Inventory JSON does not contain planar_builds.{site_key}") from exc


def normalize_phase_scope_alias(value: str) -> str:
    text = norm(value).upper().replace(" ", "")
    if re.fullmatch(r"\d+", text):
        return f"DG{text}"
    return text


def phase1_scope_config_for_site(inventory: dict, site_name: str) -> dict:
    if site_name.lower() != DEFAULT_SITE_PREFIX:
        return {}
    build = inventory_build_for_site(inventory, site_name)
    phase_scopes = build.get("phase_scopes", {})
    for key in ("Phase 1", "phase_1", "phase1", "1"):
        value = phase_scopes.get(key)
        if isinstance(value, dict):
            return value
    return {}


def phase1_dg_scope_for_site(inventory: dict, site_name: str) -> dict[str, dict[str, str]]:
    if site_name.lower() != DEFAULT_SITE_PREFIX:
        return {}

    phase_config = phase1_scope_config_for_site(inventory, site_name)
    configured_groups = phase_config.get("deployment_groups", {})
    if isinstance(configured_groups, dict) and configured_groups:
        result = {}
        for dg, metadata in configured_groups.items():
            dg_alias = normalize_phase_scope_alias(dg)
            result[dg_alias] = metadata if isinstance(metadata, dict) else {}
        return result
    if isinstance(configured_groups, list) and configured_groups:
        return {normalize_phase_scope_alias(dg): {} for dg in configured_groups}

    build = inventory_build_for_site(inventory, site_name)
    from_inventory: dict[str, dict[str, str]] = {}
    for group in build.get("placement_groups", []):
        phase = compact_name(group.get("phase", "") or group.get("customer_handoff_phase", ""))
        if phase not in {"phase1", "1"}:
            continue
        dg = norm(group.get("deployment_group_alias", "")) or f"DG{norm(group.get('placement_group', ''))}"
        dg = normalize_phase_scope_alias(dg)
        if not dg:
            continue
        from_inventory[dg] = {
            "milestone": norm(group.get("milestone", "")),
            "data_hall": norm(group.get("data_hall", "")),
        }

    return from_inventory or PHASE1_DEFAULT_DG_SCOPE


def phase1_special_scopes_for_site(inventory: dict, site_name: str) -> set[str]:
    phase_config = phase1_scope_config_for_site(inventory, site_name)
    configured_special = phase_config.get("special_scopes", {})
    if isinstance(configured_special, dict) and configured_special:
        return {normalize_phase_scope_alias(scope) for scope in configured_special}
    if isinstance(configured_special, list) and configured_special:
        return {normalize_phase_scope_alias(scope) for scope in configured_special}
    return set(PHASE1_SPECIAL_OPTICS_SCOPES) if site_name.lower() == DEFAULT_SITE_PREFIX else set()


def add_rack_scope_alias(rack_to_scope: dict[str, str], rack: str, scope_alias: str) -> None:
    rack_text = rack_number_from_text(rack) or norm(rack)
    if not rack_text:
        return
    rack_to_scope[rack_text] = scope_alias
    try:
        rack_to_scope[str(int(rack_text))] = scope_alias
        rack_to_scope[rack_text.zfill(4)] = scope_alias
    except ValueError:
        pass


def phase1_scope_by_network_rack(inventory: dict, site_name: str) -> dict[str, str]:
    if site_name.lower() != DEFAULT_SITE_PREFIX:
        return {}

    build = inventory_build_for_site(inventory, site_name)
    phase1_dgs = set(phase1_dg_scope_for_site(inventory, site_name))
    rack_to_scope: dict[str, str] = {}

    for group in build.get("placement_groups", []):
        dg = norm(group.get("deployment_group_alias", "")) or f"DG{norm(group.get('placement_group', ''))}"
        dg = normalize_phase_scope_alias(dg)
        if dg not in phase1_dgs:
            continue
        for rack in group.get("qfabt0_racks", []):
            add_rack_scope_alias(rack_to_scope, rack, dg)

    for group in build.get("t1_spine_groups", []):
        for rack in group.get("rack_positions", []):
            add_rack_scope_alias(rack_to_scope, rack, "SPINE")

    for rack in build.get("ipr_rack", {}).get("rack_positions", []):
        add_rack_scope_alias(rack_to_scope, rack, "IPR")

    return rack_to_scope


def phase1_scope_aliases(inventory: dict, site_name: str) -> set[str]:
    if site_name.lower() != DEFAULT_SITE_PREFIX:
        return set()
    return set(phase1_dg_scope_for_site(inventory, site_name)) | phase1_special_scopes_for_site(inventory, site_name)


def optics_power_kind_from_row(row, source_kind: str | None = None) -> str | None:
    if source_kind in {"rx", "tx"}:
        return source_kind

    power_type = compact_name(row.get("Power Type", "")) if hasattr(row, "get") else ""
    if power_type == "rx":
        return "rx"
    if power_type == "tx":
        return "tx"

    metric = compact_name(row.get("Metric", "")) if hasattr(row, "get") else ""
    if "output" in metric or metric == "tx" or "txpower" in metric:
        return "tx"
    if "input" in metric or metric == "rx" or "rxpower" in metric:
        return "rx"

    return None


def optics_measured_dbm_from_row(row) -> float | None:
    for column in (
        "Measured (dBm)",
        "Value",
        "RX power",
        "TX power",
        "Input Power",
        "Output Power",
        "Measured",
        "Power",
    ):
        if not hasattr(row, "get") or column not in row:
            continue
        measured = numeric_value(row.get(column, ""))
        if measured is not None:
            return measured
    return None


def phase_scope_for_row(row, rack_to_phase_scope: dict[str, str]) -> str:
    if hasattr(row, "get"):
        dg = normalize_phase_scope_alias(row.get("DG", ""))
        if dg:
            return dg
        for column in ("Rack", "rack"):
            rack = rack_number_from_text(row.get(column, ""))
            if rack:
                return normalize_phase_scope_alias(rack_to_phase_scope.get(rack, ""))
    return ""


def optics_row_outside_threshold_range(
    row,
    source_kind: str | None,
    rx_min: float,
    tx_min: float,
    max_value: float,
) -> bool:
    power_kind = optics_power_kind_from_row(row, source_kind)
    if power_kind is None:
        return True

    measured = optics_measured_dbm_from_row(row)
    if measured is None:
        return True

    if power_kind == "rx":
        return measured < rx_min or measured > max_value
    if power_kind == "tx":
        return measured < tx_min or measured > max_value
    return True


def optics_row_passes_phase1_threshold(row, scope_alias: str, source_kind: str | None = None) -> bool:
    del scope_alias
    rx_min = NON_IPR_RELAX_RX_MIN_DBM if RELAX_NON_IPR_OPTICS else PHASE1_RX_INPUT_WORSE_THAN_DBM
    tx_min = NON_IPR_RELAX_TX_MIN_DBM if RELAX_NON_IPR_OPTICS else PHASE1_TX_OUTPUT_WORSE_THAN_DBM
    max_value = NON_IPR_RELAX_MAX_DBM if RELAX_NON_IPR_OPTICS else NON_IPR_MAX_DBM
    return optics_row_outside_threshold_range(row, source_kind, rx_min, tx_min, max_value)


def optics_row_passes_ipr_tx_threshold(row, source_kind: str | None = None) -> bool:
    rx_min = IPR_RELAX_RX_MIN_DBM if RELAX_IPR_OPTICS else IPR_RX_MIN_DBM
    tx_min = IPR_RELAX_TX_MIN_DBM if RELAX_IPR_OPTICS else IPR_TX_MIN_DBM
    max_value = IPR_RELAX_MAX_DBM if RELAX_IPR_OPTICS else IPR_MAX_DBM
    return optics_row_outside_threshold_range(row, source_kind, rx_min, tx_min, max_value)


def optics_row_passes_jbp19_threshold(row, source_kind: str | None = None) -> bool:
    rx_min = NON_IPR_RELAX_RX_MIN_DBM if RELAX_NON_IPR_OPTICS else PHASE1_RX_INPUT_WORSE_THAN_DBM
    tx_min = NON_IPR_RELAX_TX_MIN_DBM if RELAX_NON_IPR_OPTICS else PHASE1_TX_OUTPUT_WORSE_THAN_DBM
    max_value = NON_IPR_RELAX_MAX_DBM if RELAX_NON_IPR_OPTICS else NON_IPR_MAX_DBM
    return optics_row_outside_threshold_range(row, source_kind, rx_min, tx_min, max_value)


def filter_phase1_optics_rows(
    df: pd.DataFrame,
    inventory: dict,
    site_name: str,
    source_kind: str | None = None,
) -> pd.DataFrame:
    if df.empty or site_name.lower() != DEFAULT_SITE_PREFIX:
        return df

    rack_to_phase_scope = phase1_scope_by_network_rack(inventory, site_name)
    phase1_aliases = phase1_scope_aliases(inventory, site_name)

    keep_indexes: list[bool] = []
    for _, row in df.iterrows():
        scope_alias = phase_scope_for_row(row, rack_to_phase_scope)
        if scope_alias not in phase1_aliases:
            keep_indexes.append(True)
            continue
        keep_indexes.append(optics_row_passes_phase1_threshold(row, scope_alias, source_kind))

    return df.loc[keep_indexes].reset_index(drop=True)


def filter_site_optics_rows(
    df: pd.DataFrame,
    inventory: dict,
    site_name: str,
    source_kind: str | None = None,
    report_side: str | None = None,
) -> pd.DataFrame:
    if df.empty:
        return df

    if report_side == "ipr":
        keep_indexes = [
            optics_row_passes_ipr_tx_threshold(row, source_kind)
            for _, row in df.iterrows()
        ]
        return df.loc[keep_indexes].reset_index(drop=True)

    if site_name.lower() == JBP19_SITE_PREFIX:
        keep_indexes = [
            optics_row_passes_jbp19_threshold(row, source_kind)
            for _, row in df.iterrows()
        ]
        return df.loc[keep_indexes].reset_index(drop=True)

    return filter_phase1_optics_rows(df, inventory, site_name, source_kind)


def dashboard_interface_candidates(transceiver: str, channel: str) -> list[str]:
    transceiver_match = re.search(r"(\d+)", norm(transceiver))
    if not transceiver_match:
        return []
    try:
        channel_num = int(float(norm(channel)))
    except Exception:
        return []

    base = f"swp{transceiver_match.group(1)}"
    candidates: list[str] = []

    if 1 <= channel_num <= 4:
        candidates.append(f"{base}s0")
    elif 5 <= channel_num <= 8:
        candidates.append(f"{base}s1")

    if 1 <= channel_num <= 2:
        candidates.append(f"{base}s0")
    elif 3 <= channel_num <= 4:
        candidates.append(f"{base}s1")
    elif 5 <= channel_num <= 6:
        candidates.append(f"{base}s2")
    elif 7 <= channel_num <= 8:
        candidates.append(f"{base}s3")

    return ordered_unique(candidates)


def resolve_dashboard_interface(
    device: str,
    transceiver: str,
    channel: str,
    a_lookup: dict[str, DashboardPanelRecord],
    z_lookup: dict[str, DashboardPanelRecord],
) -> str:
    candidates = dashboard_interface_candidates(transceiver, channel)
    for interface in candidates:
        key = workbook_patch.endpoint_key(device, interface)
        if key in a_lookup or key in z_lookup:
            return interface
    return candidates[0] if candidates else ""


def dashboard_panel_enrichment(
    device: str,
    interface: str,
    a_lookup: dict[str, DashboardPanelRecord],
    z_lookup: dict[str, DashboardPanelRecord],
    dg_by_qfabt0_rack: dict[str, str],
) -> dict[str, str]:
    key = workbook_patch.endpoint_key(device, interface)
    matched_endpoint = f"{norm(device)} {norm(interface)}".strip()
    if not key:
        return {
            "DG": "Unknown",
            "Local PP": "",
            "Peer PP": "",
            "Peer Endpoint": "",
            "Peer Rack": "",
            "PP_A": "",
            "PP_Z": "",
            "Matched Endpoint": matched_endpoint,
        }

    if key in z_lookup:
        record = z_lookup[key]
        local_pp = record.pp_z
        peer_pp = record.pp_a
        peer_endpoint = record.a_endpoint
        peer_rack = record.a_rack
    elif key in a_lookup:
        record = a_lookup[key]
        local_pp = record.pp_a
        peer_pp = record.pp_z
        peer_endpoint = record.z_endpoint
        peer_rack = record.z_rack
    else:
        return {
            "DG": "Unknown",
            "Local PP": workbook_patch.PATCH_NOT_FOUND,
            "Peer PP": workbook_patch.PATCH_NOT_FOUND,
            "Peer Endpoint": workbook_patch.PATCH_NOT_FOUND,
            "Peer Rack": "",
            "PP_A": workbook_patch.PATCH_NOT_FOUND,
            "PP_Z": workbook_patch.PATCH_NOT_FOUND,
            "Matched Endpoint": matched_endpoint,
        }

    dg = dg_by_qfabt0_rack.get(peer_rack) or dg_by_qfabt0_rack.get(str(int(peer_rack)) if peer_rack else "")
    return {
        "DG": dg or "Unknown",
        "Local PP": local_pp,
        "Peer PP": peer_pp,
        "Peer Endpoint": peer_endpoint,
        "Peer Rack": peer_rack,
        "PP_A": record.pp_a,
        "PP_Z": record.pp_z,
        "Matched Endpoint": matched_endpoint,
    }


def dashboard_empty_report() -> pd.DataFrame:
    return pd.DataFrame(columns=DASHBOARD_REPORT_COLUMNS)


def enrich_dashboard_optics_csv(
    source: DashboardCsvSource,
    a_lookup: dict[str, DashboardPanelRecord],
    z_lookup: dict[str, DashboardPanelRecord],
    dg_by_qfabt0_rack: dict[str, str],
) -> pd.DataFrame:
    source_df = read_dashboard_csv(source.path)
    power_type = "RX" if source.kind == "rx" else "TX"
    value_column = f"{power_type} power"
    rows: list[dict[str, str]] = []
    for _, row in source_df.iterrows():
        device = norm(row.get("device", ""))
        interface = resolve_dashboard_interface(device, row.get("name", ""), row.get("channel", ""), a_lookup, z_lookup)
        output = {
            "Time": norm(row.get("Time", "")),
            "Issue Type": "Optics",
            "Power Type": power_type,
            "fabric_instance": norm(row.get("fabric_instance", "")),
            "device": device,
            "rack": norm(row.get("rack", "")),
            "elevation": norm(row.get("elevation", "")),
            "name": norm(row.get("name", "")),
            "channel": norm(row.get("channel", "")),
            "interface": interface,
            "serial": norm(row.get("serial", "")),
            "Value": norm(row.get(value_column, row.get("Value", ""))),
            "remote_device": "",
            "remote_interface": "",
            "Source File": compact_spaces(source.path.name),
        }
        output.update(dashboard_panel_enrichment(device, interface, a_lookup, z_lookup, dg_by_qfabt0_rack))
        rows.append(output)
    return pd.DataFrame(rows, columns=DASHBOARD_REPORT_COLUMNS)


def enrich_dashboard_endpoint_csv(
    source: DashboardCsvSource,
    issue_type: str,
    a_lookup: dict[str, DashboardPanelRecord],
    z_lookup: dict[str, DashboardPanelRecord],
    dg_by_qfabt0_rack: dict[str, str],
) -> pd.DataFrame:
    source_df = read_dashboard_csv(source.path)
    rows: list[dict[str, str]] = []
    for _, row in source_df.iterrows():
        device = norm(row.get("device", ""))
        interface = norm(row.get("interface", ""))
        output = {
            "Time": norm(row.get("Time", "")),
            "Issue Type": issue_type,
            "Power Type": "",
            "fabric_instance": norm(row.get("fabric_instance", "")),
            "device": device,
            "rack": norm(row.get("rack", "")),
            "elevation": norm(row.get("elevation", "")),
            "name": norm(row.get("name", "")),
            "channel": norm(row.get("channel", "")),
            "interface": interface,
            "serial": norm(row.get("serial", "")),
            "Value": norm(row.get("Value", "")),
            "remote_device": norm(row.get("remote_device", "")),
            "remote_interface": norm(row.get("remote_interface", "")),
            "Source File": compact_spaces(source.path.name),
        }
        output.update(dashboard_panel_enrichment(device, interface, a_lookup, z_lookup, dg_by_qfabt0_rack))
        rows.append(output)
    return pd.DataFrame(rows, columns=DASHBOARD_REPORT_COLUMNS)


def dashboard_dg_sort_key(value: str) -> tuple[int, str]:
    match = re.search(r"\d+", norm(value))
    if match:
        return int(match.group(0)), norm(value)
    return 9999, norm(value)


def build_dashboard_summary(all_df: pd.DataFrame) -> pd.DataFrame:
    if all_df.empty:
        return pd.DataFrame(columns=DASHBOARD_SUMMARY_COLUMNS)
    dgs = sorted(
        [dg for dg in all_df["DG"].map(norm).unique() if dg],
        key=dashboard_dg_sort_key,
    )
    rows: list[dict[str, int | str]] = []
    for dg in dgs:
        dg_df = all_df[all_df["DG"].map(norm).eq(dg)]
        rx = int(((dg_df["Issue Type"] == "Optics") & (dg_df["Power Type"] == "RX")).sum())
        tx = int(((dg_df["Issue Type"] == "Optics") & (dg_df["Power Type"] == "TX")).sum())
        pre_fec = int((dg_df["Issue Type"] == "Pre-FEC").sum())
        fec_bin = int((dg_df["Issue Type"] == "FEC Bin").sum())
        rows.append(
            {
                "DG": dg,
                "RX Optics": rx,
                "TX Optics": tx,
                "Total Optics": rx + tx,
                "Pre-FEC": pre_fec,
                "FEC Bin": fec_bin,
                "Grand Total": rx + tx + pre_fec + fec_bin,
            }
        )

    if rows:
        rows.append({"DG": "Total", **{column: sum(int(row[column]) for row in rows) for column in DASHBOARD_SUMMARY_COLUMNS[1:]}})
    return pd.DataFrame(rows, columns=DASHBOARD_SUMMARY_COLUMNS)


def dashboard_dg_part(summary_df: pd.DataFrame) -> str:
    dgs = [dg for dg in summary_df.get("DG", pd.Series(dtype=str)).map(norm).tolist() if dg and dg != "Total"]
    if not dgs:
        return "DG-unknown"
    parts = [dgs[0]]
    parts.extend(re.sub(r"^DG", "", dg, flags=re.IGNORECASE) for dg in dgs[1:])
    return "-".join(parts)


def next_available_path(path: Path) -> Path:
    if not path.exists():
        return path
    index = 1
    while True:
        candidate = path.with_name(f"{path.stem} ({index}){path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def dashboard_output_path(batch_dir: Path, site_name: str, summary_df: pd.DataFrame, group: DashboardCsvGroup) -> Path:
    timestamp = group.output_timestamp.strftime("%Y-%m-%d-%H%M")
    return next_available_path(
        batch_dir / f"{site_name}-{dashboard_dg_part(summary_df)}-t1-t0-rx-tx-pre-fec-fec-bin-{timestamp}-with-pp.xlsx"
    )


def style_dashboard_workbook(output_path: Path) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(output_path)
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    pp_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    dg_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        header_map = {cell.value: cell.column for cell in worksheet[1]}
        for column_name in ["Local PP", "Peer PP", "Peer Endpoint", "PP_A", "PP_Z", "Matched Endpoint"]:
            column_index = header_map.get(column_name)
            if column_index is None:
                continue
            for row_index in range(1, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).fill = pp_fill

        dg_column = header_map.get("DG")
        if dg_column is not None:
            for row_index in range(1, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=dg_column).fill = dg_fill

        for column_index in range(1, worksheet.max_column + 1):
            values = [
                norm(worksheet.cell(row=row_index, column=column_index).value)
                for row_index in range(1, min(worksheet.max_row, 100) + 1)
            ]
            width = min(max(max((len(value) for value in values), default=0) + 2, 10), 42)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
    workbook.save(output_path)


def build_dashboard_report(
    group: DashboardCsvGroup,
    panels_path: Path,
    inventory: dict,
) -> tuple[str, dict[str, pd.DataFrame]]:
    site_name = dashboard_site_name(group)
    a_lookup, z_lookup = load_dashboard_panel_lookup(panels_path)
    dg_by_qfabt0_rack = dashboard_dg_by_qfabt0_rack(inventory, site_name)

    rx_df = (
        enrich_dashboard_optics_csv(group.sources["rx"], a_lookup, z_lookup, dg_by_qfabt0_rack)
        if "rx" in group.sources
        else dashboard_empty_report()
    )
    tx_df = (
        enrich_dashboard_optics_csv(group.sources["tx"], a_lookup, z_lookup, dg_by_qfabt0_rack)
        if "tx" in group.sources
        else dashboard_empty_report()
    )
    rx_df = filter_site_optics_rows(rx_df, inventory, site_name, "rx")
    tx_df = filter_site_optics_rows(tx_df, inventory, site_name, "tx")
    optics_df = pd.concat([rx_df, tx_df], ignore_index=True)
    pre_fec_df = (
        enrich_dashboard_endpoint_csv(group.sources["pre_fec"], "Pre-FEC", a_lookup, z_lookup, dg_by_qfabt0_rack)
        if "pre_fec" in group.sources
        else dashboard_empty_report()
    )
    fec_bin_df = (
        enrich_dashboard_endpoint_csv(group.sources["fec_bin"], "FEC Bin", a_lookup, z_lookup, dg_by_qfabt0_rack)
        if "fec_bin" in group.sources
        else dashboard_empty_report()
    )
    all_df = pd.concat([optics_df, pre_fec_df, fec_bin_df], ignore_index=True)
    summary_df = build_dashboard_summary(all_df)

    generated: dict[str, pd.DataFrame] = {
        "Summary": summary_df,
        "All": all_df,
        "Optics": optics_df,
        "Pre-FEC": pre_fec_df,
        "FEC Bin": fec_bin_df,
    }
    for dg in [dg for dg in summary_df.get("DG", pd.Series(dtype=str)).map(norm).tolist() if dg and dg != "Total"]:
        sheet_name = dg[:31]
        generated[sheet_name] = all_df[all_df["DG"].map(norm).eq(dg)]
    return site_name, generated


def write_dashboard_report(output_path: Path, generated: dict[str, pd.DataFrame]) -> Path:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in generated.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    style_dashboard_workbook(output_path)
    return output_path


def is_count_summary_workbook(path: Path, site_prefix: str = DEFAULT_SITE_PREFIX) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    name = path.name
    name_key = compact_name(name)
    lower_name = name.lower()
    if not path_matches_site_prefix(path, site_prefix):
        return False
    if is_generated_or_support_file(path) and "_with_pp" not in path.stem:
        return False
    if "filtered" in name_key:
        return False
    if "t1t0rxtxprefecfecbin" in name_key:
        return False
    if "-q2-p-" not in lower_name and "-q2-ip-" not in lower_name:
        return False
    return "_with_pp" in lower_name or lower_name.endswith("_full_report.xlsx")


def count_summary_side(path: Path) -> str | None:
    lower_name = path.name.lower()
    if "-q2-ip-" in lower_name:
        return "ipr"
    if "-q2-p-" in lower_name:
        return "t0"
    return None


def count_summary_site_name(path: Path) -> str | None:
    return filename_site_name(path)


def site_name_from_path_or_frame(path: Path, df: pd.DataFrame) -> str | None:
    site_name = count_summary_site_name(path)
    if site_name:
        return site_name

    for column in ("Hostname", "device"):
        if column not in df.columns:
            continue
        for value in df[column].map(norm):
            match = re.search(r"\b(jbp\d+)-", value, re.IGNORECASE)
            if match:
                return match.group(1).lower()
    return None


def count_summary_sheet_name(workbook, count_key: str) -> str | None:
    for sheet_name in COUNT_SUMMARY_SHEETS[count_key]:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    return None


def header_index(headers: list[str], *wanted_names: str) -> int | None:
    wanted = {compact_name(name) for name in wanted_names}
    for index, header in enumerate(headers):
        if compact_name(header) in wanted:
            return index
    return None


def dg_alias_sort_key(value: str) -> tuple[int, str]:
    match = re.search(r"\d+", norm(value))
    if match:
        return int(match.group(0)), norm(value)
    return 9999, norm(value)


def zero_count_summary() -> dict[str, int]:
    return {key: 0 for key in COUNT_SUMMARY_KEYS}


def add_count_summary_rows(
    counts: dict[str, dict[str, int]],
    count_key: str,
    path: Path,
    sheet_name: str,
    side: str,
    rack_to_dg: dict[str, str],
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [norm(value) for value in next(rows)]
        except StopIteration:
            return

        rack_index = header_index(headers, "Rack")
        if rack_index is None:
            return

        for row in rows:
            if not any(norm(value) for value in row):
                continue
            rack = rack_number_from_text(row[rack_index] if rack_index < len(row) else "")
            if side == "ipr":
                dg = "IPR"
            else:
                dg = rack_to_dg.get(rack) or rack_to_dg.get(str(int(rack)) if rack else "")
            if not dg:
                continue
            counts.setdefault(dg, zero_count_summary())[count_key] += 1
    finally:
        workbook.close()


def count_summary_for_workbook(path: Path, inventory: dict) -> tuple[str, dict[str, dict[str, int]]] | None:
    side = count_summary_side(path)
    site_name = count_summary_site_name(path)
    if side is None or site_name is None:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        sheet_names = {
            count_key: count_summary_sheet_name(workbook, count_key)
            for count_key in COUNT_SUMMARY_KEYS
        }
    finally:
        workbook.close()

    if not any(sheet_names.values()):
        return None

    rack_to_dg = dashboard_dg_by_qfabt0_rack(inventory, site_name)
    counts: dict[str, dict[str, int]] = {}
    for count_key, sheet_name in sheet_names.items():
        if sheet_name is None:
            continue
        add_count_summary_rows(counts, count_key, path, sheet_name, side, rack_to_dg)
    return side, counts


def latest_count_summaries(
    batch_dir: Path,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> dict[str, dict[str, dict[str, int]]]:
    inventory = load_planar_inventory(inventory_path)
    latest: dict[str, dict[str, tuple[float, dict[str, int]]]] = {"t0": {}, "ipr": {}}

    for path in sorted(batch_dir.iterdir(), key=lambda item: item.stat().st_mtime):
        if not path.is_file() or not is_count_summary_workbook(path, site_prefix):
            continue
        summary = count_summary_for_workbook(path, inventory)
        if summary is None:
            continue
        side, counts_by_dg = summary
        mtime = path.stat().st_mtime
        for dg, counts in counts_by_dg.items():
            current = latest.setdefault(side, {}).get(dg)
            if current is None or mtime >= current[0]:
                latest[side][dg] = (mtime, counts)

    return {
        side: {dg: counts for dg, (_mtime, counts) in summaries.items()}
        for side, summaries in latest.items()
    }


def count_summary_capture_timestamp(path: Path) -> datetime:
    match = re.search(r"jbp\d+-q2-(?:p|ip)-(\d{4}-\d{2}-\d{2})_(\d{6})", path.name, re.IGNORECASE)
    if match:
        return datetime.strptime(f"{match.group(1)}_{match.group(2)}", "%Y-%m-%d_%H%M%S")
    return datetime.fromtimestamp(path.stat().st_mtime)


def detailed_count_summary_rank(path: Path) -> tuple[datetime, int, float]:
    name = path.name.lower()
    if name.endswith("_full_report.xlsx"):
        detail_priority = 3
    elif "_with_pp" not in name:
        detail_priority = 2
    else:
        detail_priority = 1
    return count_summary_capture_timestamp(path), detail_priority, path.stat().st_mtime


def zero_detailed_count_summary() -> dict[str, int | set[tuple[str, str]]]:
    return {
        "lldp": 0,
        "interface": 0,
        "optics_tx": 0,
        "optics_rx": 0,
        "optics_temp": 0,
        "fec_bin": 0,
        "pre_fec": 0,
        "_endpoints": set(),
    }


def numeric_value(value) -> float | None:
    text = norm(value)
    if not text:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?(?:e[-+]?\d+)?", text, re.IGNORECASE)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def detailed_summary_count_key(count_key: str, headers: list[str], row: tuple) -> str | None:
    if count_key == "optics":
        metric_index = header_index(headers, "Metric")
        metric = compact_name(row[metric_index]) if metric_index is not None and metric_index < len(row) else ""
        if "output" in metric or metric == "tx" or "txpower" in metric:
            return "optics_tx"
        return "optics_rx"
    if count_key == "combined_fec":
        lock_index = header_index(headers, "Lock Status")
        pre_fec_index = header_index(headers, "Pre-FEC BER", "Pre‑FEC BER")
        lock_status = compact_name(row[lock_index]) if lock_index is not None and lock_index < len(row) else ""
        if "fecbin" in lock_status:
            return "fec_bin"
        pre_fec = numeric_value(row[pre_fec_index]) if pre_fec_index is not None and pre_fec_index < len(row) else None
        if pre_fec is not None and pre_fec > 1e-7:
            return "pre_fec"
        if "rawber" in lock_status:
            return "pre_fec"
        return None
    return count_key


def add_detailed_count_summary_rows(
    counts: dict[str, dict[str, int | set[tuple[str, str]]]],
    count_key: str,
    path: Path,
    sheet_name: str,
    site_name: str,
    side: str,
    rack_to_dg: dict[str, str],
    rack_to_phase_scope: dict[str, str],
    phase1_aliases: set[str],
) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [norm(value) for value in next(rows)]
        except StopIteration:
            return

        rack_index = header_index(headers, "Rack")
        hostname_index = header_index(headers, "Hostname", "device")
        interface_index = header_index(headers, "Interface", "Interface (not up / not enabled)")
        if rack_index is None:
            return

        for row in rows:
            if not any(norm(value) for value in row):
                continue
            rack = rack_number_from_text(row[rack_index] if rack_index < len(row) else "")
            if side == "ipr":
                dg = "IPR"
            else:
                dg = rack_to_dg.get(rack) or rack_to_dg.get(str(int(rack)) if rack else "")
            if not dg:
                continue

            row_map = {
                headers[index]: row[index] if index < len(row) else ""
                for index in range(len(headers))
            }
            scope_alias = (
                "IPR"
                if side == "ipr"
                else normalize_phase_scope_alias(rack_to_phase_scope.get(rack, "") or dg)
            )
            if count_key == "optics":
                if side == "ipr":
                    if not optics_row_passes_ipr_tx_threshold(row_map):
                        continue
                elif site_name.lower() == JBP19_SITE_PREFIX:
                    if not optics_row_passes_jbp19_threshold(row_map):
                        continue
                elif scope_alias in phase1_aliases and not optics_row_passes_phase1_threshold(row_map, scope_alias):
                    continue

            target_key = detailed_summary_count_key(count_key, headers, row)
            if target_key is None:
                continue

            dg_counts = counts.setdefault(dg, zero_detailed_count_summary())
            dg_counts[target_key] = int(dg_counts.get(target_key, 0)) + 1
            if hostname_index is not None and interface_index is not None:
                hostname = norm(row[hostname_index] if hostname_index < len(row) else "")
                interface = norm(row[interface_index] if interface_index < len(row) else "")
                if hostname or interface:
                    endpoints = dg_counts["_endpoints"]
                    assert isinstance(endpoints, set)
                    endpoints.add((hostname, interface))
    finally:
        workbook.close()


def detailed_count_summary_for_workbook(
    path: Path,
    inventory: dict,
) -> tuple[str, dict[str, dict[str, int | set[tuple[str, str]]]]] | None:
    side = count_summary_side(path)
    site_name = count_summary_site_name(path)
    if side is None or site_name is None:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    try:
        sheet_names = {
            count_key: count_summary_sheet_name(workbook, count_key)
            for count_key in COUNT_SUMMARY_KEYS
        }
    finally:
        workbook.close()

    if not any(sheet_names.values()):
        return None

    rack_to_dg = dashboard_dg_by_qfabt0_rack(inventory, site_name)
    rack_to_phase_scope = phase1_scope_by_network_rack(inventory, site_name)
    phase1_aliases = phase1_scope_aliases(inventory, site_name)
    counts: dict[str, dict[str, int | set[tuple[str, str]]]] = {}
    for count_key, sheet_name in sheet_names.items():
        if sheet_name is None:
            continue
        add_detailed_count_summary_rows(
            counts,
            count_key,
            path,
            sheet_name,
            site_name,
            side,
            rack_to_dg,
            rack_to_phase_scope,
            phase1_aliases,
        )
    return side, counts


def latest_detailed_count_summaries(
    batch_dir: Path,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> dict[str, dict[str, dict[str, int]]]:
    inventory = load_planar_inventory(inventory_path)
    latest: dict[str, dict[str, tuple[tuple[datetime, int, float], dict[str, int | set[tuple[str, str]]]]]] = {
        "t0": {},
        "ipr": {},
    }

    for path in sorted(batch_dir.iterdir(), key=lambda item: item.stat().st_mtime):
        if not path.is_file() or not is_count_summary_workbook(path, site_prefix):
            continue
        summary = detailed_count_summary_for_workbook(path, inventory)
        if summary is None:
            continue
        side, counts_by_dg = summary
        rank = detailed_count_summary_rank(path)
        for dg, counts in counts_by_dg.items():
            current = latest.setdefault(side, {}).get(dg)
            if current is None or rank >= current[0]:
                latest[side][dg] = (rank, counts)

    result: dict[str, dict[str, dict[str, int]]] = {"t0": {}, "ipr": {}}
    for side, summaries in latest.items():
        for dg, (_rank, counts) in summaries.items():
            endpoints = counts.get("_endpoints", set())
            result[side][dg] = {
                "lldp": int(counts.get("lldp", 0)),
                "interface": int(counts.get("interface", 0)),
                "optics_tx": int(counts.get("optics_tx", 0)),
                "optics_rx": int(counts.get("optics_rx", 0)),
                "optics_temp": int(counts.get("optics_temp", 0)),
                "fec_bin": int(counts.get("fec_bin", 0)),
                "pre_fec": int(counts.get("pre_fec", 0)),
                "present_total": len(endpoints) if isinstance(endpoints, set) else 0,
            }
    return result


def count_summary_site_prefixes(batch_dir: Path, site_prefix: str) -> list[str]:
    prefixes = site_prefix_values(site_prefix)
    if prefixes:
        return prefixes

    discovered: set[str] = set()
    for path in batch_dir.iterdir():
        if not path.is_file() or not is_count_summary_workbook(path, ""):
            continue
        site_name = count_summary_site_name(path)
        if site_name:
            discovered.add(site_name)
    return sorted(discovered)


def latest_detailed_count_summaries_by_site(
    batch_dir: Path,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> list[tuple[str, dict[str, dict[str, dict[str, int]]]]]:
    prefixes = count_summary_site_prefixes(batch_dir, site_prefix)
    if not prefixes:
        return [("", latest_detailed_count_summaries(batch_dir, inventory_path, ""))]
    return [
        (prefix, latest_detailed_count_summaries(batch_dir, inventory_path, prefix))
        for prefix in prefixes
    ]


def count_summary_site_prefix_from_tag(tag: str) -> str:
    raw_tag = str(tag or "").strip()
    if not raw_tag:
        return ""

    result: list[str] = []
    for value in re.split(r"[\s,]+", raw_tag):
        key = compact_name(value)
        if not key:
            continue
        if key in {"all", "both", "jbp", "jbpall", "alljbp"}:
            return ""
        if re.fullmatch(r"\d+", key):
            result.append(f"jbp{int(key)}")
            continue
        result.append(key)
    return ",".join(ordered_unique(result))


def expand_count_summary_filter_value(value: str) -> list[str]:
    text = str(value).strip()
    if not text:
        return []

    key = compact_name(text)
    if key in {"all", "alltargets", "allsummary", "summaries"}:
        return []
    if key in {"ipr", "ip", "qfabip", "qfabipr", "t1ipr", "t1toipr"}:
        return ["IPR"]

    range_match = re.fullmatch(
        r"(?:jbp\d+)?(?:pg|dg)?([0-9]+)\s*-\s*(?:jbp\d+)?(?:pg|dg)?([0-9]+)",
        text,
        flags=re.IGNORECASE,
    )
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        step = 1 if end >= start else -1
        return [f"DG{number}" for number in range(start, end + step, step)]

    dg_match = re.fullmatch(r"(?:jbp\d+)?(?:pg|dg)?([0-9]+)", key)
    if dg_match:
        return [f"DG{int(dg_match.group(1))}"]

    return [text.upper()]


def count_summary_filter_targets(summary_tag: str) -> set[str] | None:
    raw_tag = str(summary_tag or "").strip()
    if not raw_tag:
        return None

    targets: list[str] = []
    for value in re.split(r"[\s,]+", raw_tag):
        key = compact_name(value)
        if key in {"all", "alltargets", "allsummary", "summaries"}:
            return None
        targets.extend(expand_count_summary_filter_value(value))
    return {target.upper() for target in targets if target}


def filter_count_summaries(
    summaries: dict[str, dict[str, dict[str, int]]],
    summary_tag: str = "",
) -> dict[str, dict[str, dict[str, int]]]:
    targets = count_summary_filter_targets(summary_tag)
    if targets is None:
        return summaries

    filtered: dict[str, dict[str, dict[str, int]]] = {}
    for side, side_counts in summaries.items():
        filtered[side] = {
            dg: counts
            for dg, counts in side_counts.items()
            if dg.upper() in targets
        }
    return filtered


def detailed_count_summary_total(counts: dict[str, int]) -> int:
    return sum(int(counts.get(key, 0)) for key in DETAILED_COUNT_SUMMARY_KEYS)


def format_count_summary_line(dg: str, counts: dict[str, int]) -> str:
    total = detailed_count_summary_total(counts)
    return (
        f"{dg} (lldp - {counts.get('lldp', 0)}, "
        f"interface-{counts.get('interface', 0)}, "
        f"optics(tx)-{counts.get('optics_tx', 0)}, "
        f"optics(rx)-{counts.get('optics_rx', 0)}, "
        f"optics(temp)-{counts.get('optics_temp', 0)}, "
        f"fec bin-{counts.get('fec_bin', 0)}, "
        f"pre fec-{counts.get('pre_fec', 0)}, "
        f"Total-{total})"
    )


def print_count_summaries(summaries: dict[str, dict[str, dict[str, int]]]) -> None:
    if not any(summaries.values()):
        return

    for side in ("t0", "ipr"):
        side_counts = summaries.get(side, {})
        if not side_counts:
            continue
        print(f"{COUNT_SUMMARY_LABELS[side]} :")
        for dg in sorted(side_counts, key=dg_alias_sort_key):
            print(format_count_summary_line(dg, side_counts[dg]))


def print_latest_count_summaries(
    batch_dir: Path,
    inventory_path: Path = DEFAULT_INVENTORY_PATH,
    site_prefix: str = DEFAULT_SITE_PREFIX,
) -> None:
    print_count_summaries(latest_detailed_count_summaries(batch_dir, inventory_path, site_prefix))


def count_summary_table_rows(
    summaries: dict[str, dict[str, dict[str, int]]]
) -> list[dict[str, int | str]]:
    rows: list[dict[str, int | str]] = []
    for side in ("t0", "ipr"):
        side_counts = summaries.get(side, {})
        for dg in sorted(side_counts, key=dg_alias_sort_key):
            counts = side_counts[dg]
            total = detailed_count_summary_total(counts)
            rows.append(
                {
                    "Side": COUNT_SUMMARY_LABELS[side],
                    "DG": dg,
                    "lldp": counts.get("lldp", 0),
                    "interface": counts.get("interface", 0),
                    "optics(tx)": counts.get("optics_tx", 0),
                    "optics(rx)": counts.get("optics_rx", 0),
                    "optics(temp)": counts.get("optics_temp", 0),
                    "fec bin": counts.get("fec_bin", 0),
                    "pre fec": counts.get("pre_fec", 0),
                    "Total": total,
                }
            )
    return rows


def build_count_summary_sheet(
    workbook,
    summaries: dict[str, dict[str, dict[str, int]]],
    thread_url: str = COUNT_SUMMARY_THREAD_URL,
) -> bool:
    if not any(summaries.values()):
        return False

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if COUNT_SUMMARY_SHEET in workbook.sheetnames:
        del workbook[COUNT_SUMMARY_SHEET]
    worksheet = workbook.create_sheet(COUNT_SUMMARY_SHEET, 0)

    worksheet["A1"] = "Multi-Planar Error Count by DGs across Phases, including T0, T1, and IPR"
    worksheet["A2"] = "Source Slack Thread"
    worksheet["B2"] = thread_url
    worksheet["A3"] = "Generated At"
    worksheet["B3"] = datetime.now().isoformat(timespec="seconds")

    headers = COUNT_SUMMARY_PHASE_HEADERS
    header_row = 5
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    phase_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    dg_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    improvement_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    regression_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row_index = header_row + 1
    t0_counts = summaries.get("t0", {})
    for phase_name, baselines in COUNT_SUMMARY_PHASE_BASELINES.items():
        phase_start_row = row_index
        phase_dgs = sorted(set(baselines) | set(t0_counts), key=dg_alias_sort_key)
        totals = {
            "past": 0,
            "lldp": 0,
            "interface": 0,
            "optics_tx": 0,
            "optics_rx": 0,
            "optics_temp": 0,
            "fec_bin": 0,
            "pre_fec": 0,
            "present": 0,
        }
        for dg in phase_dgs:
            counts = t0_counts.get(dg, {})
            past_total = int(baselines.get(dg, 0))
            category_sum = (
                counts.get("lldp", 0)
                + counts.get("interface", 0)
                + counts.get("optics_tx", 0)
                + counts.get("optics_rx", 0)
                + counts.get("optics_temp", 0)
                + counts.get("fec_bin", 0)
                + counts.get("pre_fec", 0)
            )
            present_total = int(counts.get("present_total", 0) or category_sum)
            delta = present_total - past_total
            values = [
                phase_name,
                dg,
                past_total,
                counts.get("lldp", 0),
                counts.get("interface", 0),
                counts.get("optics_tx", 0),
                counts.get("optics_rx", 0),
                counts.get("optics_temp", 0),
                counts.get("fec_bin", 0),
                counts.get("pre_fec", 0),
                present_total,
                delta,
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            worksheet.cell(row=row_index, column=1).fill = phase_fill
            worksheet.cell(row=row_index, column=2).fill = dg_fill if delta <= 0 else regression_fill
            worksheet.cell(row=row_index, column=12).fill = improvement_fill if delta <= 0 else regression_fill
            if delta > 0:
                worksheet.cell(row=row_index, column=2).font = Font(color="9C0006")
                worksheet.cell(row=row_index, column=12).font = Font(color="9C0006")
            else:
                worksheet.cell(row=row_index, column=2).font = Font(color="006100")
                worksheet.cell(row=row_index, column=12).font = Font(color="006100")

            totals["past"] += past_total
            totals["lldp"] += int(counts.get("lldp", 0))
            totals["interface"] += int(counts.get("interface", 0))
            totals["optics_tx"] += int(counts.get("optics_tx", 0))
            totals["optics_rx"] += int(counts.get("optics_rx", 0))
            totals["optics_temp"] += int(counts.get("optics_temp", 0))
            totals["fec_bin"] += int(counts.get("fec_bin", 0))
            totals["pre_fec"] += int(counts.get("pre_fec", 0))
            totals["present"] += present_total
            row_index += 1

        if row_index - 1 > phase_start_row:
            worksheet.merge_cells(start_row=phase_start_row, start_column=1, end_row=row_index - 1, end_column=1)
            worksheet.cell(row=phase_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

        delta_total = totals["present"] - totals["past"]
        total_values = [
            "",
            "Total",
            totals["past"],
            totals["lldp"],
            totals["interface"],
            totals["optics_tx"],
            totals["optics_rx"],
            totals["optics_temp"],
            totals["fec_bin"],
            totals["pre_fec"],
            totals["present"],
            delta_total,
        ]
        for column_index, value in enumerate(total_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        row_index += 1

        percent = (delta_total / totals["past"]) if totals["past"] else 0
        for column_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        worksheet.cell(row=row_index, column=2, value="Percentage Difference")
        worksheet.cell(row=row_index, column=2).font = Font(bold=True, color="006100")
        worksheet.cell(row=row_index, column=2).fill = total_fill
        worksheet.cell(row=row_index, column=12, value=percent)
        worksheet.cell(row=row_index, column=12).number_format = "0.00%"
        worksheet.cell(row=row_index, column=12).font = Font(bold=True, color="006100" if percent <= 0 else "9C0006")
        worksheet.cell(row=row_index, column=12).fill = improvement_fill if percent <= 0 else regression_fill
        row_index += 2

    ipr_counts = summaries.get("ipr", {})
    if ipr_counts:
        phase_start_row = row_index
        for dg in sorted(ipr_counts, key=dg_alias_sort_key):
            counts = ipr_counts[dg]
            category_sum = (
                counts.get("lldp", 0)
                + counts.get("interface", 0)
                + counts.get("optics_tx", 0)
                + counts.get("optics_rx", 0)
                + counts.get("optics_temp", 0)
                + counts.get("fec_bin", 0)
                + counts.get("pre_fec", 0)
            )
            present_total = int(counts.get("present_total", 0) or category_sum)
            values = [
                "T1<>IPR",
                dg,
                "",
                counts.get("lldp", 0),
                counts.get("interface", 0),
                counts.get("optics_tx", 0),
                counts.get("optics_rx", 0),
                counts.get("optics_temp", 0),
                counts.get("fec_bin", 0),
                counts.get("pre_fec", 0),
                present_total,
                "",
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            worksheet.cell(row=row_index, column=1).fill = phase_fill
            worksheet.cell(row=row_index, column=2).fill = dg_fill
            row_index += 1
        if row_index - 1 > phase_start_row:
            worksheet.merge_cells(start_row=phase_start_row, start_column=1, end_row=row_index - 1, end_column=1)
            worksheet.cell(row=phase_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A6"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet["A1"].font = Font(bold=True, size=13)
    worksheet["A2"].font = Font(bold=True)
    worksheet["A3"].font = Font(bold=True)
    for column_index in range(1, len(headers) + 1):
        values = [
            norm(worksheet.cell(row=row_index, column=column_index).value)
            for row_index in range(1, worksheet.max_row + 1)
        ]
        width = min(max(max((len(value) for value in values), default=0) + 2, 12), 64)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    return True


def add_count_summary_table_sheet(
    workbook,
    summaries: dict[str, dict[str, dict[str, int]]],
) -> None:
    rows = count_summary_table_rows(summaries)
    if not rows:
        return

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if COUNT_SUMMARY_TABLE_SHEET in workbook.sheetnames:
        del workbook[COUNT_SUMMARY_TABLE_SHEET]
    worksheet = workbook.create_sheet(COUNT_SUMMARY_TABLE_SHEET)
    headers = list(rows[0])
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_index, row in enumerate(rows, start=2):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row[header])
            cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index in range(1, len(headers) + 1):
        values = [
            norm(worksheet.cell(row=row_index, column=column_index).value)
            for row_index in range(1, worksheet.max_row + 1)
        ]
        width = min(max(max((len(value) for value in values), default=0) + 2, 12), 40)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def count_summary_workbook_path(batch_dir: Path, site_prefix: str) -> Path:
    prefix = site_prefix.upper() if site_prefix else "JBP"
    timestamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    return next_available_path(batch_dir / f"{prefix}-dg-error-count-summary-{timestamp}.xlsx")


def write_count_summary_workbook(
    output_path: Path,
    detailed_summaries: dict[str, dict[str, dict[str, int]]],
    thread_url: str = COUNT_SUMMARY_THREAD_URL,
) -> Path | None:
    if not any(detailed_summaries.values()):
        return None

    workbook = Workbook()
    workbook.remove(workbook.active)
    build_count_summary_sheet(workbook, detailed_summaries, thread_url)
    add_count_summary_table_sheet(workbook, detailed_summaries)
    workbook.save(output_path)
    workbook.close()
    return output_path


def process_source(
    source_path: Path,
    panels_path: Path,
    panels_df: pd.DataFrame,
    output_path: Path,
    inventory: dict | None = None,
) -> Path:
    if source_path.suffix.lower() == ".csv":
        return process_csv(source_path, panels_path, panels_df, output_path, inventory)
    return process_excel(source_path, panels_path, panels_df, output_path, inventory)


def resolve_inputs(
    inputs: Iterable[str],
    batch_dir: Path,
    force_refresh: bool,
    site_prefix: str,
) -> tuple[list[Path], list[Path]]:
    if inputs:
        paths = [Path(item).expanduser().resolve() for item in inputs]
        missing = [path for path in paths if not path.exists()]
        if missing:
            raise FileNotFoundError("Missing input file(s): " + ", ".join(str(path) for path in missing))
        wrong_site = [
            path
            for path in paths
            if filename_site_name(path) and not path_matches_site_prefix(path, site_prefix)
        ]
        if wrong_site:
            raise ValueError(
                f"This script is for {site_prefix.upper()} only; wrong-site input(s): "
                + ", ".join(path.name for path in wrong_site)
            )
        return paths, []
    return discover_batch_sources(batch_dir, force_refresh, site_prefix)


def main() -> int:
    global RELAX_NON_IPR_OPTICS, RELAX_IPR_OPTICS

    parser = argparse.ArgumentParser(
        description=(
            "JBP15 report and qcli operations helper. With no inputs, batch-process pending "
            "qcli jbp15 *_full_report.xlsx files and JBP15 dashboard Rack All CSV groups "
            "in the script directory."
        ),
        allow_abbrev=False,
    )
    parser.add_argument("inputs", nargs="*", help="Optional report workbook(s) or CSV export(s)")
    parser.add_argument("-p", "--panels", help="PP matrix workbook override")
    parser.add_argument("-o", "--output", help="Output path override; only valid with one explicit input")
    parser.add_argument("--batch-dir", type=Path, default=SCRIPT_DIR, help="Directory scanned in default no-arg mode")
    parser.add_argument(
        "--print-qcli-command",
        action="store_true",
        help="Print JBP15 qcli hc-summary command(s) for --targets, then exit.",
    )
    parser.add_argument(
        "--targets",
        dest="qcli_targets",
        default="all",
        help=(
            "Comma-separated or range targets for --print-qcli-command. "
            "Examples: all, 16-19, DG5,DG8, ipr, plane1."
        ),
    )
    parser.add_argument(
        "--inventory-json",
        type=Path,
        default=DEFAULT_INVENTORY_PATH,
        help="JBP planar AI2ND inventory JSON used by qcli command printers.",
    )
    parser.add_argument(
        "--qcli-state",
        default="deployed",
        help="State filter to include in the printed qcli command; pass an empty string to omit.",
    )
    parser.add_argument(
        "--qcli-no-apex-update",
        action="store_true",
        help="Append --no-apex-update to the printed qcli command.",
    )
    parser.add_argument(
        "--relax",
        action="store_true",
        help="Relax non-IPR optics RX/TX min/max thresholds by 1 dBm.",
    )
    parser.add_argument(
        "--relax-ipr",
        action="store_true",
        help="Relax IPR optics thresholds to RX -5..3 dBm and TX -3..3 dBm.",
    )
    parser.add_argument(
        "--force-refresh",
        action="store_true",
        help="Regenerate even if a timestamped *_with_pp_YYYYMMDD_HHMMSS.xlsx output exists",
    )
    parser.add_argument(
        "--dashboard-group-window-minutes",
        type=int,
        default=DEFAULT_DASHBOARD_GROUP_WINDOW_MINUTES,
        help=(
            "Maximum timestamp span used to cluster RX/TX/Pre-FEC/FEC Bin dashboard CSV exports "
            "from the same run in default no-input mode."
        ),
    )
    parser.add_argument(
        "--generate-summary-report",
        action="store_true",
        help="Write a standalone DG error count summary workbook in default no-input mode.",
    )
    parser.add_argument(
        "--count-summary-tag",
        default="",
        help=(
            "Print latest JBP15 count summaries. Counts are not printed by default. "
            "Examples: jbp15, 15."
        ),
    )
    parser.add_argument(
        "--print-summary-tag",
        default="",
        help=(
            "Filter printed latest count summaries to specific DG/IPR targets. "
            "Examples: 1-3, DG5,DG8, IPR."
        ),
    )
    parser.add_argument("--dry-run", action="store_true", help="Show pending work without writing files")
    args = parser.parse_args()
    RELAX_NON_IPR_OPTICS = args.relax
    RELAX_IPR_OPTICS = args.relax_ipr

    batch_dir = args.batch_dir.expanduser().resolve()
    if args.output and len(args.inputs) != 1:
        print("Error: --output is only valid with exactly one explicit input.", file=sys.stderr)
        return 2

    try:
        if args.print_qcli_command:
            commands = jbp15_qcli_commands_text(
                args.qcli_targets,
                args.inventory_json,
                args.qcli_state.strip(),
                args.qcli_no_apex_update,
            )
            print(qcli_commands_with_optics_relax(commands, args.relax))
            return 0

        batch_site_prefix = DEFAULT_BATCH_SITE_PREFIX
        pending, skipped = resolve_inputs(args.inputs, batch_dir, args.force_refresh, batch_site_prefix)
        should_print_count_summaries = bool(args.count_summary_tag.strip() or args.print_summary_tag.strip())
        if args.count_summary_tag.strip():
            count_summary_site_prefix = count_summary_site_prefix_from_tag(args.count_summary_tag)
            requested_prefixes = site_prefix_values(count_summary_site_prefix)
            if not requested_prefixes or any(prefix != batch_site_prefix for prefix in requested_prefixes):
                raise ValueError(f"This script is for {batch_site_prefix.upper()} only; use the JBP19 script for JBP19.")
        else:
            count_summary_site_prefix = batch_site_prefix
        dashboard_pending: list[DashboardCsvGroup] = []
        dashboard_skipped: list[DashboardCsvGroup] = []
        if not args.inputs:
            dashboard_pending, dashboard_skipped = discover_dashboard_csv_groups(
                batch_dir,
                args.force_refresh,
                args.dashboard_group_window_minutes,
                batch_site_prefix,
            )
        output_timestamp = datetime.now().strftime(OUTPUT_TIMESTAMP_FORMAT)

        if not args.inputs and batch_site_prefix and not should_print_count_summaries:
            print(f"Site prefix: {batch_site_prefix}")
        if not pending and not dashboard_pending:
            if not args.inputs and (should_print_count_summaries or args.generate_summary_report):
                detailed_count_summaries_by_site = latest_detailed_count_summaries_by_site(
                    batch_dir, args.inventory_json, count_summary_site_prefix
                )
                for prefix, detailed_count_summaries in detailed_count_summaries_by_site:
                    printed_count_summaries = filter_count_summaries(
                        detailed_count_summaries,
                        args.print_summary_tag,
                    )
                    if should_print_count_summaries:
                        if prefix and any(printed_count_summaries.values()):
                            print(f"{prefix.upper()} Latest Counts:")
                        print_count_summaries(printed_count_summaries)
                    if args.generate_summary_report and not args.dry_run:
                        summary_path = write_count_summary_workbook(
                            count_summary_workbook_path(batch_dir, prefix or count_summary_site_prefix),
                            detailed_count_summaries,
                        )
                        if summary_path is not None:
                            print(f"Summary workbook: {summary_path}")
            if not should_print_count_summaries:
                print("No pending reports found.")
            return 0

        if pending:
            inventory = load_planar_inventory(args.inventory_json)
            panels_cache: dict[Path, pd.DataFrame] = {}
            print(f"Processing {len(pending)} source(s):")
            for source in pending:
                panels_path = discover_source_panels_path(batch_dir, source, args.panels)
                if panels_path not in panels_cache:
                    panels_cache[panels_path] = load_panels_df(panels_path)
                panels_df = panels_cache[panels_path]
                output_path = (
                    Path(args.output).expanduser().resolve()
                    if args.output
                    else default_output_path(source, output_timestamp)
                )
                print(f"  {source.name} -> {output_path.name}")
                print(f"    PP matrix: {panels_path}")
                if args.dry_run:
                    continue
                result = process_source(source, panels_path, panels_df, output_path, inventory)
                print(f"    wrote {result}")

        if dashboard_pending:
            inventory = load_planar_inventory(args.inventory_json)
            print(f"Processing {len(dashboard_pending)} dashboard CSV group(s):")
            for group in dashboard_pending:
                site_name = dashboard_site_name(group)
                panels_path = discover_dashboard_panels_path(batch_dir, site_name, args.panels)
                report_site_name, generated = build_dashboard_report(group, panels_path, inventory)
                output_path = dashboard_output_path(batch_dir, report_site_name, generated["Summary"], group)
                print(f"  {dashboard_group_label(group)}")
                print(f"    PP matrix: {panels_path}")
                print(f"    -> {output_path.name}")
                if args.dry_run:
                    continue
                result = write_dashboard_report(output_path, generated)
                print(f"    wrote {result}")

        if not args.inputs:
            if should_print_count_summaries or args.generate_summary_report:
                detailed_count_summaries_by_site = latest_detailed_count_summaries_by_site(
                    batch_dir, args.inventory_json, count_summary_site_prefix
                )
                for prefix, detailed_count_summaries in detailed_count_summaries_by_site:
                    printed_count_summaries = filter_count_summaries(
                        detailed_count_summaries,
                        args.print_summary_tag,
                    )
                    if args.generate_summary_report and not args.dry_run:
                        summary_path = write_count_summary_workbook(
                            count_summary_workbook_path(batch_dir, prefix or count_summary_site_prefix),
                            detailed_count_summaries,
                        )
                        if summary_path is not None:
                            print(f"Summary workbook: {summary_path}")
                    if should_print_count_summaries:
                        if prefix and any(printed_count_summaries.values()):
                            print(f"{prefix.upper()} Latest Counts:")
                        print_count_summaries(printed_count_summaries)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
