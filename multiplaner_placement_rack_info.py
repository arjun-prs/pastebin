#!/usr/bin/env python3
"""
Script: multiplaner_placement_rack_info.py
Owner: Surjeet Singh (Surjeet.Singh@oracle.com)
Team: Ai2ND

Script overview
---------------
This script ingests rack-location CSV exports, consolidates them, and generates console tables
that summarize racks by placement group and platform (QFAB T0/T1, IPR, GPU).

Outputs
-------
1) Consolidated CSV: merges all input CSVs into consolidated.csv (adds source_file column).
2) Network Summary: total rack counts and device totals by role.
3) Placement Group Details: per-PG/per-platform rack positions, CFAB block, sector, and link counts.
4) Excel Report: workbook with the same summary/detail tables shown in the console output.

Link count logic (high level)
-----------------------------
- Core PGs (151–154): link count is based on QFABT1 rack count (T1<>IPR rule).
- Non-core PGs: GPU link counts depend on GPU type (b300/gb300), and T0<>T1 link counts are
  computed from the number of QFABT0 racks in that PG using a multiplier derived from PG151.

"""

import pandas as pd
from pathlib import Path
from typing import Optional, Iterable, Set, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rich import print
from rich.table import Table
from rich.console import Console

# === CONFIGURABLE CONSTANTS ===
PLATFORM_COLUMN = "PLATFORM"
PLACEMENT_GROUP_COLUMN = "PLACEMENT_GROUP"
CFAB_BLOCK_COLUMN = "CFAB_FABRIC_BLOCK"
QFAB_INSTANCE_ID = "QFAB_INSTANCE_ID"
BLOCK_NAME = "BLOCK_NAME"
RACK_NUMBER = "RACK_NUMBER"
ROOM_NAME = "ROOM_NAME"
SUMMARY_HEADERS = ["Role", "RACK Count", "Notes", "Total Device Count"]
DETAIL_HEADERS = [
    "Placement Group",
    "Activity",
    "Rack SKU",
    "Rack position",
    "Total Racks",
    "CFAB Block",
    "DH Sector",
    "Link Count",
]

# ---------- Filters / classification ----------
def _is_missing(value: object) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except TypeError:
        return False

def _to_clean_str(value: object) -> str:
    if _is_missing(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "<na>"} else text

def _to_clean_lower(value: object) -> str:
    return _to_clean_str(value).lower()

def _unique_non_empty_strings(values: Iterable[object]) -> list[str]:
    unique_values: list[str] = []
    seen: Set[str] = set()
    for value in values:
        text = _to_clean_str(value)
        if not text or text in seen:
            continue
        seen.add(text)
        unique_values.append(text)
    return unique_values

def is_allowed_platform(platform: str) -> bool:
    """
    Only keep platforms that contain one of:
      qfab_t0, qfab_t1, ipr, gpu
    (also supports qfabt0/qfabt1 without underscore)
    """
    p = _to_clean_lower(platform)
    return (
        ("qfab_t0" in p) or ("qfabt0" in p) or
        ("qfab_t1" in p) or ("qfabt1" in p) or
        ("ipr" in p) or
        ("gpu" in p)
    )

def _role_from_platform(platform: str) -> Optional[str]:
    p = _to_clean_lower(platform)
    if "ipr" in p or "inter" in p:
        return "QFABIP"
    if "qfab_t1" in p or "qfabt1" in p:
        return "QFABT1"
    if "qfab_t0" in p or "qfabt0" in p:
        return "QFABT0"
    if "gpu" in p:
        return "GPU"
    return None

def activity_from_platform(platform: str) -> str:
    p = _to_clean_lower(platform)
    if "aux" in p or "oad_aux" in p:
        return "AUX racks\nto be\nbootstrap"
    if "ipr" in p:
        return "Bootstrap\nIPR rack"
    if "qfab_t1" in p or "qfabt1" in p:
        return "QFABT1\nspine(t1) racks\nto be\nbootstrap"
    if "qfab_t0" in p or "qfabt0" in p:
        return "QFABT0\nleaf(t0) racks\nto be\nbootstrap"
    if "gpu" in p:
        return "GPU racks\nto be\nbootstrapp"
    return "Bootstrap"

# ---------- Link count calculators ----------
def calculate_t1_ipr_link_count(platforms_dict: Dict[str, dict]) -> int:
    """T1<>IPR link count per PG: (#unique QFABT1 racks) * 8"""
    t1_racks = set()
    for platform, pinfo in (platforms_dict or {}).items():
        if _role_from_platform(platform) != "QFABT1":
            continue
        racks = (pinfo or {}).get("rack_number", []) or []
        t1_racks.update(_unique_non_empty_strings(racks))
    return len(t1_racks) * 8

def calculate_gpu_link_count(platforms_dict: Dict[str, dict]) -> int:
    """
    GPU link count per PG:
      - b300:  32 * (#unique racks for b300 platforms in that PG)
      - gb300: 288 * (#unique racks for gb300 platforms in that PG)
    """
    b300_racks = set()
    gb300_racks = set()
    for platform, pinfo in (platforms_dict or {}).items():
        if _role_from_platform(platform) != "GPU":
            continue
        p = _to_clean_lower(platform)
        racks = (pinfo or {}).get("rack_number", []) or []
        racks_set = set(_unique_non_empty_strings(racks))
        if "gb300" in p:
            gb300_racks |= racks_set
        elif "b300" in p:
            b300_racks |= racks_set
    return (len(b300_racks) * 32) + (len(gb300_racks) * 288)

def get_pg151_t1_reference_multiplier(
    df: pd.DataFrame,
    placement_group_column: str,
    platform_column: str,
    rack_number: str
) -> int:
    """
    Reference rule:
      - Count QFABT1 racks in placement group 151
      - If > 8 => multiplier = 128
      - else  => multiplier = 64
    """
    pg151 = df[df[placement_group_column].astype(str) == "151"]
    if pg151.empty:
        return 64

    t1 = pg151[pg151[platform_column].apply(lambda x: _role_from_platform(x) == "QFABT1")]
    if t1.empty:
        return 64

    t1_racks = set(_unique_non_empty_strings(t1[rack_number].tolist()))
    return 128 if len(t1_racks) > 8 else 64

def calculate_t0_t1_link_count_for_pg(platforms_dict: Dict[str, dict], reference_multiplier: int) -> int:
    """
    For an individual PG:
      t0<>t1 link count = reference_multiplier * 8 * (#unique QFABT0 racks in that PG)
    Where reference_multiplier is determined ONLY from PG 151's T1 rack count (64 or 128).
    """
    t0_racks = set()
    for platform, pinfo in (platforms_dict or {}).items():
        if _role_from_platform(platform) != "QFABT0":
            continue
        racks = (pinfo or {}).get("rack_number", []) or []
        t0_racks.update(_unique_non_empty_strings(racks))
    return reference_multiplier * 8 * len(t0_racks)

# ---------- CSV consolidation ----------
def consolidate_csv(folder: Path, pattern: str, out_file: Path) -> None:
    """Consolidate CSV files in a folder into a single CSV file."""
    if out_file.exists():
        out_file.unlink()
    dfs = []
    for f in folder.glob(pattern):
        if f.resolve() == out_file.resolve():
            continue
        df = pd.read_csv(f, dtype=str, keep_default_na=False)
        df["source_file"] = f.name
        dfs.append(df)
    if not dfs:
        raise FileNotFoundError(f"No CSV files found in {folder} matching {pattern}")
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    combined.to_csv(out_file, index=False)
    print(f"Wrote {len(combined):,} rows to {out_file}")

# ---------- Data extraction ----------
def location_file_reader(
    input_file: Path,
    platform_column: str = PLATFORM_COLUMN,
    placement_group_column: str = PLACEMENT_GROUP_COLUMN,
    cfab_block_column: str = CFAB_BLOCK_COLUMN,
    instance_id: str = QFAB_INSTANCE_ID,
    block_name: str = BLOCK_NAME,
    rack_number: str = RACK_NUMBER,
    room_name: str = ROOM_NAME,
) -> tuple:
    """
    Read a CSV file and analyze the data.
    Returns: (gpu_pg_dict, core_network_pg_dict)
    """
    try:
        df = pd.read_csv(input_file, dtype=str, keep_default_na=False)

        required_columns = {
            platform_column,
            placement_group_column,
            cfab_block_column,
            rack_number,
            room_name,
        }
        missing_columns = sorted(required_columns - set(df.columns))
        if missing_columns:
            raise KeyError(f"Missing required columns: {', '.join(missing_columns)}")

        columns_to_convert = [placement_group_column, cfab_block_column, instance_id, block_name]
        for column in columns_to_convert:
            if column in df.columns:
                df[column] = df[column].apply(lambda value: _to_clean_str(value).removesuffix(".0"))

        for column in [platform_column, rack_number, room_name]:
            df[column] = df[column].apply(_to_clean_str)

        gpu_dict, core_dict = get_placement_group_info(
            df,
            placement_group_column,
            platform_column,
            cfab_block_column,
            rack_number,
            room_name,
        )
        return gpu_dict, core_dict
    except FileNotFoundError:
        print(f"File {input_file} not found.")
        return None, None
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None, None

def get_placement_group_info(
    df: pd.DataFrame,
    placement_group_column: str,
    platform_column: str,
    cfab_block_column: str,
    rack_number: str,
    room_name: str,
) -> tuple:
    core_network_pg_set = {151, 152, 153, 154, 201, 202}
    core_link_pgs = {151, 152, 153, 154}

    ref_multiplier = get_pg151_t1_reference_multiplier(
        df, placement_group_column, platform_column, rack_number
    )

    core_network_pg_dict: Dict[str, dict] = {}
    gpu_pg_dict: Dict[str, dict] = {}

    for pg in df[placement_group_column].unique():
        pg_str = _to_clean_str(pg)
        if not pg_str.isdigit():
            continue
        pg_num = int(pg_str)

        pg_df = df[df[placement_group_column] == pg_str]
        gpu_platform_dict: Dict[str, dict] = {}
        core_network_platform_dict: Dict[str, dict] = {}

        for platform in pg_df[platform_column].unique():
            platform_name = _to_clean_str(platform)
            if not is_allowed_platform(platform_name):
                continue

            platform_df = pg_df[pg_df[platform_column] == platform_name]
            if platform_df.empty:
                continue

            cfab_values = _unique_non_empty_strings(platform_df[cfab_block_column].tolist())
            rack_numbers = _unique_non_empty_strings(platform_df[rack_number].tolist())
            room_values = _unique_non_empty_strings(platform_df[room_name].tolist())

            if not rack_numbers:
                continue

            cfab_block = cfab_values[0] if cfab_values else None
            room_name_val = room_values[0] if room_values else None

            if pg_num in core_network_pg_set:
                core_network_platform_dict[platform_name] = {
                    "cfab_fabric_block": cfab_block,
                    "rack_number": rack_numbers,
                    "room_name": room_name_val,
                    "count": len(rack_numbers),
                }
            else:
                role = _role_from_platform(platform_name)
                if role in {"GPU", "QFABT0"}:
                    gpu_platform_dict[platform_name] = {
                        "cfab_fabric_block": cfab_block,
                        "rack_number": rack_numbers,
                        "room_name": room_name_val,
                        "count": len(rack_numbers),
                    }

        # Non-core PGs
        if gpu_platform_dict:
            pg_entry = {"platforms": gpu_platform_dict}
            pg_entry["gpu_link_count"] = calculate_gpu_link_count(gpu_platform_dict)
            pg_entry["t0_t1_link_count"] = calculate_t0_t1_link_count_for_pg(
                gpu_platform_dict, reference_multiplier=ref_multiplier
            )
            gpu_pg_dict[pg_str] = pg_entry

        # Core PGs
        if core_network_platform_dict:
            pg_entry = {"platforms": core_network_platform_dict}
            if pg_num in core_link_pgs:
                pg_entry["t1_ipr_link_count"] = calculate_t1_ipr_link_count(core_network_platform_dict)
            core_network_pg_dict[pg_str] = pg_entry

    return gpu_pg_dict, core_network_pg_dict

# ---------- Summary table ----------
QFAB_DEVICES_PER_RACK = {
    "QFABIP": 8,
    "QFABT1": 8,
    "QFABT0": 8,
}

def qfab_notes(role: str, total_devices: int, devices_per_rack: int = 8) -> str:
    if role == "QFABIP":
        return "Inter Planar Rack"
    if role in ("QFABT1", "QFABT0"):
        devices_per_plane = total_devices / 4
        if float(devices_per_plane).is_integer():
            devices_per_plane = int(devices_per_plane)
        return f"{devices_per_plane} devices per plane\n{devices_per_rack} devices per rack"
    return ""

def gpu_per_rack_from_platform(platform: str, default: int = 8) -> int:
    """Decide GPUs per rack based on shape keywords in platform name."""
    p = _to_clean_lower(platform)
    if "gb300" in p:
        return 72
    if "b300" in p:
        return 8
    return default

def _collect_racks_by_platform(pg_dict: dict) -> Dict[str, Set[str]]:
    """Return: { platform: set(unique_rack_numbers) }"""
    racks_by_platform: Dict[str, Set[str]] = {}
    for _pg, pg_info in (pg_dict or {}).items():
        platforms = (pg_info or {}).get("platforms", {}) or {}
        for platform, pinfo in platforms.items():
            if not is_allowed_platform(platform):
                continue
            racks = (pinfo or {}).get("rack_number", []) or []
            rack_set = racks_by_platform.setdefault(platform, set())
            rack_set.update(_unique_non_empty_strings(racks))
    return racks_by_platform

def build_summary_rows(core_network_pg_dict: dict, gpu_pg_dict: dict) -> list[dict[str, str]]:
    racks_by_platform: Dict[str, Set[str]] = {}
    for src in (_collect_racks_by_platform(core_network_pg_dict),
                _collect_racks_by_platform(gpu_pg_dict)):
        for platform, racks in src.items():
            racks_by_platform.setdefault(platform, set()).update(racks)

    racks_by_role: Dict[str, Set[str]] = {"QFABIP": set(), "QFABT1": set(), "QFABT0": set(), "GPU": set()}
    gpu_device_total = 0

    for platform, racks in racks_by_platform.items():
        role = _role_from_platform(platform)
        if not role:
            continue
        racks_by_role.setdefault(role, set()).update(racks)
        if role == "GPU":
            per_rack = gpu_per_rack_from_platform(platform, default=8)
            gpu_device_total += len(racks) * per_rack

    total_racks = 0
    total_devices = 0
    rows: list[dict[str, str]] = []

    for role in ["QFABIP", "QFABT1", "QFABT0"]:
        rack_count = len(racks_by_role.get(role, set()))
        device_count = rack_count * QFAB_DEVICES_PER_RACK[role]
        total_racks += rack_count
        total_devices += device_count
        rows.append({
            "Role": role,
            "RACK Count": str(rack_count),
            "Notes": qfab_notes(role, device_count, devices_per_rack=QFAB_DEVICES_PER_RACK[role]),
            "Total Device Count": str(device_count),
        })

    gpu_rack_count = len(racks_by_role.get("GPU", set()))
    total_racks += gpu_rack_count
    total_devices += gpu_device_total
    rows.append({
        "Role": "GPU",
        "RACK Count": str(gpu_rack_count),
        "Notes": "GPUs per rack depends on shape",
        "Total Device Count": str(gpu_device_total),
    })
    rows.append({
        "Role": "Total",
        "RACK Count": str(total_racks),
        "Notes": "",
        "Total Device Count": str(total_devices),
        "_is_total": "true",
    })
    return rows

def summary_table(core_network_pg_dict: dict, gpu_pg_dict: dict) -> Table:
    rows = build_summary_rows(core_network_pg_dict, gpu_pg_dict)

    table = Table(title="Network Summary")
    table.add_column("Role", style="bold")
    table.add_column("RACK Count", justify="right")
    table.add_column("Notes")
    table.add_column("Total Device Count", justify="right")

    for row in rows:
        table.add_row(
            row["Role"],
            row["RACK Count"],
            row["Notes"],
            row["Total Device Count"],
            style="bold" if row.get("_is_total") == "true" else None,
        )
    return table

# ---------- Details table (core or gpu) ----------
def _normalize_pg_set(values: Optional[Iterable]) -> Optional[Set[str]]:
    if values is None:
        return None
    return set(str(v) for v in values)

def build_platform_rack_detail_rows(
    pg_dict: dict,
    include_pgs: Optional[Iterable[int]] = None,
) -> list[dict[str, str]]:
    include_set = _normalize_pg_set(include_pgs)
    agg: Dict[Tuple[str, str], Dict[str, Set[str]]] = {}

    for pg, pg_info in (pg_dict or {}).items():
        pg_str = str(pg)
        if include_set is not None and pg_str not in include_set:
            continue

        platforms = (pg_info or {}).get("platforms", {}) or {}
        for platform, pinfo in platforms.items():
            if not is_allowed_platform(platform):
                continue

            racks = (pinfo or {}).get("rack_number", []) or []
            cfab = (pinfo or {}).get("cfab_fabric_block", None)
            room = (pinfo or {}).get("room_name", None)

            key = (pg_str, platform)
            if key not in agg:
                agg[key] = {"racks": set(), "cfab_blocks": set(), "rooms": set()}

            agg[key]["racks"].update(_unique_non_empty_strings(racks))

            cfab_text = _to_clean_str(cfab)
            room_text = _to_clean_str(room)
            if cfab_text:
                agg[key]["cfab_blocks"].add(cfab_text)
            if room_text:
                agg[key]["rooms"].add(room_text)

    rows: list[dict[str, str]] = []

    def sort_key(item):
        (pg_str, platform), _ = item
        try:
            pg_num = int(pg_str)
        except ValueError:
            pg_num = 10**9
        return (pg_num, platform)

    for (pg_str, platform), info in sorted(agg.items(), key=sort_key):
        racks_sorted = sorted(info["racks"], key=lambda x: int(x) if x.isdigit() else x)
        rack_positions = "\n".join(racks_sorted)
        cfab_block = "\n".join(sorted(info["cfab_blocks"])) if info["cfab_blocks"] else ""
        dh_sector = "\n".join(sorted(info["rooms"])) if info["rooms"] else ""

        pg_level = (pg_dict or {}).get(pg_str, {}) or {}
        t1_ipr = pg_level.get("t1_ipr_link_count", "")
        gpu_lc = pg_level.get("gpu_link_count", "")
        t0_t1 = pg_level.get("t0_t1_link_count", "")

        if _role_from_platform(platform) == "QFABT0" and t0_t1 != "":
            link_count_cell = str(t0_t1)
        else:
            primary = t1_ipr if t1_ipr != "" else gpu_lc
            link_count_cell = str(primary) if primary != "" else ""

        rows.append({
            "Placement Group": pg_str,
            "Activity": activity_from_platform(platform),
            "Rack SKU": platform,
            "Rack position": rack_positions,
            "Total Racks": str(len(racks_sorted)),
            "CFAB Block": cfab_block,
            "DH Sector": dh_sector,
            "Link Count": link_count_cell,
        })

    return rows

def render_platform_rack_details(
    pg_dict: dict,
    include_pgs: Optional[Iterable[int]] = None,
    title: str = "Platform Rack Details",
    console: Optional[Console] = None,
) -> None:
    """
    Single generic Link Count behavior:
      - If row is QFABT0 => show t0_t1_link_count (PG-level)
      - Else => show t1_ipr_link_count if present else gpu_link_count if present
    """
    if console is None:
        console = Console()
    if not pg_dict:
        console.print("[bold yellow]No data to display.[/bold yellow]")
        return

    rows = build_platform_rack_detail_rows(pg_dict, include_pgs=include_pgs)
    if not rows:
        console.print("[bold yellow]No matching placement groups found (after filtering).[/bold yellow]")
        return

    table = Table(title=title, show_lines=True)
    table.add_column("Placement Group", style="bold", no_wrap=False)
    table.add_column("Activity", no_wrap=False)
    table.add_column("Rack SKU", no_wrap=False)
    table.add_column("Rack position", no_wrap=False, justify="center")
    table.add_column("Total Racks", justify="center")
    table.add_column("CFAB Block", justify="center")
    table.add_column("DH Sector", no_wrap=False)
    table.add_column("Link Count", justify="center")

    for row in rows:
        table.add_row(
            row["Placement Group"],
            row["Activity"],
            row["Rack SKU"],
            row["Rack position"],
            row["Total Racks"],
            row["CFAB Block"],
            row["DH Sector"],
            row["Link Count"],
        )

    console.print(table)

def _apply_excel_table_format(
    ws,
    title: str,
    headers: list[str],
    rows: list[dict[str, str]],
    center_columns: Optional[Set[str]] = None,
    right_columns: Optional[Set[str]] = None,
) -> None:
    center_columns = center_columns or set()
    right_columns = right_columns or set()

    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    title_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_fill = PatternFill(fill_type="solid", fgColor="B8CCE4")
    total_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill
    title_cell.border = border
    ws.row_dimensions[1].height = 24

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = border
    ws.row_dimensions[2].height = 22

    if not rows:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        cell = ws.cell(row=3, column=1, value="No data to display.")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.row_dimensions[3].height = 20
    else:
        for row_idx, row in enumerate(rows, start=3):
            row_values = [row.get(header, "") for header in headers]
            line_count = max(max(str(value).count("\n") + 1, 1) for value in row_values)
            ws.row_dimensions[row_idx].height = max(18, line_count * 15)

            is_total = row.get("_is_total") == "true"
            for col_idx, header in enumerate(headers, start=1):
                value = row.get(header, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if header in center_columns:
                    horizontal = "center"
                elif header in right_columns:
                    horizontal = "right"
                else:
                    horizontal = "left"
                cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=True)
                cell.border = border
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill

    for col_idx, header in enumerate(headers, start=1):
        longest = len(header)
        for row in rows:
            value = _to_clean_str(row.get(header, ""))
            if not value:
                continue
            longest = max(longest, max(len(line) for line in value.splitlines()))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 12), 40)

    last_row = max(3, len(rows) + 2)
    last_col = get_column_letter(len(headers))
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}{last_row}"

def export_report_to_excel(
    output_file: Path,
    core_network_pg_dict: dict,
    gpu_pg_dict: dict,
    core_pgs: Optional[Iterable[int]] = None,
    gpu_pgs: Optional[Iterable[int]] = None,
) -> None:
    workbook = Workbook()

    summary_rows = build_summary_rows(core_network_pg_dict, gpu_pg_dict)
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _apply_excel_table_format(
        summary_sheet,
        title="Network Summary",
        headers=SUMMARY_HEADERS,
        rows=summary_rows,
        right_columns={"RACK Count", "Total Device Count"},
    )

    core_rows = build_platform_rack_detail_rows(core_network_pg_dict, include_pgs=core_pgs)
    core_title = (
        "Core Network Details (ALL PGs)"
        if core_pgs is None
        else f"Core Network Details for PGs: {', '.join(map(str, core_pgs))}"
    )
    core_sheet = workbook.create_sheet("Core Details")
    _apply_excel_table_format(
        core_sheet,
        title=core_title,
        headers=DETAIL_HEADERS,
        rows=core_rows,
        center_columns={"Placement Group", "Rack position", "Total Racks", "CFAB Block", "Link Count"},
    )

    gpu_rows = build_platform_rack_detail_rows(gpu_pg_dict, include_pgs=gpu_pgs)
    gpu_title = (
        "GPU Details (ALL PGs)"
        if gpu_pgs is None
        else f"GPU Details for PGs: {', '.join(map(str, gpu_pgs))}"
    )
    gpu_sheet = workbook.create_sheet("GPU Details")
    _apply_excel_table_format(
        gpu_sheet,
        title=gpu_title,
        headers=DETAIL_HEADERS,
        rows=gpu_rows,
        center_columns={"Placement Group", "Rack position", "Total Racks", "CFAB Block", "Link Count"},
    )

    workbook.save(output_file)
    print(f"Wrote Excel report to {output_file}")

def get_available_pgs(pg_dict: dict) -> list[int]:
    """Return sorted list of placement groups present in the dict."""
    pgs = []
    for pg in (pg_dict or {}).keys():
        try:
            pgs.append(int(str(pg)))
        except ValueError:
            continue
    return sorted(set(pgs))

def prompt_pg_selection(available_pgs: list[int], prompt_text: str) -> Optional[list[int]]:
    """
    User can:
      - press Enter => return None (meaning: run for ALL available PGs)
      - enter comma-separated list (e.g., 151,152,201)
      - enter ranges (e.g., 151-154,201)
    Returns list[int] or None.
    """
    if not available_pgs:
        return None

    print(f"\n{prompt_text}")
    print(f"Available PGs: {', '.join(map(str, available_pgs))}")
    raw = input("Enter PGs (comma/range) or press Enter for ALL: ").strip()

    if raw == "":
        return None  # ALL

    selected: Set[int] = set()
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    for part in parts:
        if "-" in part:
            a, b = part.split("-", 1)
            a, b = a.strip(), b.strip()
            if a.isdigit() and b.isdigit():
                start, end = int(a), int(b)
                if start > end:
                    start, end = end, start
                selected.update(range(start, end + 1))
        else:
            if part.isdigit():
                selected.add(int(part))

    selected = {pg for pg in selected if pg in set(available_pgs)}
    return sorted(selected)

# ---------- Main ----------
if __name__ == "__main__":
    dir_path = input("Please enter the directory path containing the location files:").strip()
    folder = Path(dir_path).expanduser()
    if not folder.exists() or not folder.is_dir():
        raise ValueError(f"Invalid directory: {folder}")

    pattern = "*.csv"
    out_file = folder / "consolidated.csv"
    excel_report_file = folder / "multiplaner_placement_rack_info.xlsx"

    consolidate_csv(folder, pattern, out_file)
    gpu_pg_dict, core_network_pg_dict = location_file_reader(out_file)
    if gpu_pg_dict is None or core_network_pg_dict is None:
        raise SystemExit("Failed to analyze the consolidated CSV. See the error above.")

    # Summary
    print(summary_table(core_network_pg_dict, gpu_pg_dict))

    # Core network details
    available_core_pgs = get_available_pgs(core_network_pg_dict)
    core_pgs = prompt_pg_selection(available_core_pgs, "Core Network Placement Groups")
    render_platform_rack_details(
        core_network_pg_dict,
        include_pgs=core_pgs,  # None means ALL
        title=("Core Network Details (ALL PGs)" if core_pgs is None
               else f"Core Network Details for PGs: {', '.join(map(str, core_pgs))}"),
    )

    # GPU details
    available_gpu_pgs = get_available_pgs(gpu_pg_dict)
    gpu_pgs = prompt_pg_selection(available_gpu_pgs, "GPU Placement Groups")
    render_platform_rack_details(
        gpu_pg_dict,
        include_pgs=gpu_pgs,  # None means ALL
        title=("GPU Details (ALL PGs)" if gpu_pgs is None
               else f"GPU Details for PGs: {', '.join(map(str, gpu_pgs))}"),
    )

    export_report_to_excel(
        excel_report_file,
        core_network_pg_dict,
        gpu_pg_dict,
        core_pgs=core_pgs,
        gpu_pgs=gpu_pgs,
    )
