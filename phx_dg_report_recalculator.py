#!/usr/bin/env python3
"""Build a PHX DG progress report from qcli full_report workbooks.

By default this script reads every *_full_report.xlsx in ~/qclihcdata/phx and
maps them to DG1, DG2, ... in filename order. Use --excel DG=path when the file
order should be explicit.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence


DEFAULT_DATA_DIR = Path("~/qclihcdata/phx").expanduser()
DEFAULT_PREVIOUS_REPORT = DEFAULT_DATA_DIR / "previous.txt"
TOTAL_TARGET = "TOTAL"
ISSUE_METRICS = (
    "LLDP",
    "Interface",
    "Optics TX/RX",
    "Combined FEC",
    "T1->T0 Optics RX/TX",
    "T1->T0 Combined FEC",
)
SPLIT_FEC_METRICS = ("Pre Fec", "Fec Bin")
ALL_METRICS = (*ISSUE_METRICS, *SPLIT_FEC_METRICS, "Total")
PRE_FEC_BER_THRESHOLD = 1e-7
FEC_BIN_MIN = 7
FEC_BIN_MAX = 15
NON_IPR_RELAX_RX_MIN_DBM = -4.0
NON_IPR_RELAX_TX_MIN_DBM = -1.0
NON_IPR_RELAX_MAX_DBM = 5.5
IPR_RELAX_RX_MIN_DBM = -5.0
IPR_RELAX_TX_MIN_DBM = -3.0
IPR_RELAX_MAX_DBM = 3.0

COMPARISON_HEADERS = ("Metric", "Previous", "Current", "Absolute", "Progress %")
TARGET_HEADERS = ("Target", "Previous", "Current", "Absolute", "Progress %")

FEC_BIN_COUNT_RE = re.compile(
    r"FEC[_\s-]*BIN[_\s-]*(\d+)[_\s-]*COUNT\s*=\s*([-+]?\d+)",
    re.I,
)
FEC_BIN_RE = re.compile(r"FEC[_\s-]*BIN[_\s-]*(\d+)", re.I)
FLOAT_RE = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?")
TARGET_RE = re.compile(r"DG\d+", re.I)
DEFAULT_INSIGHTS_PROVIDER = os.environ.get("PHX_DG_INSIGHTS_PROVIDER", "auto")
DEFAULT_INSIGHTS_MODEL = os.environ.get("PHX_DG_INSIGHTS_MODEL", "gpt-5.5")
DEFAULT_INSIGHTS_BASE_URL = os.environ.get("PHX_DG_INSIGHTS_BASE_URL") or os.environ.get(
    "OPENAI_BASE_URL",
    "https://api.openai.com/v1",
)
DEFAULT_INSIGHTS_TIMEOUT = float(os.environ.get("PHX_DG_INSIGHTS_TIMEOUT", "30"))


@dataclass(frozen=True)
class SiteConfig:
    generation: int
    dg_racks: Mapping[str, tuple[str, ...]]


DEFAULT_SITE_TAG = "phx23"
PHX20_QCLI_TARGETS = ("DG1", "DG2", "DG3", "DG4")
SITE_CONFIGS: dict[str, SiteConfig] = {
    "phx23": SiteConfig(
        generation=23,
        dg_racks={
            "DG1": ("4206", "4207", "4306", "4307"),
            "DG2": ("4217", "4218", "4317", "4318"),
            "DG3": ("4226", "4227", "4326", "4327"),
            "DG4": ("5817", "5818", "5917", "5918"),
            "DG5": ("5826", "5827", "5926", "5927"),
        },
    ),
    "phx20": SiteConfig(
        generation=20,
        dg_racks={
            "DG1": ("1307", "1308"),
            "DG2": ("1407", "1408"),
        },
    ),
}
RACK_TO_DG_BY_SITE = {
    site_tag: {rack: target for target, racks in config.dg_racks.items() for rack in racks}
    for site_tag, config in SITE_CONFIGS.items()
}

GPU_REMOTE_HOST_COLUMNS = (
    "Remote Host",
    "Mapped Remote Host",
    "Z_end_host",
    "Z End Host",
    "Remote Hostname",
    "Mapped Remote Hostname",
)


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def normalize_site_tag(value: str | None) -> str:
    normalized = str(value or DEFAULT_SITE_TAG).strip().lower()
    return normalized or DEFAULT_SITE_TAG


def site_config(site_tag: str | None) -> SiteConfig:
    normalized = normalize_site_tag(site_tag)
    try:
        return SITE_CONFIGS[normalized]
    except KeyError as exc:
        valid = ", ".join(sorted(SITE_CONFIGS))
        raise ValueError(f"Unknown site tag {site_tag!r}; expected one of: {valid}") from exc


def dg_racks_for_site(site_tag: str | None) -> Mapping[str, tuple[str, ...]]:
    return site_config(site_tag).dg_racks


def all_dg_targets(site_tag: str | None) -> tuple[str, ...]:
    return tuple(dg_racks_for_site(site_tag))


def qcli_targets_for_site(site_tag: str | None) -> tuple[str, ...]:
    if normalize_site_tag(site_tag) == "phx20":
        return PHX20_QCLI_TARGETS
    return all_dg_targets(site_tag)


def rack_to_dg_for_site(site_tag: str | None) -> Mapping[str, str]:
    normalized = normalize_site_tag(site_tag)
    site_config(normalized)
    return RACK_TO_DG_BY_SITE[normalized]


def normalize_target(value: str) -> str:
    match = re.fullmatch(r"(?:DG|PG)?\s*(\d+)", str(value).strip().upper().rstrip(":"))
    if not match:
        raise ValueError(f"Invalid target {value!r}; expected values like DG1 or DG4")
    return f"DG{int(match.group(1))}"


def parse_target_list(
    raw: str | None,
    site_tag: str | None = None,
    all_targets: Sequence[str] | None = None,
) -> list[str] | None:
    if not raw:
        return None

    targets: list[str] = []
    for item in raw.split(","):
        item = item.strip()
        if not item:
            continue
        if item.lower() == "all":
            for target in all_targets or all_dg_targets(site_tag):
                if target not in targets:
                    targets.append(target)
            continue
        target = normalize_target(item)
        if target not in targets:
            targets.append(target)

    if not targets:
        raise ValueError("No valid targets were provided")
    return targets


def racks_for_targets(targets: Sequence[str], site_tag: str | None = None) -> list[str]:
    dg_racks = dg_racks_for_site(site_tag)
    racks: list[str] = []
    for target in targets:
        if target not in dg_racks:
            raise ValueError(f"No {normalize_site_tag(site_tag).upper()} rack mapping exists for {target}")
        for rack in dg_racks[target]:
            if rack not in racks:
                racks.append(rack)
    return racks


def combined_qcli_command(targets: Sequence[str], site_tag: str | None = None, relax: bool = False) -> str:
    relax_flag = " --optics-relax" if relax else ""
    if normalize_site_tag(site_tag) == "phx20":
        unsupported = [target for target in targets if target not in PHX20_QCLI_TARGETS]
        if unsupported:
            raise ValueError(
                "No PHX20 qcli --deployment-group command mapping exists for: "
                f"{', '.join(unsupported)}"
            )
        dg_values = ",".join(str(int(target.removeprefix("DG"))) for target in targets)
        return (
            "qcli hc-summary --region phx --building 20 --instance 2 --state deployed "
            f"--spectrum --customtag test_spectrum_fec_ber --failures-only --filter_pg --t1-reports --nvidia-linkflap-clear --slack{relax_flag} --deployment-group {dg_values}"
        )

    config = site_config(site_tag)
    dg_racks = dg_racks_for_site(site_tag)
    dg_values: list[str] = []
    for target in targets:
        if target not in dg_racks:
            raise ValueError(f"No {normalize_site_tag(site_tag).upper()} rack mapping exists for {target}")
        dg_values.append(str(int(target.removeprefix("DG"))))
    return (
        f"qcli hc-summary --region phx --building {config.generation} --instance 2 --state deployed "
        f"--spectrum --customtag test_spectrum_fec_ber --failures-only --filter_pg --t1-reports --nvidia-linkflap-clear --slack{relax_flag} --deployment-group {','.join(dg_values)}"
    )


def metric_template() -> dict[str, int]:
    return {metric: 0 for metric in ALL_METRICS}


def parse_int(value: object) -> int:
    text = str(value or "").replace(",", "").strip()
    match = re.search(r"[+-]?\d+", text)
    if not match:
        return 0
    return int(match.group(0))


def parse_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    match = FLOAT_RE.search(text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def optics_power_type(
    header: Sequence[object],
    row: Sequence[object],
    source_kind: str | None = None,
) -> str:
    if source_kind in {"rx", "tx"}:
        return source_kind.upper()
    for value in (
        cell(header, row, ("Power Type",)),
        cell(header, row, ("Metric",)),
    ):
        normalized = normalize(value)
        if normalized in {"inputpower", "rx", "rxoptics"}:
            return "RX"
        if normalized in {"outputpower", "tx", "txoptics"}:
            return "TX"
    return ""


def optics_measured_dbm(header: Sequence[object], row: Sequence[object]) -> float | None:
    for value in (
        cell(header, row, ("Measured (dBm)",)),
        cell(header, row, ("Value",)),
        cell(header, row, ("RX power",)),
        cell(header, row, ("TX power",)),
        cell(header, row, ("Input Power",)),
        cell(header, row, ("Output Power",)),
        cell(header, row, ("Measured",)),
        cell(header, row, ("Power",)),
    ):
        measured = parse_float(value)
        if measured is not None:
            return measured
    return None


def optics_row_passes_relaxed_threshold(
    header: Sequence[object],
    row: Sequence[object],
    *,
    relax: bool = False,
    relax_ipr: bool = False,
    is_ipr: bool = False,
    source_kind: str | None = None,
) -> bool:
    if is_ipr:
        if not relax_ipr:
            return True
        rx_min = IPR_RELAX_RX_MIN_DBM
        tx_min = IPR_RELAX_TX_MIN_DBM
        max_value = IPR_RELAX_MAX_DBM
    else:
        if not relax:
            return True
        rx_min = NON_IPR_RELAX_RX_MIN_DBM
        tx_min = NON_IPR_RELAX_TX_MIN_DBM
        max_value = NON_IPR_RELAX_MAX_DBM

    power_type = optics_power_type(header, row, source_kind)
    measured = optics_measured_dbm(header, row)
    if not power_type or measured is None:
        return True
    if power_type == "RX":
        return measured < rx_min or measured > max_value
    if power_type == "TX":
        return measured < tx_min or measured > max_value
    return True


def signed_int(value: int) -> str:
    if value > 0:
        return f"+{value}"
    return str(value)


def progress(previous: int, current: int) -> str:
    if previous == 0:
        if current == 0:
            return "0.0%"
        return "N/A"
    pct = ((previous - current) / previous) * 100
    if abs(pct) < 0.05:
        return "0.0%"
    return f"{pct:+.1f}%"


def split_pipe_row(line: str) -> list[str] | None:
    stripped = line.strip()
    if not stripped.startswith("|"):
        return None
    cells = [cell.strip() for cell in stripped.strip("|").split("|")]
    if len(cells) < 2:
        return None
    return cells


def canonical_metric(value: str) -> str | None:
    mapping = {
        "lldp": "LLDP",
        "interface": "Interface",
        "interfacedown": "Interface",
        "opticstxrx": "Optics TX/RX",
        "optics": "Optics TX/RX",
        "prefec": "Pre Fec",
        "prefecber": "Pre Fec",
        "fecbin": "Fec Bin",
        "combinedfec": "Combined FEC",
        "t1t0opticsrxtx": "T1->T0 Optics RX/TX",
        "t1tot0opticsrxtx": "T1->T0 Optics RX/TX",
        "t1t0combinedfec": "T1->T0 Combined FEC",
        "t1tot0combinedfec": "T1->T0 Combined FEC",
        "total": "Total",
    }
    return mapping.get(normalize(value))


def discover_targets_from_previous(text: str) -> list[str]:
    targets: list[str] = []
    for line in text.splitlines():
        stripped = line.strip().upper().rstrip(":")
        if TARGET_RE.fullmatch(stripped) and stripped not in targets:
            targets.append(stripped)
    return targets


def first_metric_table_after(lines: Sequence[str], start: int) -> dict[str, int] | None:
    for idx in range(start, len(lines)):
        header = split_pipe_row(lines[idx])
        if not header or normalize(header[0]) != "metric":
            continue

        normalized_header = [normalize(cell) for cell in header]
        if "current" not in normalized_header:
            continue
        current_idx = normalized_header.index("current")

        result = metric_template()
        found = False
        for row_line in lines[idx + 1 :]:
            row = split_pipe_row(row_line)
            if row is None:
                if found:
                    break
                continue
            metric = canonical_metric(row[0])
            if metric is None:
                if found:
                    break
                continue
            if current_idx < len(row):
                result[metric] = parse_int(row[current_idx])
                found = True
        return result if found else None
    return None


def previous_current_metrics(previous_report: str, targets: Sequence[str]) -> dict[str, dict[str, int]]:
    lines = previous_report.splitlines()
    previous: dict[str, dict[str, int]] = {}

    for target in targets:
        for idx, line in enumerate(lines):
            if line.strip().upper().rstrip(":") != target:
                continue
            table = first_metric_table_after(lines, idx + 1)
            if table is not None:
                previous[target] = table
                break

    return previous


def sum_metrics(*items: Mapping[str, int]) -> dict[str, int]:
    total = metric_template()
    for item in items:
        for metric in ALL_METRICS:
            total[metric] += int(item.get(metric, 0))
    return total


def finish_totals(metrics: dict[str, int]) -> dict[str, int]:
    metrics["Total"] = sum(metrics.get(metric, 0) for metric in ISSUE_METRICS)
    return metrics


def table(headers: Sequence[object], rows: Sequence[Sequence[object]], left_columns: set[int] | None = None) -> str:
    left_columns = left_columns or set()
    string_rows = [[str(value) for value in row] for row in rows]
    header_cells = [str(header) for header in headers]
    widths = [
        max(len(header_cells[col]), *(len(row[col]) for row in string_rows))
        for col in range(len(header_cells))
    ]

    def border() -> str:
        return "+" + "+".join("-" * (width + 2) for width in widths) + "+"

    def format_row(row: Sequence[str], header: bool = False) -> str:
        cells = []
        for col, value in enumerate(row):
            if header and col in left_columns:
                rendered = value.ljust(widths[col])
            elif header:
                rendered = value.center(widths[col])
            elif col in left_columns:
                rendered = value.ljust(widths[col])
            else:
                rendered = value.rjust(widths[col])
            cells.append(f" {rendered} ")
        return "|" + "|".join(cells) + "|"

    lines = [border(), format_row(header_cells, header=True), border()]
    lines.extend(format_row(row) for row in string_rows)
    lines.append(border())
    return "\n".join(lines)


def metric_rows(previous: Mapping[str, int], current: Mapping[str, int]) -> list[list[object]]:
    rows: list[list[object]] = []
    for metric in (*ISSUE_METRICS, "Total"):
        prev = previous.get(metric, 0)
        cur = current.get(metric, 0)
        rows.append([metric, prev, cur, signed_int(cur - prev), progress(prev, cur)])
    return rows


def target_rows(
    previous: Mapping[str, Mapping[str, int]],
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
) -> list[list[object]]:
    rows: list[list[object]] = []
    for target in [*targets, TOTAL_TARGET]:
        prev = previous[target]["Total"]
        cur = current[target]["Total"]
        rows.append([target, prev, cur, signed_int(cur - prev), progress(prev, cur)])
    return rows


def human_list(values: Sequence[str]) -> str:
    items = list(values)
    if not items:
        return ""
    if len(items) == 1:
        return items[0]
    if len(items) == 2:
        return f"{items[0]} and {items[1]}"
    return f"{', '.join(items[:-1])}, and {items[-1]}"


def summarize_changes_rule_based(
    previous: Mapping[str, Mapping[str, int]],
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
) -> list[str]:
    deltas = {target: current[target]["Total"] - previous[target]["Total"] for target in targets}
    improved = [target for target, delta in deltas.items() if delta < 0]
    regressed = [target for target, delta in deltas.items() if delta > 0]
    unchanged = [target for target, delta in deltas.items() if delta == 0]

    lines: list[str] = []
    if improved and not regressed:
        lines.append("All targets improved.")
    elif regressed and not improved:
        lines.append("All targets regressed.")
    elif improved or regressed:
        parts = []
        if improved:
            parts.append(f"{human_list(improved)} improved")
        if regressed:
            parts.append(f"{human_list(regressed)} regressed")
        if unchanged:
            parts.append(f"{human_list(unchanged)} unchanged")
        lines.append("Mixed result: " + ", while ".join(parts) + ".")
    else:
        lines.append("No target total changed.")

    if improved:
        best_delta = min(deltas[target] for target in improved)
        best_targets = [target for target in improved if deltas[target] == best_delta]
        lines.append(f"Largest improvement: {human_list(best_targets)} decreased by {abs(best_delta)} issues.")
    if regressed:
        worst_delta = max(deltas[target] for target in regressed)
        worst_targets = [target for target in regressed if deltas[target] == worst_delta]
        lines.append(f"Largest regression: {human_list(worst_targets)} increased by {worst_delta} issues.")

    total_delta = current[TOTAL_TARGET]["Total"] - previous[TOTAL_TARGET]["Total"]
    if total_delta < 0:
        lines.append(f"Overall improvement: {total_delta}.")
    elif total_delta > 0:
        lines.append(f"Overall regression: +{total_delta}.")
    else:
        lines.append("Overall unchanged.")

    for metric in ("Combined FEC", "T1->T0 Optics RX/TX", "T1->T0 Combined FEC"):
        lines.append(
            f"{metric} changed overall: "
            f"{previous[TOTAL_TARGET][metric]} -> {current[TOTAL_TARGET][metric]}."
        )
    return lines


def insight_payload(
    previous: Mapping[str, Mapping[str, int]],
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
) -> dict[str, object]:
    return {
        "targets": [
            {
                "target": target,
                "previous_total": previous[target]["Total"],
                "current_total": current[target]["Total"],
                "delta": current[target]["Total"] - previous[target]["Total"],
            }
            for target in targets
        ],
        "overall": {
            "previous_total": previous[TOTAL_TARGET]["Total"],
            "current_total": current[TOTAL_TARGET]["Total"],
            "delta": current[TOTAL_TARGET]["Total"] - previous[TOTAL_TARGET]["Total"],
        },
        "metric_changes": [
            {
                "metric": metric,
                "previous": previous[TOTAL_TARGET][metric],
                "current": current[TOTAL_TARGET][metric],
                "delta": current[TOTAL_TARGET][metric] - previous[TOTAL_TARGET][metric],
            }
            for metric in ("Combined FEC", "T1->T0 Optics RX/TX", "T1->T0 Combined FEC")
        ],
    }


def insights_api_key() -> str | None:
    return (
        os.environ.get("PHX_DG_INSIGHTS_API_KEY")
        or os.environ.get("OPENAI_API_KEY")
        or os.environ.get("AIDER_API_KEY")
    )


def normalize_insights_provider(provider: str) -> str:
    normalized = provider.strip().lower()
    aliases = {
        "aider": "openai",
        "chatgpt": "openai",
        "codex": "openai",
        "none": "off",
        "rule": "off",
        "rule-based": "off",
    }
    return aliases.get(normalized, normalized)


def parse_response_output_text(data: Mapping[str, object]) -> str:
    output_text = data.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text

    parts: list[str] = []
    output = data.get("output")
    if isinstance(output, list):
        for item in output:
            if not isinstance(item, Mapping):
                continue
            content = item.get("content")
            if not isinstance(content, list):
                continue
            for part in content:
                if not isinstance(part, Mapping):
                    continue
                text = part.get("text")
                if isinstance(text, str):
                    parts.append(text)
    return "\n".join(parts)


def strip_markdown_code_fence(text: str) -> str:
    stripped = text.strip()
    if not stripped.startswith("```"):
        return stripped
    stripped = re.sub(r"^```(?:json)?\s*", "", stripped, flags=re.I)
    stripped = re.sub(r"\s*```$", "", stripped)
    return stripped.strip()


def json_insight_lines(text: str) -> list[str]:
    candidates = [strip_markdown_code_fence(text)]
    start = text.find("[")
    end = text.rfind("]")
    if start != -1 and end != -1 and end > start:
        candidates.append(text[start : end + 1])

    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(parsed, Mapping):
            parsed = parsed.get("insights")
        if isinstance(parsed, list) and all(isinstance(item, str) for item in parsed):
            return [item.strip() for item in parsed if item.strip()]
    return []


def split_sentence_insight_lines(text: str) -> list[str]:
    sentence_parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text.strip())
    return [part.strip() for part in sentence_parts if part.strip()]


def clean_insight_lines(text: str) -> list[str]:
    json_lines = json_insight_lines(text)
    if json_lines:
        return json_lines

    split_text = re.sub(r"\s+(?=[1-6][.)]\s+)", "\n", text.strip())
    lines: list[str] = []
    for raw_line in split_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^(?:[-*]\s*|[1-6][.)]\s*)", "", line).strip()
        if line:
            lines.append(line)
    if len(lines) == 1:
        sentence_lines = split_sentence_insight_lines(lines[0])
        if len(sentence_lines) > 1:
            return sentence_lines
    return lines


def call_openai_insights(
    payload: Mapping[str, object],
    model: str,
    base_url: str,
    timeout: float,
) -> list[str]:
    api_key = insights_api_key()
    if not api_key:
        raise RuntimeError(
            "set PHX_DG_INSIGHTS_API_KEY, OPENAI_API_KEY, or AIDER_API_KEY"
        )

    request_body = {
        "model": model,
        "max_output_tokens": 300,
        "input": [
            {
                "role": "developer",
                "content": (
                    "You analyze PHX DG health-check report deltas. Return only a JSON array "
                    "containing exactly six concise strings. Do not wrap it in markdown. Preserve "
                    "exact metric names and counts from the JSON. Item 1: target trend. Item 2: "
                    "largest target movement. Item 3: overall total movement. Items 4-6: Combined "
                    "FEC, T1->T0 Optics RX/TX, and T1->T0 Combined FEC."
                ),
            },
            {
                "role": "user",
                "content": json.dumps(payload, sort_keys=True),
            },
        ],
    }
    request = urllib.request.Request(
        f"{base_url.rstrip('/')}/responses",
        data=json.dumps(request_body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            response_body = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise RuntimeError(f"insights API HTTP {exc.code}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"insights API request failed: {exc.reason}") from exc

    data = json.loads(response_body)
    lines = clean_insight_lines(parse_response_output_text(data))
    if len(lines) < 6:
        raise RuntimeError(f"insights API returned {len(lines)} parseable lines; expected 6")
    if len(lines) > 6:
        return lines[:6]
    return lines


def summarize_changes(
    previous: Mapping[str, Mapping[str, int]],
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
    insights_provider: str = DEFAULT_INSIGHTS_PROVIDER,
    insights_model: str = DEFAULT_INSIGHTS_MODEL,
    insights_base_url: str = DEFAULT_INSIGHTS_BASE_URL,
    insights_timeout: float = DEFAULT_INSIGHTS_TIMEOUT,
) -> list[str]:
    provider = normalize_insights_provider(insights_provider)
    if provider == "off":
        return summarize_changes_rule_based(previous, current, targets)
    if provider not in {"auto", "openai"}:
        raise ValueError(f"Unknown insights provider {insights_provider!r}")
    if provider == "auto" and not insights_api_key():
        return summarize_changes_rule_based(previous, current, targets)

    payload = insight_payload(previous, current, targets)
    fallback_lines = summarize_changes_rule_based(previous, current, targets)
    try:
        return call_openai_insights(payload, insights_model, insights_base_url, insights_timeout)
    except Exception as exc:
        print(f"AI insight generation unavailable ({exc}); using rule-based fallback.", file=sys.stderr)
        return fallback_lines


def index_by_normalized_header(header: Sequence[object]) -> dict[str, int]:
    indexes: dict[str, int] = {}
    for idx, value in enumerate(header):
        key = normalize(value)
        if key and key not in indexes:
            indexes[key] = idx
    return indexes


def find_index(header: Sequence[object], candidates: Sequence[str]) -> int | None:
    indexes = index_by_normalized_header(header)
    for candidate in candidates:
        idx = indexes.get(normalize(candidate))
        if idx is not None:
            return idx
    return None


def cell(header: Sequence[object], row: Sequence[object], candidates: Sequence[str]) -> object | None:
    idx = find_index(header, candidates)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def iter_nonempty_rows(workbook: object, sheet_name: str) -> Iterable[tuple[list[object], tuple[object, ...]]]:
    if sheet_name not in workbook.sheetnames:
        return
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return
    for row in rows:
        if any(value not in (None, "") for value in row):
            yield header, row


def row_dict(header: Sequence[object], row: Sequence[object]) -> dict[str, object]:
    result: dict[str, object] = {}
    for idx, raw_name in enumerate(header):
        name = str(raw_name or "").strip() or f"Column {idx + 1}"
        if name in result:
            name = f"{name} {idx + 1}"
        result[name] = row[idx] if idx < len(row) else None
    return result


def detail_row(
    target: str,
    path: Path,
    sheet_name: str,
    issue_type: str,
    header: Sequence[object],
    row: Sequence[object],
    counts_fec_bin: int = 0,
    counts_pre_fec: int = 0,
) -> dict[str, object]:
    detail: dict[str, object] = {
        "Source File": path.name,
        "Source Sheet": sheet_name,
        "DG": target,
        "Issue Type": issue_type,
        "Counts Fec Bin": counts_fec_bin,
        "Counts Pre Fec": counts_pre_fec,
    }
    for key, value in row_dict(header, row).items():
        out_key = key if key not in detail else f"Original {key}"
        detail[out_key] = value
    return detail


def is_lldp_interface_down(header: Sequence[object], row: Sequence[object]) -> bool:
    values = [
        cell(header, row, ("Act. Interface", "Active Interface", "Interface State", "Status")),
        cell(header, row, ("failed_tests", "message", "Failure")),
    ]
    combined = " ".join(str(value or "") for value in values).lower()
    return "interfacedown" in combined or "interface down" in combined


def lock_status_has_fec_bin(lock_status: object) -> bool:
    text = str(lock_status or "")
    matched_explicit_count = False
    for match in FEC_BIN_COUNT_RE.finditer(text):
        matched_explicit_count = True
        bin_num = int(match.group(1))
        count = int(match.group(2))
        if FEC_BIN_MIN <= bin_num <= FEC_BIN_MAX and count > 0:
            return True
    if matched_explicit_count:
        return False

    return any(
        FEC_BIN_MIN <= int(match.group(1)) <= FEC_BIN_MAX
        for match in FEC_BIN_RE.finditer(text)
    )


def row_counts_pre_fec(header: Sequence[object], row: Sequence[object]) -> bool:
    pre_fec_ber = cell(
        header,
        row,
        (
            "Pre-FEC BER",
            "Pre FEC BER",
            "Pre_FEC_BER",
            "RAW BER",
            "RAW_BER",
            "RAW_BER_MAX",
        ),
    )
    parsed = parse_float(pre_fec_ber)
    return parsed is not None and parsed > PRE_FEC_BER_THRESHOLD


def is_gpu_remote_host_row(header: Sequence[object], row: Sequence[object]) -> bool:
    for column in GPU_REMOTE_HOST_COLUMNS:
        value = cell(header, row, (column,))
        if value is not None and "gpu" in str(value).lower():
            return True
    return False


def target_from_placement_group(value: object) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"\d+", text)
    if not match:
        return None
    return f"DG{int(match.group(0))}"


def target_for_row(
    header: Sequence[object],
    row: Sequence[object],
    site_tag: str | None = None,
    fallback_target: str | None = None,
) -> str | None:
    placement_group = cell(header, row, ("Placement Group", "placement_group", "PG"))
    target = target_from_placement_group(placement_group)
    if target and (target in dg_racks_for_site(site_tag) or target == fallback_target):
        return target

    rack = cell(header, row, ("Rack", "rack_a", "DeviceA_Rack"))
    if rack is not None:
        target = rack_to_dg_for_site(site_tag).get(str(rack).strip())
        if target:
            return target

    return fallback_target


def row_matches_target(
    header: Sequence[object],
    row: Sequence[object],
    target: str,
    site_tag: str | None = None,
    fallback_target: str | None = None,
) -> bool:
    return target_for_row(header, row, site_tag, fallback_target) == target


def unique_sheet_rows(
    workbook: object,
    sheet_name: str,
    key_columns: Sequence[str],
) -> list[tuple[list[object], tuple[object, ...]]]:
    unique: dict[tuple[object, ...], tuple[list[object], tuple[object, ...]]] = {}
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        key = tuple(cell(header, row, (column,)) for column in key_columns)
        unique.setdefault(key, (header, row))
    return list(unique.values())


def combined_fec_rows(workbook: object) -> list[tuple[str, list[object], tuple[object, ...]]]:
    source_sheet = "combined_fec_with_pp"
    if source_sheet not in workbook.sheetnames:
        return []

    rows: list[tuple[str, list[object], tuple[object, ...]]] = []
    for header, row in iter_nonempty_rows(workbook, source_sheet):
        if is_gpu_remote_host_row(header, row):
            continue
        rows.append((source_sheet, header, row))
    return rows


def read_workbook_metrics(
    path: Path,
    target: str,
    site_tag: str | None = None,
    relax: bool = False,
    relax_ipr: bool = False,
) -> tuple[dict[str, int], list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Workbook mode requires openpyxl. Install it with: python3 -m pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    metrics = metric_template()
    details: list[dict[str, object]] = []

    sheet_name = "full_path_lldp_with_int_down"
    seen_lldp: set[tuple[object, ...]] = set()
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        if is_lldp_interface_down(header, row):
            continue
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        key = tuple(
            cell(header, row, (column,))
            for column in (
                "Hostname",
                "Interface",
                "Rack",
                "Elevation",
                "Expected Hostname",
                "Exp. Interface",
            )
        )
        if key in seen_lldp:
            continue
        seen_lldp.add(key)
        metrics["LLDP"] += 1
        details.append(detail_row(target, path, sheet_name, "LLDP", header, row))

    sheet_name = "interfaces_sp_with_pp"
    seen_interfaces: set[tuple[object, ...]] = set()
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        key = tuple(
            cell(header, row, (column,))
            for column in (
                "Hostname",
                "Interface",
                "Rack",
                "Elevation",
                "Z_end_host",
                "Z_end_intf",
                "Placement Group",
            )
        )
        if key in seen_interfaces:
            continue
        seen_interfaces.add(key)
        metrics["Interface"] += 1
        details.append(detail_row(target, path, sheet_name, "Interface", header, row))

    sheet_name = "optics_rx_tx_threshold_with_pp"
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        if not optics_row_passes_relaxed_threshold(header, row, relax=relax, relax_ipr=relax_ipr):
            continue
        metrics["Optics TX/RX"] += 1
        details.append(detail_row(target, path, sheet_name, "Optics TX/RX", header, row))

    for sheet_name, header, row in combined_fec_rows(workbook):
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        metrics["Combined FEC"] += 1
        lock_status = cell(header, row, ("Lock Status", "Lock_status", "LockStatus"))
        counts_fec_bin = 0
        counts_pre_fec = 0
        if lock_status_has_fec_bin(lock_status):
            metrics["Fec Bin"] += 1
            counts_fec_bin = 1
        if row_counts_pre_fec(header, row):
            metrics["Pre Fec"] += 1
            counts_pre_fec = 1
        details.append(
            detail_row(
                target,
                path,
                sheet_name,
                "Combined FEC",
                header,
                row,
                counts_fec_bin,
                counts_pre_fec,
            )
        )

    sheet_name = "t1_t0_optics_rx_tx"
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        if not optics_row_passes_relaxed_threshold(header, row, relax=relax, relax_ipr=relax_ipr):
            continue
        metrics["T1->T0 Optics RX/TX"] += 1
        details.append(detail_row(target, path, sheet_name, "T1->T0 Optics RX/TX", header, row))

    sheet_name = "t1_t0_combined_fec"
    for header, row in iter_nonempty_rows(workbook, sheet_name):
        if not row_matches_target(header, row, target, site_tag, target):
            continue
        metrics["T1->T0 Combined FEC"] += 1
        details.append(detail_row(target, path, sheet_name, "T1->T0 Combined FEC", header, row))

    return finish_totals(metrics), details


def discover_workbooks(data_dir: Path) -> list[Path]:
    return sorted(data_dir.expanduser().glob("*_full_report.xlsx"))


def parse_excel_args(excel_args: Sequence[str] | None, data_dir: Path) -> list[tuple[str | None, Path]]:
    if not excel_args:
        return [(None, path) for path in discover_workbooks(data_dir)]

    parsed: list[tuple[str | None, Path]] = []
    for raw_item in excel_args:
        if "=" in raw_item:
            raw_target, raw_path = raw_item.split("=", 1)
            target = normalize_target(raw_target)
        else:
            target = None
            raw_path = raw_item
        path = Path(raw_path).expanduser()
        if not path.is_absolute():
            path = data_dir / path
        parsed.append((target, path))
    return parsed


def workbook_target_pairs(excel_args: Sequence[str] | None, targets: Sequence[str] | None, data_dir: Path) -> list[tuple[str, Path]]:
    parsed = parse_excel_args(excel_args, data_dir)
    if not parsed:
        raise FileNotFoundError(f"No *_full_report.xlsx files found in {data_dir}")

    explicit_targets = [target for target, _path in parsed if target is not None]
    if explicit_targets and len(explicit_targets) != len(parsed):
        raise ValueError("Use either DG=path for every --excel value or no DG= prefix values")

    if explicit_targets:
        pairs = [(target or "", path) for target, path in parsed]
    else:
        if targets and len(parsed) == 1:
            pairs = [(target, parsed[0][1]) for target in targets]
        elif targets and len(targets) != len(parsed):
            raise ValueError(
                f"Target count ({len(targets)}) does not match workbook count ({len(parsed)}); "
                "use one combined workbook, one workbook per target, or DG=path for explicit mapping"
            )
        else:
            inferred_targets = list(targets) if targets else [f"DG{idx}" for idx in range(1, len(parsed) + 1)]
            pairs = list(zip(inferred_targets, [path for _target, path in parsed]))

    for _target, path in pairs:
        if not path.exists():
            raise FileNotFoundError(path)
    return pairs


def read_current_metrics(
    pairs: Sequence[tuple[str, Path]],
    site_tag: str | None = None,
    relax: bool = False,
    relax_ipr: bool = False,
) -> tuple[dict[str, dict[str, int]], list[dict[str, object]]]:
    current: dict[str, dict[str, int]] = {}
    details: list[dict[str, object]] = []
    for target, path in pairs:
        metrics, workbook_details = read_workbook_metrics(path, target, site_tag, relax, relax_ipr)
        current[target] = metrics
        details.extend(workbook_details)
    current[TOTAL_TARGET] = sum_metrics(*(current[target] for target, _path in pairs))
    return current, details


def resolve_previous_path(path: Path | None, data_dir: Path) -> Path | None:
    if path is None:
        return None
    if str(path) == "-":
        return None
    resolved = path.expanduser()
    if not resolved.is_absolute():
        resolved = data_dir / resolved
    return resolved


def read_previous(path: Path | None, data_dir: Path) -> str:
    if path is None:
        return ""
    if str(path) == "-":
        return sys.stdin.read()
    resolved = resolve_previous_path(path, data_dir)
    if resolved is None:
        return ""
    if not resolved.exists():
        return ""
    return resolved.read_text(encoding="utf-8")


def write_report(path: Path, report: str) -> Path:
    resolved = path.expanduser()
    resolved.parent.mkdir(parents=True, exist_ok=True)
    resolved.write_text(report + "\n", encoding="utf-8")
    return resolved


def unique_columns(rows: Sequence[Mapping[str, object]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", name).strip() or "Sheet"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def write_detail_sheet(
    workbook: object,
    name: str,
    rows: Sequence[Mapping[str, object]],
    used: set[str],
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_name(name, used))
    if not rows:
        worksheet.append(["No rows"])
        return
    columns = unique_columns(rows)
    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column) for column in columns])


def write_excel_report(
    path: Path,
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
    details: Sequence[Mapping[str, object]],
    site_tag: str | None = None,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("Writing workbook output requires openpyxl. Install it with: python3 -m pip install openpyxl") from exc

    output_path = path.expanduser()
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    used = {summary.title}
    dg_racks = dg_racks_for_site(site_tag)
    summary.append(["DG", "Racks", *ISSUE_METRICS, "Total"])
    for target in targets:
        values = current.get(target, {})
        summary.append(
            [
                target,
                ",".join(dg_racks.get(target, ())),
                *(values.get(metric, 0) for metric in ISSUE_METRICS),
                values.get("Total", 0),
            ]
        )
    total_values = sum_metrics(*(current.get(target, metric_template()) for target in targets))
    summary.append(
        [
            TOTAL_TARGET,
            "",
            *(total_values.get(metric, 0) for metric in ISSUE_METRICS),
            total_values.get("Total", 0),
        ]
    )

    write_detail_sheet(
        wb,
        "DG Racks",
        [{"Site": normalize_site_tag(site_tag), "DG": target, "Racks": ",".join(racks)} for target, racks in dg_racks.items()],
        used,
    )
    write_detail_sheet(wb, "All", details, used)
    write_detail_sheet(wb, "LLDP", [row for row in details if row.get("Issue Type") == "LLDP"], used)
    write_detail_sheet(wb, "Interface", [row for row in details if row.get("Issue Type") == "Interface"], used)
    write_detail_sheet(wb, "Optics TX/RX", [row for row in details if row.get("Issue Type") == "Optics TX/RX"], used)
    write_detail_sheet(wb, "Pre Fec", [row for row in details if row.get("Counts Pre Fec") == 1], used)
    write_detail_sheet(wb, "Fec Bin", [row for row in details if row.get("Counts Fec Bin") == 1], used)
    write_detail_sheet(wb, "Combined FEC", [row for row in details if row.get("Issue Type") == "Combined FEC"], used)

    for target in targets:
        write_detail_sheet(wb, target, [row for row in details if row.get("DG") == target], used)
        write_detail_sheet(
            wb,
            f"{target}_combined_fec",
            [
                row
                for row in details
                if row.get("DG") == target and row.get("Issue Type") == "Combined FEC"
            ],
            used,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def same_path(left: Path | None, right: Path | None) -> bool:
    if left is None or right is None:
        return False
    try:
        return left.expanduser().resolve() == right.expanduser().resolve()
    except FileNotFoundError:
        return left.expanduser().absolute() == right.expanduser().absolute()


def build_report(
    previous: Mapping[str, Mapping[str, int]],
    current: Mapping[str, Mapping[str, int]],
    targets: Sequence[str],
    insights_provider: str = DEFAULT_INSIGHTS_PROVIDER,
    insights_model: str = DEFAULT_INSIGHTS_MODEL,
    insights_base_url: str = DEFAULT_INSIGHTS_BASE_URL,
    insights_timeout: float = DEFAULT_INSIGHTS_TIMEOUT,
) -> str:
    sections = [
        "Target Total Summary",
        "",
        table(TARGET_HEADERS, target_rows(previous, current, targets), left_columns={0}),
        "",
        *summarize_changes(
            previous,
            current,
            targets,
            insights_provider,
            insights_model,
            insights_base_url,
            insights_timeout,
        ),
        "",
        "Issue Breakdown By Target",
    ]

    for target in targets:
        sections.extend(
            [
                "",
                target,
                "",
                table(COMPARISON_HEADERS, metric_rows(previous[target], current[target]), left_columns={0}),
            ]
        )

    sections.extend(
        [
            "",
            TOTAL_TARGET,
            "",
            table(COMPARISON_HEADERS, metric_rows(previous[TOTAL_TARGET], current[TOTAL_TARGET]), left_columns={0}),
        ]
    )
    return "\n".join(sections)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a PHX DG progress report from qcli *_full_report.xlsx workbooks."
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=DEFAULT_DATA_DIR,
        help=f"Directory to scan for *_full_report.xlsx files. Default: {DEFAULT_DATA_DIR}",
    )
    parser.add_argument(
        "--excel",
        action="append",
        help=(
            "Workbook input. Use a bare path to map files in order, or DG=path for explicit mapping. "
            "Repeat for multiple workbooks. Defaults to all *_full_report.xlsx files in --data-dir."
        ),
    )
    parser.add_argument(
        "--site-tag",
        "--site",
        "--tag",
        choices=sorted(SITE_CONFIGS),
        default=DEFAULT_SITE_TAG,
        help=f"PHX site/build tag for rack mapping and qcli generation. Default: {DEFAULT_SITE_TAG}.",
    )
    parser.add_argument(
        "--targets",
        help=(
            "Comma-separated target names for bare --excel paths, for example DG1,DG2,DG3,DG4,DG5. "
            "Use 'all' for every target in the selected site rack map, or every supported qcli target "
            "when used with --print-qcli-command."
        ),
    )
    parser.add_argument(
        "--print-qcli-command",
        action="store_true",
        help="Print the combined qcli hc-summary command for --targets, then exit.",
    )
    parser.add_argument(
        "--relax",
        action="store_true",
        help="Relax non-IPR optics RX/TX min/max thresholds by 1 dBm.",
    )
    parser.add_argument(
        "--relax-ipr",
        action="store_true",
        help="Relax IPR optics thresholds to RX -5..3 dBm and TX -3..3 dBm.",
    )
    parser.add_argument(
        "--previous",
        nargs="?",
        const=DEFAULT_PREVIOUS_REPORT,
        type=Path,
        help=(
            "Previous formatted report to compare against. If --previous is given without a path, "
            f"uses {DEFAULT_PREVIOUS_REPORT}. Use --previous - to read from stdin."
        ),
    )
    parser.add_argument("--output", type=Path, help="Write the generated report to this file.")
    parser.add_argument("--output-excel", type=Path, help="Write grouped workbook rows and split FEC sheets to this Excel file.")
    parser.add_argument(
        "--insights-provider",
        default=DEFAULT_INSIGHTS_PROVIDER,
        help=(
            "Insight generator: auto, openai, aider, chatgpt, codex, or off. "
            "Aliases aider/chatgpt/codex use the OpenAI-compatible Responses API. "
            f"Default: {DEFAULT_INSIGHTS_PROVIDER}."
        ),
    )
    parser.add_argument(
        "--insights-model",
        default=DEFAULT_INSIGHTS_MODEL,
        help=f"Model used for API-backed insight generation. Default: {DEFAULT_INSIGHTS_MODEL}.",
    )
    parser.add_argument(
        "--insights-base-url",
        default=DEFAULT_INSIGHTS_BASE_URL,
        help=f"OpenAI-compatible API base URL. Default: {DEFAULT_INSIGHTS_BASE_URL}.",
    )
    parser.add_argument(
        "--insights-timeout",
        type=float,
        default=DEFAULT_INSIGHTS_TIMEOUT,
        help=f"Insight API request timeout in seconds. Default: {DEFAULT_INSIGHTS_TIMEOUT:g}.",
    )
    parser.add_argument("--debug", action="store_true", help="Print workbook-to-DG mapping to stderr.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    data_dir = args.data_dir.expanduser()
    site_tag = normalize_site_tag(args.site_tag)
    qcli_all_targets = qcli_targets_for_site(site_tag) if args.print_qcli_command else None
    requested_targets = parse_target_list(args.targets, site_tag, qcli_all_targets)

    previous_text = read_previous(args.previous, data_dir)
    if requested_targets is None and previous_text.strip():
        previous_targets = discover_targets_from_previous(previous_text)
        requested_targets = previous_targets or None

    if args.print_qcli_command:
        if requested_targets is None:
            raise ValueError("--targets is required unless --previous provides DG sections")
        print(combined_qcli_command(requested_targets, site_tag, args.relax))
        return 0

    pairs = workbook_target_pairs(args.excel, requested_targets, data_dir)
    targets = [target for target, _path in pairs]

    if args.debug:
        for target, path in pairs:
            print(f"{target}: {path}", file=sys.stderr)

    current, details = read_current_metrics(pairs, site_tag, args.relax, args.relax_ipr)
    previous = previous_current_metrics(previous_text, targets) if previous_text.strip() else {}
    for target in targets:
        previous.setdefault(target, dict(current[target]))
    previous[TOTAL_TARGET] = sum_metrics(*(previous[target] for target in targets))

    report = build_report(
        previous,
        current,
        targets,
        args.insights_provider,
        args.insights_model,
        args.insights_base_url,
        args.insights_timeout,
    )
    print(report)

    written_output: Path | None = None
    if args.output:
        output = args.output.expanduser()
        written_output = write_report(output, report)
        print(f"\nReport written to: {written_output}", file=sys.stderr)

    if args.output_excel:
        written_excel = write_excel_report(args.output_excel, current, targets, details, site_tag)
        print(f"\nExcel report written to: {written_excel}", file=sys.stderr)

    previous_path = resolve_previous_path(args.previous, data_dir)
    if previous_path and not same_path(previous_path, written_output):
        written_previous = write_report(previous_path, report)
        print(f"\nPrevious report updated: {written_previous}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
