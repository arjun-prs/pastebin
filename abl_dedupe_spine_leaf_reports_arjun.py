#!/usr/bin/env python3

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


INPUT_VALIDATION_SHEET = "Validation Errors"
VALIDATION_SHEET = "Unique Interface and Optics"
LLDP_SHEET = "LLDP Mismatch"
INTERFACE_DUPLICATES_SHEET = "Interface Down Dups"
OPTICS_DUPLICATES_SHEET = "Optics Issues Dups"
LLDP_DUPLICATES_SHEET = "LLDP Dups"
RACK_TOTAL_SHEET = "Rack Total"
PIVOT_SHEET = "Pivot Table"
REPORT_CATEGORIES = ["Interface Down", "Optics_Issues", "LLDP"]
DEDUPED_DUPLICATE_CATEGORIES = {"Interface Down", "LLDP"}
DISCOVERY_WINDOW_SECONDS = 60 * 60
DEFAULT_REPORT_TAG = "ABL18"
REPORT_PARENT_DIRS = (
    Path("~/tools/ROMA/roma-deployment-scripts/reports").expanduser(),
    Path("~/tools/roma/roma-deployment-scripts/reports").expanduser(),
)
REPORT_TAGS = {
    "ABL18": {
        "fabric_number": "18",
        "repository_subdir": Path("ABL18"),
        "search_subdir": Path("ABL18") / "Bleaf",
        "leaf_search_subdirs": (Path("."),),
        "spine_search_subdirs": (Path("."),),
        "default_leaf_dh_number": "3",
        "both_latest_leaf_dh_numbers": ("3", "4"),
        "spine_dh_numbers": ("1", "2"),
    },
    "ABL19": {
        "fabric_number": "19",
        "repository_subdir": Path("ABL19") / "Bleaf",
        "search_subdir": Path("ABL19") / "Bleaf",
        "leaf_search_subdirs": (Path("."),),
        "spine_search_subdirs": (Path("."),),
        "default_leaf_dh_number": "1",
        "both_latest_leaf_dh_numbers": ("1", "2"),
        "spine_dh_numbers": ("1", "2"),
    },
}
ACTIVE_REPORT_TAG = DEFAULT_REPORT_TAG
REPORT_COLUMNS = [
    "device_rack",
    "device_elevation",
    "spine_dh",
    "hostname",
    "interface",
    "remote_device",
    "remote_interface",
    "remote_device_elevation",
    "remote_device_rack",
    "bleaf_pod",
    "link_failure_details",
    "oper_status",
    "actions_taken",
    "category",
]


class ColorArgumentParser(argparse.ArgumentParser):
    def error(self, message):
        self.print_usage(sys.stderr)
        error_text = f"{self.prog}: error: {message}\n"
        if sys.stderr.isatty():
            error_text = f"\033[91m{error_text}\033[0m"
        self.exit(2, error_text)


def normalize_report_tag(value):
    tag = str(value).strip().upper()
    if tag not in REPORT_TAGS:
        choices = ", ".join(sorted(REPORT_TAGS))
        raise argparse.ArgumentTypeError(
            f"unsupported report tag {value!r}; choose one of: {choices}"
        )
    return tag


def set_report_tag(report_tag):
    global ACTIVE_REPORT_TAG
    ACTIVE_REPORT_TAG = normalize_report_tag(report_tag)


def report_tag():
    return ACTIVE_REPORT_TAG


def report_tag_config():
    return REPORT_TAGS[ACTIVE_REPORT_TAG]


def report_fabric_number():
    return report_tag_config()["fabric_number"]


def default_leaf_dh_number():
    return report_tag_config()["default_leaf_dh_number"]


def both_latest_leaf_dh_numbers():
    return report_tag_config()["both_latest_leaf_dh_numbers"]


def spine_dh_numbers():
    return report_tag_config()["spine_dh_numbers"]


def script_dir():
    return Path(__file__).resolve().parent


def report_parent_dir():
    for path in REPORT_PARENT_DIRS:
        if path.is_dir():
            return path
    return REPORT_PARENT_DIRS[0]


def report_repository_dir():
    return report_parent_dir() / report_tag_config()["repository_subdir"]


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def normalize_interface(value):
    interface = normalize_text(value)
    interface = interface.replace("ethernet", "et")
    interface = re.sub(r"\s+", "", interface)
    return interface


def normalize_detail_text(value):
    return normalize_text(value)


def canonical_link_key(row):
    category = normalize_text(row.get("category"))
    endpoint_a = (
        normalize_text(row.get("hostname")),
        normalize_interface(row.get("interface")),
    )
    endpoint_b = (
        normalize_text(row.get("remote_device")),
        normalize_interface(row.get("remote_interface")),
    )
    left, right = sorted([endpoint_a, endpoint_b])
    key_parts = [
        category,
        left[0],
        left[1],
        right[0],
        right[1],
    ]
    return tuple(key_parts)


def extract_bleaf_pod(row):
    hostname = normalize_text(row.get("hostname"))
    remote_device = normalize_text(row.get("remote_device"))

    if "bleaf" in hostname:
        pod_value = row.get("POD")
        if not pd.isna(pod_value):
            return pod_value

    if "bleaf" in remote_device:
        pod_value = row.get("remote_device_pod_number")
        if not pd.isna(pod_value):
            return pod_value

    return pd.NA


def extract_dh_label(report_path):
    dh_number = extract_dh_number(report_path)
    if dh_number == "UNKNOWN":
        return pd.NA
    return f"DH{dh_number}"


def normalize_dh_number(value):
    match = re.fullmatch(r"(?:DH)?(\d+)", str(value).strip(), re.IGNORECASE)
    if not match:
        raise argparse.ArgumentTypeError(
            f"invalid datahall {value!r}; expected a number like 2 or a label like DH2"
        )
    return match.group(1)


def load_report(path, source_side):
    dataframe = pd.read_excel(path, sheet_name=INPUT_VALIDATION_SHEET)
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    dataframe["bleaf_pod"] = dataframe.apply(extract_bleaf_pod, axis=1)
    dataframe["spine_dh"] = extract_dh_label(path) if source_side == "spine" else pd.NA

    for column in REPORT_COLUMNS:
        if column not in dataframe.columns:
            dataframe[column] = pd.NA

    dataframe = dataframe[REPORT_COLUMNS].copy()
    dataframe["_source_side"] = source_side
    dataframe["_source_report"] = Path(path).name
    dataframe["_canonical_link_key"] = dataframe.apply(canonical_link_key, axis=1)
    return dataframe


def choose_row_to_keep(group):
    side_priority = {"leaf": 0, "spine": 1}
    ordered = group.copy()
    ordered["_side_priority"] = ordered["_source_side"].map(side_priority).fillna(9)
    ordered = ordered.sort_values(
        by=["_side_priority", "_source_report", "hostname", "interface"],
        kind="stable",
    )
    return ordered.iloc[0]


def is_spine_to_leaf_row(row):
    hostname = normalize_text(row.get("hostname"))
    remote_device = normalize_text(row.get("remote_device"))
    return "bspine" in hostname and "bleaf" in remote_device


def orient_row_to_leaf_side(row):
    oriented = row.copy()
    if not is_spine_to_leaf_row(oriented):
        return oriented

    oriented["device_rack"], oriented["remote_device_rack"] = (
        oriented.get("remote_device_rack"),
        oriented.get("device_rack"),
    )
    oriented["device_elevation"], oriented["remote_device_elevation"] = (
        oriented.get("remote_device_elevation"),
        oriented.get("device_elevation"),
    )
    oriented["hostname"], oriented["remote_device"] = (
        oriented.get("remote_device"),
        oriented.get("hostname"),
    )
    oriented["interface"], oriented["remote_interface"] = (
        oriented.get("remote_interface"),
        oriented.get("interface"),
    )
    return oriented


def build_rack_total(dataframe):
    rack_totals = (
        dataframe.groupby("device_rack", dropna=False)
        .size()
        .reset_index(name="total_count")
    )

    rack_categories = (
        dataframe.pivot_table(
            index="device_rack",
            columns="category",
            values="link_failure_details",
            aggfunc="count",
            fill_value=0,
        )
        .reindex(columns=REPORT_CATEGORIES, fill_value=0)
        .reset_index()
    )

    rack_totals_combined = pd.merge(
        rack_totals,
        rack_categories,
        on="device_rack",
        how="left",
    ).fillna(0)

    rack_totals_combined["total_count"] = pd.to_numeric(
        rack_totals_combined["total_count"], errors="coerce"
    ).fillna(0).astype(int)

    rack_totals_combined = rack_totals_combined.sort_values(
        by="total_count",
        ascending=False,
        kind="stable",
    ).reset_index(drop=True)

    rack_totals_combined["priority_comment"] = ""
    rack_totals_combined.loc[:9, "priority_comment"] = "High priority rack"
    return rack_totals_combined


def build_pivot_table(dataframe):
    final_table = pd.pivot_table(
        dataframe,
        values="link_failure_details",
        index=["device_rack", "category"],
        aggfunc="count",
    )

    grand_total = final_table.values.sum()
    grand_total_row = pd.DataFrame(
        [[grand_total]],
        columns=final_table.columns,
        index=pd.MultiIndex.from_tuples(
            [("Grand Total", "")],
            names=final_table.index.names,
        ),
    )
    return pd.concat([final_table, grand_total_row])


def build_duplicate_matches(combined):
    duplicate_frames = []
    duplicate_group_id = 1

    for _, group in combined.groupby("_canonical_link_key", sort=False):
        if len(group) < 2:
            continue

        category = str(group.iloc[0]["category"]).strip()
        if category not in DEDUPED_DUPLICATE_CATEGORIES:
            continue

        duplicate_group = group.copy()
        duplicate_group["duplicate_group_id"] = duplicate_group_id
        duplicate_group["duplicate_count"] = len(duplicate_group)
        duplicate_group["duplicate_error_type"] = category
        duplicate_frames.append(duplicate_group)
        duplicate_group_id += 1

    if not duplicate_frames:
        return pd.DataFrame(
            columns=[
                "duplicate_group_id",
                "duplicate_count",
                "duplicate_error_type",
                "source_side",
                "source_report",
                *REPORT_COLUMNS,
            ]
        )

    duplicates_df = pd.concat(duplicate_frames, ignore_index=True, sort=False)
    duplicates_df = duplicates_df.rename(
        columns={
            "_source_side": "source_side",
            "_source_report": "source_report",
        }
    )
    duplicates_df = duplicates_df.drop(columns=["_canonical_link_key"], errors="ignore")
    duplicates_df = duplicates_df[
        [
            "duplicate_group_id",
            "duplicate_count",
            "duplicate_error_type",
            "source_side",
            "source_report",
            *REPORT_COLUMNS,
        ]
    ]
    duplicates_df = duplicates_df.sort_values(
        by=["duplicate_group_id", "source_side", "source_report", "hostname", "interface"],
        kind="stable",
    ).reset_index(drop=True)
    return duplicates_df


def write_duplicate_sheet(writer, sheet_name, dataframe):
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
    worksheet = writer.book[sheet_name]
    worksheet["A1"] = "Total duplicate rows"
    worksheet["B1"] = int(len(dataframe))
    worksheet["A2"] = "Total duplicate groups"
    worksheet["B2"] = int(dataframe["duplicate_group_id"].nunique()) if not dataframe.empty else 0


def write_workbook(validation_df, lldp_df, duplicates_df, summary_df, xlsx_path):
    rack_total_df = build_rack_total(summary_df)
    pivot_df = build_pivot_table(summary_df)
    interface_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "Interface Down"
    ].copy()
    optics_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "Optics_Issues"
    ].copy()
    lldp_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "LLDP"
    ].copy()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        validation_df.to_excel(writer, sheet_name=VALIDATION_SHEET, index=False)
        lldp_df.to_excel(writer, sheet_name=LLDP_SHEET, index=False)
        write_duplicate_sheet(writer, INTERFACE_DUPLICATES_SHEET, interface_duplicates_df)
        write_duplicate_sheet(writer, OPTICS_DUPLICATES_SHEET, optics_duplicates_df)
        write_duplicate_sheet(writer, LLDP_DUPLICATES_SHEET, lldp_duplicates_df)
        rack_total_df.to_excel(writer, sheet_name=RACK_TOTAL_SHEET, index=False)
        pivot_df.to_excel(writer, sheet_name=PIVOT_SHEET, index=True)

        worksheet = writer.book[RACK_TOTAL_SHEET]
        highlight_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
        comment_col_idx = rack_total_df.columns.get_loc("priority_comment") + 1

        for row in range(2, min(len(rack_total_df), 10) + 2):
            for cell in worksheet[row]:
                cell.fill = highlight_fill
            worksheet.cell(row=row, column=comment_col_idx).comment = Comment(
                "High priority rack",
                "Codex",
            )

        for idx, column in enumerate(rack_total_df.columns, start=1):
            max_data_len = max(rack_total_df[column].astype(str).map(len).max(), len(str(column)))
            worksheet.column_dimensions[get_column_letter(idx)].width = max_data_len + 2

        writer.book[INTERFACE_DUPLICATES_SHEET].sheet_state = "hidden"
        writer.book[OPTICS_DUPLICATES_SHEET].sheet_state = "hidden"
        writer.book[LLDP_DUPLICATES_SHEET].sheet_state = "hidden"


def deduplicate_reports(leaf_report, spine_report_dh1=None, spine_report_dh2=None):
    leaf_df = load_report(leaf_report, "leaf")
    spine_frames = []
    if spine_report_dh1:
        spine_frames.append(load_report(spine_report_dh1, "spine"))
    if spine_report_dh2:
        spine_frames.append(load_report(spine_report_dh2, "spine"))
    spine_df = pd.concat(spine_frames, ignore_index=True, sort=False)
    combined = pd.concat(
        [
            leaf_df,
            spine_df,
        ],
        ignore_index=True,
        sort=False,
    )
    duplicates_df = build_duplicate_matches(combined)

    stats = {
        category: {"input_rows": 0, "kept_rows": 0, "removed_rows": 0}
        for category in REPORT_CATEGORIES
    }

    spine_interface_down = spine_df[spine_df["category"] == "Interface Down"].copy()
    leaf_interface_down_keys = set(
        leaf_df[leaf_df["category"] == "Interface Down"]["_canonical_link_key"]
    )
    interface_down_df = spine_interface_down[
        ~spine_interface_down["_canonical_link_key"].isin(leaf_interface_down_keys)
    ].copy()
    optics_df = spine_df[spine_df["category"] == "Optics_Issues"].copy()
    validation_df = pd.concat([interface_down_df, optics_df], ignore_index=True, sort=False)

    stats["Interface Down"]["input_rows"] = int(len(spine_interface_down))
    stats["Interface Down"]["kept_rows"] = int(len(interface_down_df))
    stats["Interface Down"]["removed_rows"] = int(
        len(spine_interface_down[spine_interface_down["_canonical_link_key"].isin(leaf_interface_down_keys)])
    )

    stats["Optics_Issues"]["input_rows"] = int(len(optics_df))
    stats["Optics_Issues"]["kept_rows"] = int(len(optics_df))
    stats["Optics_Issues"]["removed_rows"] = 0

    lldp_kept_rows = []
    stats["LLDP"]["input_rows"] = int((combined["category"] == "LLDP").sum())
    for _, group in combined[combined["category"] == "LLDP"].groupby("_canonical_link_key", sort=False):
        kept_row = choose_row_to_keep(group)
        lldp_kept_rows.append(kept_row)
        stats["LLDP"]["kept_rows"] += 1
        stats["LLDP"]["removed_rows"] += max(len(group) - 1, 0)

    lldp_df = pd.DataFrame(lldp_kept_rows).copy()

    validation_df = validation_df.drop(
        columns=["_source_side", "_source_report", "_canonical_link_key", "_side_priority"],
        errors="ignore",
    )
    lldp_df = lldp_df.drop(
        columns=["_source_side", "_source_report", "_canonical_link_key", "_side_priority"],
        errors="ignore",
    )

    for column in REPORT_COLUMNS:
        if column not in validation_df.columns:
            validation_df[column] = pd.NA
        if column not in lldp_df.columns:
            lldp_df[column] = pd.NA

    validation_df = validation_df[REPORT_COLUMNS].copy()
    lldp_df = lldp_df[REPORT_COLUMNS].copy()
    summary_df = pd.concat([validation_df, lldp_df], ignore_index=True, sort=False)

    validation_df = validation_df.sort_values(
        by=["device_rack", "hostname", "interface", "category"],
        kind="stable",
    ).reset_index(drop=True)
    lldp_df = lldp_df.sort_values(
        by=["device_rack", "hostname", "interface"],
        kind="stable",
    ).reset_index(drop=True)
    summary_df = summary_df.sort_values(
        by=["device_rack", "hostname", "interface", "category"],
        kind="stable",
    ).reset_index(drop=True)

    return validation_df, lldp_df, duplicates_df, summary_df, stats


def extract_site_prefix(report_path):
    match = re.match(r"([A-Za-z0-9]+)_", Path(report_path).name)
    if match:
        return match.group(1)
    return "REPORT"


def extract_dh_number(report_path):
    if not report_path:
        return "UNKNOWN"
    match = re.search(r"_DH(\d+)_", Path(report_path).name, re.IGNORECASE)
    if match:
        return match.group(1)
    return "UNKNOWN"


def extract_spine_leaf_number(report_path):
    if not report_path:
        return "UNKNOWN"
    name = Path(report_path).name
    fabric_number = re.escape(report_fabric_number())
    for pattern in (
        rf"_bspine_{fabric_number}_(\d+)_bleaf_link_report_",
        rf"_bleaf_{fabric_number}_(\d+)_link_report_",
    ):
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            return match.group(1)
    return "UNKNOWN"


def extract_spine_dh_number(report_path):
    if not report_path:
        return "UNKNOWN"
    name = Path(report_path).name
    fabric_number = re.escape(report_fabric_number())

    bspine_match = re.search(
        rf"_bspine_{fabric_number}_(\d+)_bleaf_link_report_",
        name,
        re.IGNORECASE,
    )
    if bspine_match:
        return bspine_match.group(1)

    return extract_dh_number(report_path)


def extract_report_timestamp(report_path):
    if not report_path:
        return None
    match = re.search(r"_(\d{1,2}[A-Za-z]{3}\d{2})_(\d{4})\.xlsx$", Path(report_path).name)
    if not match:
        return None
    return datetime.strptime("".join(match.groups()), "%d%b%y%H%M")


def build_report_label(spine_report_dh1=None, spine_report_dh2=None):
    if spine_report_dh1 and spine_report_dh2:
        return "Spine_combined_report"
    if spine_report_dh1:
        return "Spine_report_DH1"
    if spine_report_dh2:
        return "Spine_report_DH2"
    return "Spine_report"


def build_console_report_title(leaf_report, spine_report_dh1=None, spine_report_dh2=None):
    fabric_number = report_fabric_number()
    leaf_dh_number = extract_dh_number(leaf_report)

    if spine_report_dh1 and not spine_report_dh2:
        spine_dh_number = extract_spine_dh_number(spine_report_dh1)
        return (
            f"Spine report {fabric_number}.{leaf_dh_number} Bleaf <> "
            f"{fabric_number}.{spine_dh_number} Half spine"
        )

    if spine_report_dh1 and spine_report_dh2:
        return f"Spine report {fabric_number}.{leaf_dh_number} Bleaf <> full spine"

    return f"Spine report {fabric_number}.{leaf_dh_number} Bleaf"


def default_output_path(leaf_report, spine_report_dh1=None, spine_report_dh2=None, unique_error_count=0):
    site_prefix = extract_site_prefix(leaf_report)
    leaf_dh_number = extract_dh_number(leaf_report)
    report_label = build_report_label(spine_report_dh1, spine_report_dh2)
    timestamps = [
        timestamp
        for timestamp in (
            extract_report_timestamp(spine_report_dh1),
            extract_report_timestamp(spine_report_dh2),
        )
        if timestamp is not None
    ]

    if timestamps:
        chosen_timestamp = max(timestamps)
    else:
        chosen_timestamp = datetime.now()

    base_name = (
        f"{site_prefix}_DH{leaf_dh_number}_{unique_error_count}_"
        f"{report_label}_"
        f"{chosen_timestamp.strftime('%d%b%y_%H%M')}"
    )
    return output_dir() / f"{base_name}.xlsx"


def default_output_match_pattern(leaf_report, spine_report_dh1=None, spine_report_dh2=None):
    site_prefix = extract_site_prefix(leaf_report)
    leaf_dh_number = extract_dh_number(leaf_report)
    report_label = build_report_label(spine_report_dh1, spine_report_dh2)
    timestamps = [
        timestamp
        for timestamp in (
            extract_report_timestamp(spine_report_dh1),
            extract_report_timestamp(spine_report_dh2),
        )
        if timestamp is not None
    ]

    if timestamps:
        chosen_timestamp = max(timestamps)
    else:
        return None

    return (
        f"{site_prefix}_DH{leaf_dh_number}_*_{report_label}_"
        f"{chosen_timestamp.strftime('%d%b%y_%H%M')}.xlsx"
    )


def remove_existing_default_outputs(leaf_report, spine_report_dh1=None, spine_report_dh2=None):
    pattern = default_output_match_pattern(leaf_report, spine_report_dh1, spine_report_dh2)
    if pattern is None:
        return []

    removed_paths = []
    for path in output_dir().glob(pattern):
        if not path.is_file():
            continue
        path.unlink()
        removed_paths.append(path)
    return removed_paths


def reports_dir():
    return report_parent_dir() / report_tag_config()["search_subdir"]


def report_search_dirs(search_kind):
    subdir_key = f"{search_kind}_search_subdirs"
    subdirs = report_tag_config().get(subdir_key, (report_tag_config()["search_subdir"],))
    return [report_parent_dir() / subdir for subdir in subdirs]


def format_search_dirs(search_kind):
    return ", ".join(f"{path} recursively" for path in report_search_dirs(search_kind))


def spine_auto_discovery_label():
    return "/".join(f"spine-DH{spine_dh_number}" for spine_dh_number in spine_dh_numbers())


def output_dir():
    return report_repository_dir()


def report_sort_key(report_path):
    timestamp = extract_report_timestamp(report_path)
    if timestamp is None:
        timestamp = datetime.min
    return (timestamp, Path(report_path).stat().st_mtime, Path(report_path).name)


def iter_report_candidates(patterns, search_kind):
    if isinstance(patterns, str):
        patterns = [patterns]
    else:
        patterns = list(patterns)

    seen = set()
    for search_dir in report_search_dirs(search_kind):
        for item in patterns:
            for path in search_dir.rglob(item):
                if path in seen:
                    continue
                seen.add(path)
                yield path


def latest_report(pattern, search_kind="leaf"):
    candidates = []
    candidates.extend(iter_report_candidates(pattern, search_kind))
    candidates = sorted(set(candidates), key=report_sort_key)
    if not candidates:
        return None
    return str(candidates[-1])


def timestamped_candidates(patterns, search_kind):
    candidates = set()
    for path in iter_report_candidates(patterns, search_kind):
        timestamp = extract_report_timestamp(path)
        if timestamp is not None:
            candidates.add((timestamp, path))

    return sorted(
        candidates,
        key=lambda item: report_sort_key(item[1]),
    )


def leaf_report_patterns(leaf_dh_number):
    tag = report_tag()
    return [
        f"{tag}_DH{leaf_dh_number}_*_full_spine_bleaf_link_report_*.xlsx",
        f"{tag}_DH{leaf_dh_number}_*_full_all_bleaf_link_report_*.xlsx",
    ]


def spine_report_patterns(spine_dh_number, leaf_dh_number):
    tag = report_tag()
    fabric_number = report_fabric_number()
    return [
        # Output from abl_bspine_bleaf_validation.py:
        # ABL18_DH1_*_bleaf_18_3_link_report means spine DH1 to leaf DH3;
        # ABL19 uses the same shape with 19 in the bleaf token.
        f"{tag}_DH{spine_dh_number}_*_bleaf_{fabric_number}_{leaf_dh_number}_link_report_*.xlsx",
        # Output from the newer ABL/Bleaf report layout:
        # ABL18_DH3_*_bspine_18_1_bleaf_link_report means leaf DH3 to spine DH1.
        f"{tag}_DH{leaf_dh_number}_*_bspine_{fabric_number}_{spine_dh_number}_bleaf_link_report_*.xlsx",
    ]


def timestamp_candidate_map(patterns, search_kind):
    candidates = {}
    candidate_priorities = {}
    for priority, pattern in enumerate(patterns):
        for path in iter_report_candidates(pattern, search_kind):
            timestamp = extract_report_timestamp(path)
            if timestamp is None:
                continue
            existing = candidates.get(timestamp)
            existing_priority = candidate_priorities.get(timestamp)
            if (
                existing is None
                or priority < existing_priority
                or (
                    priority == existing_priority
                    and report_sort_key(path) > report_sort_key(existing)
                )
            ):
                candidates[timestamp] = path
                candidate_priorities[timestamp] = priority
    return {timestamp: str(path) for timestamp, path in candidates.items()}


def discover_latest_spine_reports(leaf_dh_number="1"):
    candidates_by_dh = {
        spine_dh_number: timestamp_candidate_map(
            spine_report_patterns(spine_dh_number, leaf_dh_number),
            "spine",
        )
        for spine_dh_number in spine_dh_numbers()
    }
    populated_timestamps = [
        set(candidates)
        for candidates in candidates_by_dh.values()
        if candidates
    ]
    if not populated_timestamps:
        return None, None

    common_timestamps = sorted(set.intersection(*populated_timestamps))
    if common_timestamps:
        timestamp = common_timestamps[-1]
        return (
            candidates_by_dh.get("1", {}).get(timestamp),
            candidates_by_dh.get("2", {}).get(timestamp),
        )

    return None, None


def discover_latest_spine_pair(leaf_dh_number="1"):
    return discover_latest_spine_reports(leaf_dh_number)


def discover_latest_job(leaf_dh_number):
    leaf_candidates = timestamped_candidates(leaf_report_patterns(leaf_dh_number), "leaf")
    latest_leaf_path = str(leaf_candidates[-1][1]) if leaf_candidates else None
    candidates_by_dh = {
        spine_dh_number: timestamp_candidate_map(
            spine_report_patterns(spine_dh_number, leaf_dh_number),
            "spine",
        )
        for spine_dh_number in spine_dh_numbers()
    }

    best_match = None
    populated_timestamps = [
        set(candidates)
        for candidates in candidates_by_dh.values()
        if candidates
    ]
    if not populated_timestamps:
        return latest_leaf_path, None, None

    for spine_timestamp in sorted(set.intersection(*populated_timestamps)):
        spine_paths = [
            Path(candidates_by_dh[spine_dh_number][spine_timestamp])
            for spine_dh_number in spine_dh_numbers()
        ]
        for leaf_timestamp, leaf_path in leaf_candidates:
            if abs((leaf_timestamp - spine_timestamp).total_seconds()) > DISCOVERY_WINDOW_SECONDS:
                continue

            match_key = (
                max(leaf_timestamp, spine_timestamp),
                spine_timestamp,
                report_sort_key(leaf_path),
                *(report_sort_key(spine_path) for spine_path in spine_paths),
            )
            if best_match is None or match_key > best_match[0]:
                best_match = (match_key, leaf_path, spine_paths)

    if best_match is not None:
        _, leaf_path, spine_paths = best_match
        spine_path_by_dh = {
            spine_dh_number: str(spine_path)
            for spine_dh_number, spine_path in zip(spine_dh_numbers(), spine_paths)
        }
        return str(leaf_path), spine_path_by_dh.get("1"), spine_path_by_dh.get("2")

    return latest_leaf_path, None, None


def is_expected_spine_report(report_path, expected_spine_dh_number, leaf_dh_number):
    report_dh_number = extract_dh_number(report_path)
    spine_leaf_number = extract_spine_leaf_number(report_path)
    name = Path(report_path).name.lower()

    if "_bspine_" in name:
        return (
            report_dh_number == leaf_dh_number
            and re.search(
                rf"_bspine_{re.escape(report_fabric_number())}_{re.escape(str(expected_spine_dh_number))}_bleaf_link_report_",
                name,
                re.IGNORECASE,
            )
            is not None
        )

    return (
        report_dh_number == expected_spine_dh_number
        and spine_leaf_number == leaf_dh_number
        and "_bleaf_" in name
        and "_bspine_" not in name
    )


def validate_report_inputs(parser, leaf_report, spine_report_dh1=None, spine_report_dh2=None):
    if not leaf_report:
        parser.error(
            "could not auto-discover a leaf-side report matching "
            "*_DH<N>_*_full_spine_bleaf_link_report_*.xlsx or "
            f"*_DH<N>_*_full_all_bleaf_link_report_*.xlsx under {format_search_dirs('leaf')}"
        )
    leaf_dh_number = extract_dh_number(leaf_report)
    if not spine_report_dh1 and not spine_report_dh2:
        expected_spines = ", ".join(f"--spine-DH{number}" for number in spine_dh_numbers())
        parser.error(
            f"at least one configured spine report is required ({expected_spines}); "
            f"searched {format_search_dirs('spine')}"
        )
    if spine_report_dh1 and not is_expected_spine_report(spine_report_dh1, "1", leaf_dh_number):
        fabric_number = report_fabric_number()
        parser.error(
            "--spine-DH1 expects a spine-to-leaf report matching "
            f"*_DH1_*_bleaf_{fabric_number}_{leaf_dh_number}_link_report_*.xlsx or "
            f"*_DH{leaf_dh_number}_*_bspine_{fabric_number}_1_bleaf_link_report_*.xlsx, but got: "
            f"{Path(spine_report_dh1).name}"
        )
    if spine_report_dh2 and not is_expected_spine_report(spine_report_dh2, "2", leaf_dh_number):
        fabric_number = report_fabric_number()
        parser.error(
            "--spine-DH2 expects a spine-to-leaf report matching "
            f"*_DH2_*_bleaf_{fabric_number}_{leaf_dh_number}_link_report_*.xlsx or "
            f"*_DH{leaf_dh_number}_*_bspine_{fabric_number}_2_bleaf_link_report_*.xlsx, but got: "
            f"{Path(spine_report_dh2).name}"
        )


def has_report_job_argument(raw_args):
    job_options = {
        "--leaf-DH",
        "--leaf-report",
        "--spine-DH1",
        "--spine-report-dh1",
        "--spine-DH2",
        "--spine-report-dh2",
        "--output-csv",
        "--output-xlsx",
        "--both",
        "--both-latest",
        "--datahall",
        "--dh",
    }
    for raw_arg in raw_args:
        option = raw_arg.split("=", 1)[0]
        if option in job_options:
            return True
    return False


def parse_args():
    parser = ColorArgumentParser(
        description="De-duplicate leaf and spine link validation reports, preferring the leaf-side row."
    )
    parser.add_argument(
        "--tag",
        "--report-tag",
        dest="report_tag",
        type=normalize_report_tag,
        required=True,
        help=(
            "Report tag for auto-discovery and default output location "
            f"({', '.join(sorted(REPORT_TAGS))})."
        ),
    )
    parser.add_argument(
        "--leaf-DH",
        "--leaf-report",
        dest="leaf_report",
        help=(
            "Leaf-side full spine/bleaf report. If omitted, the latest default "
            "leaf-DH full_spine_bleaf report for the selected tag is used."
        ),
    )
    parser.add_argument(
        "--spine-DH1",
        "--spine-report-dh1",
        dest="spine_report_dh1",
        help="Spine-side DH1 report.",
    )
    parser.add_argument(
        "--spine-DH2",
        "--spine-report-dh2",
        dest="spine_report_dh2",
        help="Spine-side DH2 report.",
    )
    parser.add_argument(
        "--output-csv",
        help="Output CSV path for the de-duplicated non-LLDP Validation Errors rows.",
    )
    parser.add_argument("--output-xlsx", help="Output XLSX path for the de-duplicated workbook.")
    parser.add_argument(
        "--both",
        "--both-latest",
        dest="both_latest",
        action="store_true",
        help=(
            "Auto-discover and run the latest dedupe jobs for the selected tag's "
            "default datahalls. "
            "This uses the latest DH-level full_spine_bleaf report plus the "
            "latest configured DH-level spine-to-leaf reports."
        ),
    )
    parser.add_argument(
        "--datahall",
        "--dh",
        dest="leaf_dh_number",
        type=normalize_dh_number,
        help=(
            "Auto-discover and run the latest reports for one leaf datahall, "
            "for example --datahall 2 or --datahall DH3."
        ),
    )
    raw_args = sys.argv[1:]
    args = parser.parse_args()
    set_report_tag(args.report_tag)

    if not raw_args or not has_report_job_argument(raw_args):
        args.both_latest = True

    if args.both_latest:
        if args.leaf_dh_number:
            parser.error("--datahall cannot be combined with --both/default auto-run")
        if any(
            [
                args.leaf_report,
                args.spine_report_dh1,
                args.spine_report_dh2,
                args.output_csv,
                args.output_xlsx,
            ]
        ):
            parser.error(
                "--both/default auto-run cannot be combined with explicit report or output paths"
            )
        args.auto_discovered = ["both"]
        return args

    selected_leaf_dh_number = args.leaf_dh_number or default_leaf_dh_number()
    explicit_report_paths = any(
        [
            args.leaf_report,
            args.spine_report_dh1,
            args.spine_report_dh2,
        ]
    )

    auto_discovered = []
    if args.leaf_dh_number and not explicit_report_paths:
        args.leaf_report, args.spine_report_dh1, args.spine_report_dh2 = discover_latest_job(
            selected_leaf_dh_number
        )
        if args.leaf_report:
            auto_discovered.append(f"leaf-DH{selected_leaf_dh_number}")
        if args.spine_report_dh1 or args.spine_report_dh2:
            auto_discovered.append(spine_auto_discovery_label())
        args.auto_discovered = auto_discovered
        validate_report_inputs(
            parser,
            args.leaf_report,
            args.spine_report_dh1,
            args.spine_report_dh2,
        )
        return args

    if not args.leaf_report:
        args.leaf_report = latest_report(leaf_report_patterns(selected_leaf_dh_number), "leaf")
        if args.leaf_report:
            auto_discovered.append(f"leaf-DH{selected_leaf_dh_number}")

    if not args.spine_report_dh1 and not args.spine_report_dh2:
        args.spine_report_dh1, args.spine_report_dh2 = discover_latest_spine_pair(
            selected_leaf_dh_number
        )
        if args.spine_report_dh1 or args.spine_report_dh2:
            auto_discovered.append(spine_auto_discovery_label())
    elif "1" in spine_dh_numbers() and not args.spine_report_dh1:
        leaf_dh_number = extract_dh_number(args.leaf_report) if args.leaf_report else selected_leaf_dh_number
        args.spine_report_dh1 = latest_report(spine_report_patterns("1", leaf_dh_number), "spine")
        if args.spine_report_dh1:
            auto_discovered.append("spine-DH1")
    elif "2" in spine_dh_numbers() and not args.spine_report_dh2:
        leaf_dh_number = extract_dh_number(args.leaf_report) if args.leaf_report else selected_leaf_dh_number
        args.spine_report_dh2 = latest_report(spine_report_patterns("2", leaf_dh_number), "spine")
        if args.spine_report_dh2:
            auto_discovered.append("spine-DH2")

    args.auto_discovered = auto_discovered

    validate_report_inputs(
        parser,
        args.leaf_report,
        args.spine_report_dh1,
        args.spine_report_dh2,
    )
    return args


def run_dedup_job(
    leaf_report,
    spine_report_dh1=None,
    spine_report_dh2=None,
    output_csv=None,
    output_xlsx=None,
    auto_discovered=None,
):
    validation_df, lldp_df, duplicates_df, summary_df, stats = deduplicate_reports(
        leaf_report=leaf_report,
        spine_report_dh1=spine_report_dh1,
        spine_report_dh2=spine_report_dh2,
    )

    xlsx_path = default_output_path(
        leaf_report,
        spine_report_dh1,
        spine_report_dh2,
        unique_error_count=len(summary_df),
    )

    if output_xlsx:
        xlsx_path = Path(output_xlsx).resolve()

    interface_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "Interface Down"
    ]
    optics_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "Optics_Issues"
    ]
    lldp_duplicates_df = duplicates_df[
        duplicates_df["duplicate_error_type"] == "LLDP"
    ]
    interface_duplicate_groups = (
        interface_duplicates_df["duplicate_group_id"].nunique() if not interface_duplicates_df.empty else 0
    )
    optics_duplicate_groups = (
        optics_duplicates_df["duplicate_group_id"].nunique() if not optics_duplicates_df.empty else 0
    )
    lldp_duplicate_groups = (
        lldp_duplicates_df["duplicate_group_id"].nunique() if not lldp_duplicates_df.empty else 0
    )
    total_duplicate_groups = interface_duplicate_groups + optics_duplicate_groups + lldp_duplicate_groups
    total_duplicate_links = len(interface_duplicates_df) + len(optics_duplicates_df) + len(lldp_duplicates_df)
    total_unique_errors = len(summary_df)

    if auto_discovered:
        print(f"Auto-discovery mode   : {', '.join(auto_discovered)}")

    print(f"Report tag            : {report_tag()}")
    print(f"Repository location   : {report_repository_dir().resolve()}")
    print(f"Leaf-side report      : {Path(leaf_report).resolve()}")
    if spine_report_dh1:
        print(f"Spine-side DH1 report : {Path(spine_report_dh1).resolve()}")
    if spine_report_dh2:
        print(f"Spine-side DH2 report : {Path(spine_report_dh2).resolve()}")

    if total_unique_errors:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        if output_csv:
            csv_path = Path(output_csv).resolve()
            csv_path.parent.mkdir(parents=True, exist_ok=True)
            validation_df.to_csv(csv_path, index=False)

        replaced_paths = []
        if not output_xlsx:
            replaced_paths = remove_existing_default_outputs(
                leaf_report,
                spine_report_dh1,
                spine_report_dh2,
            )
        write_workbook(validation_df, lldp_df, duplicates_df, summary_df, xlsx_path)
        print(f"Output XLSX           : {xlsx_path}")
        if replaced_paths:
            print(f"Replaced XLSX         : {', '.join(path.name for path in replaced_paths)}")
        if output_csv:
            print(f"Output CSV            : {csv_path}")
    else:
        print("Output XLSX           : not created (0 unique errors)")
    print()

    print(build_console_report_title(leaf_report, spine_report_dh1, spine_report_dh2))
    print("Summary after removing duplicates")
    print(f"  Total Unique Errors: {total_unique_errors}")
    print(f"  Interface Down: {stats['Interface Down']['kept_rows']}")
    print(f"  Optics Issues: {stats['Optics_Issues']['kept_rows']}")
    print(f"  LLDP Mismatch: {stats['LLDP']['kept_rows']}")
    print(
        f"  Duplicates Removed: {total_duplicate_groups}"
        f"({interface_duplicate_groups} int_down + {optics_duplicate_groups} optic + {lldp_duplicate_groups} lldp) "
        f"groups ({total_duplicate_links} links)"
    )
    print()
    print("Note:")
    print("  Interface/Optics are unique spine-side errors.")
    print("  LLDP duplicates are matched by device/port pair. If found on both sides, leaf-side is kept.")
    print("  LLDP rows found on only one side are kept as-is.")


def main():
    args = parse_args()

    if args.both_latest:
        parser = ColorArgumentParser()
        for leaf_dh_number in both_latest_leaf_dh_numbers():
            leaf_report, spine_report_dh1, spine_report_dh2 = discover_latest_job(
                leaf_dh_number
            )
            missing_reports = []
            if not leaf_report:
                missing_reports.append("leaf-DH")
            if "1" in spine_dh_numbers() and not spine_report_dh1:
                missing_reports.append("spine-DH1")
            if "2" in spine_dh_numbers() and not spine_report_dh2:
                missing_reports.append("spine-DH2")
            if missing_reports:
                parser.error(
                    "could not auto-discover the latest DH-level report(s) for "
                    f"{report_tag()} DH{leaf_dh_number}: {', '.join(missing_reports)}"
                )
            validate_report_inputs(
                parser,
                leaf_report,
                spine_report_dh1,
                spine_report_dh2,
            )
            print(f"{report_tag()} DH{leaf_dh_number} latest dedupe")
            run_dedup_job(
                leaf_report=leaf_report,
                spine_report_dh1=spine_report_dh1,
                spine_report_dh2=spine_report_dh2,
                auto_discovered=[f"leaf-DH{leaf_dh_number}", spine_auto_discovery_label()],
            )
            if leaf_dh_number != both_latest_leaf_dh_numbers()[-1]:
                print()
        return

    run_dedup_job(
        leaf_report=args.leaf_report,
        spine_report_dh1=args.spine_report_dh1,
        spine_report_dh2=args.spine_report_dh2,
        output_csv=args.output_csv,
        output_xlsx=args.output_xlsx,
        auto_discovered=args.auto_discovered,
    )


if __name__ == "__main__":
    main()
